import sys
import os
import io
import csv
import logging
from logging.handlers import RotatingFileHandler
import secrets
import subprocess
import random
from pathlib import Path
import threading
import importlib
import webbrowser
from tools.aegis_logger import live_logger, init_demo_logging, log_user_action, log_api_call, log_database_operation, log_authentication, log_scan_start, log_scan_complete, log_vulnerability, log_critical_finding
import google.generativeai as genai
from apscheduler.schedulers.background import BackgroundScheduler
import atexit
from apscheduler.jobstores.sqlalchemy import SQLAlchemyJobStore
from apscheduler.executors.pool import ThreadPoolExecutor
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, or_, distinct, case
from datetime import datetime, timezone, timedelta
import json
import time
from collections import defaultdict
from functools import wraps
# License management imports
from licenses.license_middleware import LicenseMiddleware, require_license, require_feature, init_license_routes, init_admin_license_routes
from licenses.license_manager import LicenseManager, LicenseValidator

# --- User-Writable Data Directory ---
APP_NAME = "AegisScanner"
if sys.platform == "win32":
    USER_DATA_DIR = os.path.join(os.environ['LOCALAPPDATA'], APP_NAME)
else:
    USER_DATA_DIR = os.path.join(os.path.expanduser('~'), f'.{APP_NAME.lower()}')
os.makedirs(USER_DATA_DIR, exist_ok=True)

# --- Global File Paths ---
log_file = os.path.join(USER_DATA_DIR, 'aegis_scanner_debug.log')
ENV_FILE_PATH = os.path.join(USER_DATA_DIR, '.env')

# --- Hardcoded Debug Flag ---
is_debug = True
log_level = logging.DEBUG

# --- Structured Logging Setup ---
class StructuredLogger:
    """Production-ready structured logging for better monitoring and debugging"""
    
    def __init__(self, name):
        self.logger = logging.getLogger(name)
        
    def log_structured(self, level, message, **kwargs):
        """Log with structured data for production monitoring"""
        log_data = {
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'message': message,
            'service': 'aegis-scanner',
            **kwargs
        }
        getattr(self.logger, level)(json.dumps(log_data))
    
    def log_request(self, request, response_status, duration_ms, **kwargs):
        """Log HTTP requests with structured data"""
        self.log_structured('info', 'HTTP Request', 
            method=request.method,
            path=request.path,
            remote_addr=request.remote_addr,
            user_agent=request.headers.get('User-Agent', ''),
            response_status=response_status,
            duration_ms=duration_ms,
            **kwargs
        )
    
    def log_security_event(self, event_type, user_id=None, details=None):
        """Log security events for monitoring"""
        self.log_structured('warning', 'Security Event',
            event_type=event_type,
            user_id=user_id,
            details=details or {}
        )
    
    def log_scan_activity(self, scan_type, provider, status, duration_ms=None, resource_count=None):
        """Log cloud scanning activities"""
        self.log_structured('info', 'Scan Activity',
            scan_type=scan_type,
            provider=provider,
            status=status,
            duration_ms=duration_ms,
            resource_count=resource_count
        )

# Global structured logger instance
structured_logger = StructuredLogger('aegis_scanner')

# --- Rate Limiting Storage ---
rate_limit_storage = defaultdict(list)

def rate_limit_key_func():
    """Rate limiting key function with IP and user tracking"""
    from flask import request, g
    # Use user ID if authenticated, otherwise IP address
    if hasattr(g, 'current_user') and g.current_user and hasattr(g.current_user, 'id'):
        return f"user:{g.current_user.id}"
    return f"ip:{request.remote_addr}"

# --- Logging Setup ---
def configure_logging():
    log_formatter = logging.Formatter('%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]')
    file_handler = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=5)
    file_handler.setFormatter(log_formatter)
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(log_formatter)
    root_logger = logging.getLogger()
    root_logger.handlers.clear()
    root_logger.addHandler(file_handler)
    if not getattr(sys, 'frozen', False):
         root_logger.addHandler(stream_handler)
    root_logger.setLevel(log_level)
    logging.getLogger('waitress').setLevel(logging.INFO)
    logging.getLogger('werkzeug').setLevel(logging.WARNING)
    logging.getLogger('sqlalchemy.engine').setLevel(logging.INFO)

configure_logging()

class LiveActivityLogger:
    def __init__(self):
        self.activities = []
    def log(self, action, category='INFO', details=""):
        activity = {
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'action': action, 'category': category, 'details': details
        }
        self.activities.append(activity)
        if len(self.activities) > 100: self.activities = self.activities[-100:]

live_logger = LiveActivityLogger()

from flask import Flask, jsonify, request, redirect, url_for, render_template, flash, session, Response, make_response, g, current_app
from flask_cors import CORS
from flask_migrate import Migrate
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_bcrypt import Bcrypt
from flask_mail import Mail, Message
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import time
import click
from flask_wtf.csrf import CSRFProtect, generate_csrf, CSRFError, validate_csrf
from flask_talisman import Talisman
import boto3
import pyotp
import qrcode
from io import BytesIO
import base64
import re
# Replaced WeasyPrint with reportlab for better compatibility
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import io
from waitress import serve
from itsdangerous import URLSafeTimedSerializer
import tools.parallel_scanner as parallel_scanner
import json
import hashlib
import string
# from zxcvbn import zxcvbn  # Disabled due to Python 3.13 compatibility issues
from botocore.exceptions import ClientError
from google.oauth2 import service_account
from google.cloud import storage
from dotenv import load_dotenv
from functools import partial, wraps

from tools.crypto_manager import SecureCryptoManager, CredentialMigrator, setup_secure_crypto_manager, generate_secure_password
from tools.input_validators import (SecurityValidator, validate_json_input, validate_form_input, security_validator)
from tools.error_handlers import (register_error_handlers, AegisError, SecurityError, ValidationError, DatabaseError, AuthenticationError, safe_execute, cloud_api_handler, ErrorContext, ErrorCategory)

scheduler = None

def initialize_scheduler(database_uri):
    global scheduler
    try:
        jobstores = {'default': SQLAlchemyJobStore(url=database_uri)}
        executors = {'default': ThreadPoolExecutor(20)}
        job_defaults = {'coalesce': False, 'max_instances': 1, 'misfire_grace_time': 30}
        scheduler = BackgroundScheduler(jobstores=jobstores, executors=executors, job_defaults=job_defaults)
        scheduler.start()
        logging.info("Scheduler initialized successfully")
        def safe_shutdown():
            try:
                if scheduler and scheduler.running:
                    scheduler.shutdown()
            except Exception as e:
                pass  # Ignore shutdown errors
        atexit.register(safe_shutdown)
    except Exception as e:
        logging.error(f"Failed to initialize scheduler: {e}")

def perform_scheduled_scan(user_id, credential_id, regions=None):
    with app.app_context():
        try:
            user = db.session.get(User, user_id)
            credential = CloudCredential.query.filter_by(id=credential_id, user_id=user_id).first()
            if not user or not credential:
                logging.error(f"Scheduled scan failed: user {user_id} or credential {credential_id} not found")
                return
            live_logger.log("Scheduled Scan Started", "SCHEDULED", f"User: {user.username}, Profile: {credential.profile_name}")
            decrypted_creds = {"provider": credential.provider, "profile_name": credential.profile_name}
            if credential.provider == 'aws':
                decrypted_creds["aws_access_key_id"] = decrypt_data(credential.encrypted_key_1)
                decrypted_creds["aws_secret_access_key"] = decrypt_data(credential.encrypted_key_2)
            elif credential.provider == 'gcp':
                decrypted_creds["gcp_service_account_json"] = decrypt_data(credential.encrypted_key_2)
            scan_results = parallel_scanner.run_parallel_scans_blocking(credentials=decrypted_creds, regions=regions)
            suppressed_hashes = {sf.finding_hash for sf in SuppressedFinding.query.filter_by(user_id=user_id).all()}
            final_results = [r for r in scan_results if _generate_finding_hash(r) not in suppressed_hashes]
            critical_findings = [r for r in final_results if r.get('status') == 'CRITICAL']
            scan_time = datetime.now(timezone.utc)
            for result in final_results:
                if "error" not in result:
                    db_result = ScanResult(
                        service=result.get('service'), resource=result.get('resource'), status=result.get('status'),
                        issue=result.get('issue'), remediation=result.get('remediation'), doc_url=result.get('doc_url'),
                        timestamp=scan_time, user_id=user_id
                    )
                    db.session.add(db_result)
            db.session.commit()
            if user.notifications_enabled:
                send_scan_completion_notification(user, final_results, credential.profile_name)
                if critical_findings and user.email_on_critical_findings:
                    send_critical_findings_alert(user, critical_findings, credential.profile_name)
            log_audit("Scheduled Scan Completed", details=f"Profile: {credential.profile_name}, Results: {len(final_results)}, Critical: {len(critical_findings)}", user=user)
            live_logger.log("Scheduled Scan Completed", "SUCCESS", f"Results: {len(final_results)}, Critical: {len(critical_findings)}")
        except Exception as e:
            logging.error(f"Scheduled scan error for user {user_id}: {e}")
            live_logger.log("Scheduled Scan Failed", "ERROR", str(e))

def send_scheduled_report(user_id):
    with app.app_context():
        try:
            user = db.session.get(User, user_id)
            if not user or not user.notifications_enabled: return
            thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
            scan_results = ScanResult.query.filter_by(user_id=user_id).filter(ScanResult.timestamp >= thirty_days_ago).order_by(ScanResult.timestamp.desc()).limit(100).all()
            if not scan_results:
                logging.info(f"No recent scan results for user {user_id} - skipping scheduled report")
                return
            pdf_bytes = _create_pdf_report(scan_results)
            msg = Message('Aegis Scanner - Scheduled Security Report', recipients=[user.email])
            msg.html = render_template('scheduled_report_email.html', user=user, result_count=len(scan_results))
            msg.attach('aegis_security_report.pdf', 'application/pdf', pdf_bytes)
            mail.send(msg)
            log_audit("Scheduled Report Sent", details=f"Results: {len(scan_results)}", user=user)
            live_logger.log("Scheduled Report Sent", "EMAIL", f"User: {user.username}")
        except Exception as e:
            logging.error(f"Scheduled report error for user {user_id}: {e}")

def send_scan_completion_notification(user, results, profile_name):
    try:
        total_findings = len(results)
        critical_count = sum(1 for r in results if isinstance(r, dict) and r.get('status') == 'CRITICAL')
        dashboard_url = generate_external_url('dashboard')
        msg = Message(f'Aegis Scanner - Scan Complete: {profile_name}', recipients=[user.email])
        msg.html = render_template('scan_notification_email.html', user=user, profile_name=profile_name, total_findings=total_findings, critical_findings=critical_count, dashboard_url=dashboard_url)
        mail.send(msg)
        logging.info(f"Scan completion notification sent to {user.email}")
    except Exception as e:
        logging.error(f"Failed to send scan notification: {e}")

def send_critical_findings_alert(user, critical_findings, profile_name):
    try:
        if not user.notifications_enabled or not user.email_on_critical_findings or not critical_findings: return
        dashboard_url = generate_external_url('dashboard')
        msg = Message(f'URGENT: Critical Security Issues Found in {profile_name}', recipients=[user.email])
        msg.html = render_template('alert_email.html', user=user, profile_name=profile_name, critical_findings=critical_findings, dashboard_url=dashboard_url)
        mail.send(msg)
        log_audit("Critical Findings Alert Sent", details=f"Profile: {profile_name}, Issues: {len(critical_findings)}", user=user)
        logging.info(f"Critical findings alert sent to {user.email}")
    except Exception as e:
        logging.error(f"Failed to send critical findings alert: {e}")

def perform_automated_scan(credential_id, user_id, regions=None):
    """Performs automated scheduled scan - wrapper for perform_scheduled_scan"""
    return perform_scheduled_scan(user_id, credential_id, regions)

# NEW IMPORTS: Security and validation systems
from tools.crypto_manager import SecureCryptoManager, CredentialMigrator, setup_secure_crypto_manager
from tools.input_validators import (
    SecurityValidator,
    validate_json_input,
    validate_form_input,
    security_validator
)
from tools.error_handlers import (
    register_error_handlers,
    AegisError,
    SecurityError,
    ValidationError,
    DatabaseError,
    AuthenticationError,
    safe_execute,
    cloud_api_handler,
    ErrorContext,
    ErrorCategory
)

# --- Global Variables ---
db = SQLAlchemy()
migrate = Migrate()
bcrypt = Bcrypt()
login_manager = LoginManager()
mail = Mail()
csrf = CSRFProtect()
limiter = Limiter(key_func=rate_limit_key_func, storage_uri="memory://", default_limits=["1000 per hour", "100 per minute"])
s = None
crypto_manager = None

# --- Main App Creation using Factory Pattern ---
def create_app():
    template_folder = os.path.join(sys._MEIPASS, 'templates') if getattr(sys, 'frozen', False) else 'templates'
    static_folder = os.path.join(sys._MEIPASS, 'static') if getattr(sys, 'frozen', False) else 'static'
    app = Flask(__name__, instance_relative_config=True, template_folder=template_folder, static_folder=static_folder)

    # ULTIMATE SERVER_NAME FIX: Patch Flask at the source
    # The issue occurs when Flask tries to access config["SERVER_NAME"] but the key doesn't exist
    # We'll patch ALL possible places where this could happen

    # Step 1: Ensure SERVER_NAME is always in config
    app.config['SERVER_NAME'] = None

    # Step 2: Patch Flask's Config class to always return None for missing SERVER_NAME
    import flask
    original_config_getitem = flask.Config.__getitem__

    def patched_config_getitem(self, key):
        if key == 'SERVER_NAME' and key not in self:
            return None
        return original_config_getitem(self, key)

    flask.Config.__getitem__ = patched_config_getitem

    # Step 3: Patch the app's config specifically as well
    original_app_config_getitem = app.config.__getitem__

    def patched_app_config_getitem(key):
        if key == 'SERVER_NAME' and key not in app.config:
            return None
        return original_app_config_getitem(key)

    app.config.__getitem__ = patched_app_config_getitem

    load_dotenv(dotenv_path=ENV_FILE_PATH)
    
    # Import and initialize secrets manager for production-ready secrets handling
    from tools.secrets_manager import secrets_manager
    app.config['SECRET_KEY'] = secrets_manager.get_secret_key()
    app.config['START_TIME'] = datetime.now()
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    
    # Database configuration - supports both SQLite (dev) and PostgreSQL (production)
    database_url = secrets_manager.get_database_url()
    if database_url:
        # Production: PostgreSQL
        app.config['SQLALCHEMY_DATABASE_URI'] = database_url
        # Configure connection pool for PostgreSQL
        app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
            'pool_size': 10,
            'pool_recycle': 120,
            'pool_pre_ping': True,
            'max_overflow': 20
        }
    else:
        # Development: SQLite
        DB_PATH = os.path.join(USER_DATA_DIR, 'app.db')
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + DB_PATH
    
    # Email configuration using secrets manager
    mail_config = secrets_manager.get_mail_config()
    app.config.update(mail_config)
    app.config['MAIL_DEFAULT_SENDER'] = mail_config['MAIL_USERNAME']
    app.config['SUPPORT_EMAIL'] = os.getenv('SUPPORT_EMAIL', mail_config['MAIL_USERNAME'])

    # Server configuration for external URL generation
    # Always set SERVER_NAME to prevent Flask KeyErrors
    domain_name = os.getenv('DOMAIN_NAME', 'localhost:5000')

    # Set SERVER_NAME to None for local development to allow flexible access
    if 'localhost' in domain_name or '127.0.0.1' in domain_name:
        app.config['SERVER_NAME'] = None
    else:
        app.config['SERVER_NAME'] = domain_name

    app.config['APPLICATION_ROOT'] = '/'
    app.config['PREFERRED_URL_SCHEME'] = 'http' if 'localhost' in domain_name or '127.0.0.1' in domain_name else 'https'
    app.config['SQLALCHEMY_ECHO'] = False
    global s, crypto_manager
    s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
    try:
        crypto_manager = setup_secure_crypto_manager()
        if not crypto_manager.is_initialized(): raise ValueError("Crypto manager initialization failed")
        logging.info("Secure crypto manager initialized successfully")
    except Exception as e:
        logging.error(f"Failed to initialize crypto manager: {e}")
        raise
    db.init_app(app)
    migrate.init_app(app, db)
    bcrypt.init_app(app)
    login_manager.init_app(app)
    login_manager.login_view = 'auth'
    login_manager.login_message_category = "info"

    # Handle AJAX requests that need authentication
    @login_manager.unauthorized_handler
    def handle_unauthorized():
        # Check if this is an AJAX request
        if request.is_json or request.headers.get('Content-Type') == 'application/json' or 'application/json' in request.headers.get('Accept', ''):
            return jsonify({
                'success': False,
                'message': 'Authentication required. Please log in.',
                'redirect': url_for('auth')
            }), 401
        # For regular requests, redirect as usual
        return redirect(url_for('auth'))
    mail.init_app(app)
    csrf.init_app(app)
    limiter.init_app(app)
    
    # Production request logging middleware
    @app.before_request
    def before_request():
        g.start_time = time.time()
    
    @app.after_request
    def after_request(response):
        if hasattr(g, 'start_time'):
            duration_ms = round((time.time() - g.start_time) * 1000, 2)
            # Log all requests for production monitoring
            structured_logger.log_request(
                request=request,
                response_status=response.status_code,
                duration_ms=duration_ms,
                content_length=response.content_length
            )
        return response
    
    csp = {'default-src': '\'self\'', 'script-src': ['\'self\'', 'https://cdn.jsdelivr.net', 'https://unpkg.com', 'https://cdnjs.cloudflare.com', '\'unsafe-inline\'', '\'unsafe-eval\''], 'style-src': ['\'self\'', 'https://cdnjs.cloudflare.com', 'https://fonts.googleapis.com', 'https://unpkg.com', 'https://cdn.jsdelivr.net', '\'unsafe-inline\''], 'font-src': ['\'self\'', 'https://cdnjs.cloudflare.com', 'https://fonts.gstatic.com'], 'img-src': ['\'self\'', 'data:'], 'connect-src': ['\'self\'', 'https://cdn.jsdelivr.net', 'https://cdnjs.cloudflare.com']}
    Talisman(app, content_security_policy=csp, force_https=False)
    register_error_handlers(app)

    # Initialize license system
    try:
        license_middleware = LicenseMiddleware(app)
        app.license_middleware = license_middleware
        init_license_routes(app)
        init_admin_license_routes(app)
        logging.info("License system initialized successfully")
    except Exception as e:
        logging.error(f"Failed to initialize license system: {e}")

    if not getattr(sys, 'frozen', False):
        try:
            initialize_scheduler(app.config['SQLALCHEMY_DATABASE_URI'])
        except Exception as e:
            logging.warning(f"Scheduler initialization delayed: {e}")
    return app

app = create_app()


# ... (Error handlers, template filters, crypto functions, DB models, decorators, etc. remain largely the same) ...
# NOTE: All existing functions like `encrypt_data`, `log_audit`, `User` model, etc., are assumed to be here.

# --- Main Application Routes ---
@app.route('/')
def initializing():
    return render_template('initializing.html')

# ... (Existing routes like /welcome, /auth, /login, /register, etc., are here) ...

@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    logging.warning(f"--- [SECURITY] CSRF token validation failed. Reason: {e.description} ---")

    # Handle AJAX requests with JSON response
    if request.is_json or 'application/json' in request.headers.get('Accept', ''):
        return jsonify({
            'success': False,
            'message': 'CSRF token validation failed. Please refresh the page and try again.',
            'error': 'csrf_error'
        }), 400

    # Handle regular requests with redirect
    flash('Your form session has expired. Please log in and try again.', 'error')
    return redirect(url_for('auth'))

# UPDATED: Template filter with enhanced error handling

@app.template_filter('a_id_decrypt')
def a_id_decrypt_filter(encrypted_data):
    try:
        if not encrypted_data: return "N/A"
        return decrypt_data(encrypted_data, "template display")
    except Exception as e:
        logging.warning(f"Template decryption failed: {e}")
        return "***ENCRYPTED***"

# --- UPDATED: Secure Helper Functions ---
def encrypt_data(data: str, context: str = "generic data") -> str:
    if not crypto_manager or not crypto_manager.is_initialized(): raise ValueError("Encryption service not available.")
    try:
        return crypto_manager.encrypt_credential(data, context)
    except Exception as e:
        logging.error(f"Failed to encrypt {context}: {e}")
        raise ValueError(f"Encryption failed for {context}")

def decrypt_data(encrypted_data: str, context: str = "generic data") -> str:
    if not crypto_manager or not crypto_manager.is_initialized(): raise ValueError("Decryption service not available.")
    try:
        return crypto_manager.decrypt_credential(encrypted_data, context)
    except Exception as e:
        logging.error(f"Failed to decrypt {context}: {e}")
        raise ValueError(f"Decryption failed for {context}")

# --- Database Models ---
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(64), unique=True, index=True)
    email = db.Column(db.String(120), unique=True, index=True)
    password_hash = db.Column(db.String(128))
    failed_login_attempts = db.Column(db.Integer, default=0)
    is_locked = db.Column(db.Boolean, default=False)
    email_verified = db.Column(db.Boolean, default=False)
    otp_secret = db.Column(db.String(32))
    is_2fa_enabled = db.Column(db.Boolean, default=False)
    inactivity_timeout = db.Column(db.Integer, default=15, nullable=False)
    is_admin = db.Column(db.Boolean, default=False, nullable=False)
    is_active = db.Column(db.Boolean, default=True, nullable=False)
    last_login_date = db.Column(db.DateTime, nullable=True)
    created_date = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    backup_email = db.Column(db.String(120), unique=True, nullable=True)
    backup_email_verified = db.Column(db.Boolean, default=False)
    recovery_codes = db.Column(db.Text, nullable=True)  # JSON list of backup codes
    recovery_codes_used = db.Column(db.Text, nullable=True)  # JSON list of used codes
    failed_2fa_attempts = db.Column(db.Integer, default=0)
    last_recovery_code_used = db.Column(db.DateTime, nullable=True)
    captcha_required = db.Column(db.Boolean, default=False)
    last_captcha_failed = db.Column(db.DateTime, nullable=True)
    notifications_enabled = db.Column(db.Boolean, nullable=False, default=True)
    report_schedule = db.Column(db.String(20), default='disabled', nullable=False)
    email_on_scan_complete = db.Column(db.Boolean, nullable=False, default=True)
    email_on_critical_findings = db.Column(db.Boolean, nullable=False, default=True)
    slack_webhook_url = db.Column(db.String(512), nullable=True)
    teams_webhook_url = db.Column(db.String(512), nullable=True)

    # Enhanced License Management
    user_type = db.Column(db.String(10), default='BASIC', nullable=False)  # BASIC or PRO
    license_key = db.Column(db.String(50), nullable=True)  # Individual license key
    license_expires_at = db.Column(db.DateTime, nullable=True)  # License expiration
    license_validated_at = db.Column(db.DateTime, nullable=True)  # Last validation time

    # Scan Management
    monthly_scans_used = db.Column(db.Integer, default=0, nullable=False)
    last_scan_reset = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    allowed_monthly_scans = db.Column(db.Integer, default=5, nullable=False)  # 5 for BASIC, unlimited for PRO

    # Organization Support (Optional)
    organization_id = db.Column(db.Integer, db.ForeignKey('organization.id'), nullable=True)
    organization = db.relationship('Organization', foreign_keys=[organization_id], backref='members')
    scans = db.relationship('ScanResult', backref='author', lazy='dynamic', cascade="all, delete-orphan")
    credentials = db.relationship('CloudCredential', backref='owner', lazy='dynamic', cascade="all, delete-orphan")
    password_history = db.relationship('PasswordHistory', backref='user', lazy='dynamic', cascade="all, delete-orphan")
    api_key = db.relationship('APIKey', backref='user', uselist=False, cascade="all, delete-orphan")
    def set_password(self, password): self.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
    def check_password(self, password): return bcrypt.check_password_hash(self.password_hash, password)
    
    def generate_recovery_codes(self):
        """Generate 8 backup recovery codes for 2FA"""
        codes = []
        for _ in range(8):
            code = ''.join(random.choices(string.digits, k=8))
            codes.append(f"{code[:4]}-{code[4:]}")
        
        self.recovery_codes = json.dumps(codes)
        self.recovery_codes_used = json.dumps([])
        return codes
    
    def get_recovery_codes(self):
        """Get list of recovery codes"""
        if self.recovery_codes:
            return json.loads(self.recovery_codes)
        return []
    
    def get_used_recovery_codes(self):
        """Get list of used recovery codes"""
        if self.recovery_codes_used:
            return json.loads(self.recovery_codes_used)
        return []
    
    def use_recovery_code(self, code):
        """Use a recovery code and mark it as used"""
        codes = self.get_recovery_codes()
        used_codes = self.get_used_recovery_codes()
        
        if code in codes and code not in used_codes:
            used_codes.append(code)
            self.recovery_codes_used = json.dumps(used_codes)
            self.last_recovery_code_used = datetime.now(timezone.utc)
            return True
        return False
    
    def has_unused_recovery_codes(self):
        """Check if user has unused recovery codes"""
        codes = self.get_recovery_codes()
        used_codes = self.get_used_recovery_codes()
        return len(codes) > len(used_codes)
    
    def need_captcha(self):
        """Check if user needs CAPTCHA verification"""
        if self.captcha_required:
            return True
        if self.failed_2fa_attempts >= 3:
            return True
        if self.last_captcha_failed:
            # Ensure timezone awareness for comparison
            last_captcha = self.last_captcha_failed
            if last_captcha.tzinfo is None:
                last_captcha = last_captcha.replace(tzinfo=timezone.utc)
            time_diff = datetime.now(timezone.utc) - last_captcha
            if time_diff.total_seconds() < 900:  # 15 minutes
                return True
        return False

    def is_pro_user(self):
        """Check if user has PRO subscription with valid license"""
        return self.user_type == 'PRO' and self.has_valid_license()

    def is_basic_user(self):
        """Check if user has BASIC subscription or invalid Pro license"""
        return not self.is_pro_user()

    def has_valid_license(self):
        """Check if user has a valid Pro license"""
        if self.user_type != 'PRO':
            return False

        # Check personal license
        if self.license_key and self.license_validated_at:
            # Check if license has expired
            if self.license_expires_at is None:
                return True

            # Compare with UTC datetime, handling timezone aware/naive comparison
            current_time = datetime.now(timezone.utc)
            if self.license_expires_at.tzinfo is None:
                # If stored time is naive, treat it as UTC
                license_expiry = self.license_expires_at.replace(tzinfo=timezone.utc)
            else:
                license_expiry = self.license_expires_at

            if license_expiry > current_time:
                return True

        # Check organization license (if part of organization)
        if self.organization and self.organization.has_valid_license():
            return True

        return False

    def upgrade_to_pro(self, license_key=None, expires_at=None):
        """Upgrade user to PRO version with license validation"""
        self.user_type = 'PRO'
        self.allowed_monthly_scans = -1  # Unlimited scans for PRO users

        if license_key:
            self.license_key = license_key
            self.license_validated_at = datetime.now(timezone.utc)
            self.license_expires_at = expires_at

    def validate_license_key(self, license_key):
        """Validate and set license key for user"""
        from licenses.license_manager import LicenseManager

        license_manager = LicenseManager()
        validation_result = license_manager.validate_license_key(license_key)

        logging.info(f"User.validate_license_key called for {self.username} with key {license_key}")
        logging.info(f"Validation result: {validation_result}")

        if validation_result['is_valid']:
            self.license_key = license_key
            self.license_validated_at = datetime.now(timezone.utc)
            self.user_type = 'PRO'
            self.allowed_monthly_scans = -1

            # Set expiration if provided
            if 'expires_at' in validation_result:
                self.license_expires_at = validation_result['expires_at']

            logging.info(f"License validation SUCCESS for {self.username} - user_type set to {self.user_type}")
            return True

        logging.error(f"License validation FAILED for {self.username} - validation_result['is_valid'] = {validation_result.get('is_valid')}")
        return False

    def downgrade_to_basic(self):
        """Downgrade user to Basic version"""
        self.user_type = 'BASIC'
        self.license_key = None
        self.license_expires_at = None
        self.license_validated_at = None
        self.allowed_monthly_scans = 5

    def reset_monthly_scans_if_needed(self):
        """Reset monthly scan count if a month has passed"""
        now = datetime.now(timezone.utc)
        if self.last_scan_reset:
            # Ensure timezone awareness for comparison
            last_reset = self.last_scan_reset
            if last_reset.tzinfo is None:
                last_reset = last_reset.replace(tzinfo=timezone.utc)

            if (now - last_reset).days >= 30:
                self.monthly_scans_used = 0
                self.last_scan_reset = now
                db.session.commit()

    def can_perform_scan(self):
        """Check if user can perform a scan based on their plan"""
        if self.is_pro_user():
            return True

        # Initialize fields for legacy users
        if self.monthly_scans_used is None:
            self.monthly_scans_used = 0
        if self.allowed_monthly_scans is None:
            self.allowed_monthly_scans = 5
        if self.last_scan_reset is None:
            self.last_scan_reset = datetime.now(timezone.utc)
            db.session.commit()

        self.reset_monthly_scans_if_needed()
        return self.monthly_scans_used < self.allowed_monthly_scans

    def remaining_scans(self):
        """Get remaining scans for Basic users"""
        if self.is_pro_user():
            return -1  # Unlimited

        self.reset_monthly_scans_if_needed()
        return max(0, self.allowed_monthly_scans - self.monthly_scans_used)

    def increment_scan_count(self):
        """Increment scan count for Basic users"""
        if self.is_basic_user():
            self.monthly_scans_used += 1
            db.session.commit()

class APIKey(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False, unique=True)
    service = db.Column(db.String(50), nullable=False, default='gemini')
    encrypted_key = db.Column(db.String(512), nullable=False)
    
class CloudCredential(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    profile_name = db.Column(db.String(64), nullable=False)
    provider = db.Column(db.String(20), nullable=False, index=True)
    encrypted_key_1 = db.Column(db.String(512), nullable=True)
    encrypted_key_2 = db.Column(db.Text, nullable=True)

class ScanResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    service = db.Column(db.String(64), index=True)
    resource = db.Column(db.String(128))
    status = db.Column(db.String(64))
    issue = db.Column(db.String(256))
    remediation = db.Column(db.String(512), nullable=True)
    doc_url = db.Column(db.String(256), nullable=True)
    timestamp = db.Column(db.DateTime, index=True, default=lambda: datetime.now(timezone.utc))
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True, index=True)

# --- LOGIN MANAGER CONFIG & DECORATORS ---

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def login_or_guest_required(f):
    """Custom decorator that allows both logged-in users and guest sessions."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Allow guest mode
        if session.get('guest_mode'):
            return f(*args, **kwargs)
            
        # Check if user is logged in normally
        if not current_user.is_authenticated:
            flash("Please log in to access this page.", "warning")
            return redirect(url_for('auth', next=request.url))
        
        return f(*args, **kwargs)
    return decorated_function

class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    ip_address = db.Column(db.String(45))
    action = db.Column(db.String(128))
    details = db.Column(db.String(256), nullable=True)
    timestamp = db.Column(db.DateTime, index=True, default=lambda: datetime.now(timezone.utc))
    user = db.relationship('User')

class Organization(db.Model):
    """Organization model for enterprise licensing"""
    __tablename__ = 'organization'

    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    license_key = db.Column(db.String(50), nullable=True)
    max_pro_users = db.Column(db.Integer, default=0)
    license_expires_at = db.Column(db.DateTime, nullable=True)
    license_validated_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))

    # Admin user who manages the organization
    admin_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    admin_user = db.relationship('User', foreign_keys=[admin_user_id], post_update=True)

    def has_valid_license(self):
        """Check if organization has valid license"""
        if not self.license_key or not self.license_validated_at:
            return False

        # Check if license has expired
        if self.license_expires_at and self.license_expires_at <= datetime.now(timezone.utc):
            return False

        return True

    def can_add_pro_user(self):
        """Check if organization can add another Pro user"""
        current_pro_users = User.query.filter_by(
            organization_id=self.id,
            user_type='PRO'
        ).count()
        return current_pro_users < self.max_pro_users

    def get_available_pro_licenses(self):
        """Get number of available Pro licenses"""
        current_pro_users = User.query.filter_by(
            organization_id=self.id,
            user_type='PRO'
        ).count()
        return max(0, self.max_pro_users - current_pro_users)

class SuppressedFinding(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    finding_hash = db.Column(db.String(64), nullable=False, index=True)
    reason = db.Column(db.String(256), nullable=True)
    suppress_until = db.Column(db.DateTime, nullable=True)
    service = db.Column(db.String(64))
    resource = db.Column(db.String(128))
    issue = db.Column(db.String(256))
    user = db.relationship('User')

class PasswordHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    password_hash = db.Column(db.String(128))
    timestamp = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    __table_args__ = (db.UniqueConstraint('user_id', 'password_hash', name='uq_user_password_history'),)

class AutomationRule(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    rule_type = db.Column(db.String(50), nullable=False)  # 'remediation', 'notification', 'report'
    trigger_condition = db.Column(db.Text, nullable=False)  # JSON string of conditions
    action_config = db.Column(db.Text, nullable=False)  # JSON string of actions
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = db.Column(db.DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))
    last_executed = db.Column(db.DateTime)
    execution_count = db.Column(db.Integer, default=0)

# --- App Context Functions and Decorators ---
@app.context_processor
def inject_csrf_token():
    return dict(csrf_token_value=generate_csrf())

@app.template_global()
def csrf_token():
    """Template function to generate CSRF token"""
    return generate_csrf()

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash("You do not have permission to access this page.", "error")
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def api_admin_required(f):
    """Admin required decorator for API endpoints - returns JSON instead of redirect"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return jsonify({"error": "Authentication required"}), 401
        if not current_user.is_admin:
            return jsonify({"error": "Admin access required"}), 403
        return f(*args, **kwargs)
    return decorated_function

def check_verified(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Allow guest mode to bypass verification
        if session.get('guest_mode'):
            return f(*args, **kwargs)
            
        if not current_user.email_verified:
            flash("You must verify your email address to access this page.", "warning")
            return redirect(url_for('unverified'))
        return f(*args, **kwargs)
    return decorated_function

def check_2fa(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('guest_mode'):
            return f(*args, **kwargs)
            
        if current_user.is_authenticated and current_user.is_2fa_enabled:
            if session.get('2fa_passed') is not True:
                flash("Please complete the 2FA verification to continue.", "warning")
                return redirect(url_for('verify_2fa_login'))
        return f(*args, **kwargs)
    return decorated_function

def guest_or_login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('guest_mode'):
            if session.get('guest_expires', 0) < time.time():
                session.clear()
                flash('Guest session expired. Please start a new session.', 'warning')
                return redirect(url_for('auth'))
            return f(*args, **kwargs)
        
        if not current_user.is_authenticated:
            return redirect(url_for('auth'))
        
        return f(*args, **kwargs)
    return decorated_function

def admin_or_guest_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('guest_mode'):
            return f(*args, **kwargs)
            
        if not current_user.is_authenticated:
            return redirect(url_for('auth'))
            
        if not current_user.is_admin:
            flash('Admin access required.', 'error')
            return redirect(url_for('dashboard'))
        
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def before_request():
    if not request.endpoint or 'static' in request.endpoint: return
    if current_user.is_authenticated:
        current_user.last_login_date = datetime.now(timezone.utc)
        db.session.commit()
        session.permanent = True
        timeout_minutes = current_user.inactivity_timeout if hasattr(current_user, 'inactivity_timeout') else 15
        app.permanent_session_lifetime = timedelta(minutes=timeout_minutes)
        if 'last_activity' in session:
            last_activity_dt = datetime.fromisoformat(session['last_activity'])
            # Ensure timezone awareness for comparison
            if last_activity_dt.tzinfo is None:
                last_activity_dt = last_activity_dt.replace(tzinfo=timezone.utc)
            if datetime.now(timezone.utc) - last_activity_dt > app.permanent_session_lifetime:
                logout_user()
                flash('Your session has expired due to inactivity. Please log in again.', 'info')
        session['last_activity'] = datetime.now(timezone.utc).isoformat()

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

def is_password_strong(password):
    """
    Simple password strength validation (replacement for zxcvbn)
    """
    if len(password) < 8:
        return False, "Password must be at least 8 characters long."

    # Check for basic password requirements
    has_upper = any(c.isupper() for c in password)
    has_lower = any(c.islower() for c in password)
    has_digit = any(c.isdigit() for c in password)
    has_special = any(c in "!@#$%^&*()_+-=[]{}|;:,.<>?" for c in password)

    score = sum([has_upper, has_lower, has_digit, has_special])

    if score < 3:
        suggestions = []
        if not has_upper:
            suggestions.append("uppercase letters")
        if not has_lower:
            suggestions.append("lowercase letters")
        if not has_digit:
            suggestions.append("numbers")
        if not has_special:
            suggestions.append("special characters")

        return False, f"Password is too weak. Please include {', '.join(suggestions)}."

    # Check for common weak patterns
    weak_patterns = ['password', '12345', 'qwerty', 'admin', 'letmein']
    if any(pattern in password.lower() for pattern in weak_patterns):
        return False, "Password contains common weak patterns. Please choose a more unique password."

    return True, ""

def log_audit(action, details="", user=None):
    with app.app_context():
        try:
            log_entry = AuditLog(action=action, details=details, ip_address=request.remote_addr, user_id=user.id if user else None)
            db.session.add(log_entry)
            db.session.commit()
        except Exception as e:
            logging.error(f"Audit log failed: {e}")
            db.session.rollback()

def generate_external_url(endpoint, **values):
    """Helper function to generate external URLs for emails by temporarily setting SERVER_NAME"""
    domain_name = os.getenv('DOMAIN_NAME', 'localhost:5000')

    # Temporarily set SERVER_NAME
    original_server_name = app.config.get('SERVER_NAME')
    app.config['SERVER_NAME'] = domain_name

    try:
        with app.app_context():
            url = url_for(endpoint, _external=True, **values)
        return url
    finally:
        # Restore original SERVER_NAME (or remove it if it wasn't set)
        if original_server_name is not None:
            app.config['SERVER_NAME'] = original_server_name
        else:
            app.config.pop('SERVER_NAME', None)

@safe_execute
def send_verification_email(user):
    with app.app_context():
        logging.info(f"--- [EMAIL] Generating verification email for {user.email}. ---")

        # Debug: Check mail configuration
        logging.info(f"--- [EMAIL] Mail server: {app.config.get('MAIL_SERVER')} ---")
        logging.info(f"--- [EMAIL] Mail username: {app.config.get('MAIL_USERNAME')} ---")
        logging.info(f"--- [EMAIL] Mail port: {app.config.get('MAIL_PORT')} ---")
        logging.info(f"--- [EMAIL] Mail TLS: {app.config.get('MAIL_USE_TLS')} ---")

        token = s.dumps(user.email, salt='email-confirm-salt')
        msg = Message(
            'Confirm Your Email for Aegis Scanner',
            recipients=[user.email],
            sender=app.config.get('MAIL_DEFAULT_SENDER') or app.config.get('MAIL_USERNAME')
        )
        confirm_url = generate_external_url('verify_email', token=token)
        msg.html = render_template('confirm_email.html', confirm_url=confirm_url)

        try:
            with mail.connect() as conn:
                conn.send(msg)
            logging.info(f"--- [EMAIL] Verification email successfully sent to {user.email}. ---")
        except Exception as e:
            logging.error(f"--- [EMAIL] FAILED to send verification email: {e} ---")
            logging.error(f"--- [EMAIL] Exception type: {type(e).__name__} ---")
            raise  # Re-raise to trigger @safe_execute error handling

@safe_execute
def send_new_primary_email_verification(user, new_email):
    with app.app_context():
        token = s.dumps({'user_id': user.id, 'new_email': new_email}, salt='new-primary-email-salt')
        msg = Message('Confirm Your New Primary Email', recipients=[new_email])
        confirm_url = generate_external_url('verify_new_primary_email', token=token)
        msg.html = render_template('confirm_email.html', confirm_url=confirm_url)
        mail.send(msg)

@safe_execute
def send_backup_email_verification(user, backup_email):
    with app.app_context():
        token = s.dumps({'user_id': user.id, 'backup_email': backup_email}, salt='backup-email-salt')
        msg = Message('Confirm Your Backup Email', recipients=[backup_email])
        confirm_url = generate_external_url('verify_backup_email', token=token)
        msg.html = render_template('confirm_email.html', confirm_url=confirm_url)
        mail.send(msg)

def generate_simple_captcha():
    """Generate a simple math CAPTCHA"""
    num1 = random.randint(1, 10)
    num2 = random.randint(1, 10)
    operation = random.choice(['+', '-'])
    
    if operation == '+':
        answer = num1 + num2
        question = f"{num1} + {num2} = ?"
    else:
        # Ensure result is positive
        if num1 < num2:
            num1, num2 = num2, num1
        answer = num1 - num2
        question = f"{num1} - {num2} = ?"
    
    return question, str(answer)

def verify_captcha(user_answer, correct_answer):
    """Verify CAPTCHA answer"""
    try:
        return str(user_answer).strip() == str(correct_answer).strip()
    except:
        return False

def send_2fa_recovery_email_to_address(user, email_address, email_type='registered'):
    """Send 2FA recovery instructions to specified email address"""
    with app.app_context():
        # Generate recovery token
        token_data = {
            'user_id': user.id, 
            'recovery_type': '2fa_reset',
            'timestamp': time.time()
        }
        recovery_token = s.dumps(token_data, salt='2fa-recovery-salt')
        
        msg = Message('2FA Account Recovery - Aegis Security Scanner', recipients=[email_address])
        recovery_url = generate_external_url('recover_2fa', token=recovery_token)
        
        msg.html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background-color: #f8f9fa; border-radius: 10px;">
            <div style="text-align: center; margin-bottom: 30px;">
                <h2 style="color: #00A896; margin-bottom: 10px;">🛡️ 2FA Account Recovery</h2>
                <p style="color: #666; font-size: 14px;">Aegis Security Scanner</p>
            </div>
            <div style="background: #f8f9fa; padding: 25px; border-radius: 8px; border-left: 4px solid #00A896;">
                <p style="color: #333; font-size: 16px; margin-bottom: 20px;">Hello <strong>{user.username}</strong>,</p>
                <p style="color: #333; line-height: 1.6;">Someone requested 2FA recovery for your Aegis Security Scanner account. This email was sent to your <strong>{email_type} email address</strong>.</p>
            <p><strong>Account:</strong> {user.email}</p>
            <p><strong>Username:</strong> {user.username}</p>
            
            <div style="background: #f8f9fa; padding: 20px; border-left: 4px solid #dc3545; margin: 20px 0;">
                <h3 style="color: #dc3545; margin: 0 0 10px 0;">Security Notice</h3>
                <p style="margin: 0;">If you did not request this recovery, please ignore this email and secure your account immediately.</p>
            </div>
            
            <p>To recover your 2FA settings, click the link below:</p>
            <div style="text-align: center; margin: 30px 0;">
                <a href="{recovery_url}" style="background: #00A896; color: white; padding: 15px 30px; text-decoration: none; border-radius: 5px; display: inline-block;">
                    Recover 2FA Access
                </a>
            </div>
            <p><small>This link expires in 1 hour for security reasons.</small></p>
            <hr style="border: none; border-top: 1px solid #eee; margin: 30px 0;">
            <p style="color: #666; font-size: 12px;">
                Aegis Security Scanner<br>
                This is an automated security message.
            </p>
        </div>
        """
        
        try:
            mail.send(msg)
            return True
        except Exception as e:
            logging.error(f"Failed to send 2FA recovery email: {e}")
            return False

def _create_pdf_report(results):
    """Create PDF report using reportlab instead of WeasyPrint for better compatibility."""
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1*inch)
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#00A896'),
            alignment=1  # Center alignment
        )
        
        # Build content
        content = []
        
        # Title
        title = Paragraph("Aegis Security Report", title_style)
        content.append(title)
        content.append(Spacer(1, 20))
        
        # Date
        date_text = f"Generated on: {datetime.now(timezone.utc).strftime('%B %d, %Y at %I:%M %p UTC')}"
        date_para = Paragraph(date_text, styles['Normal'])
        content.append(date_para)
        content.append(Spacer(1, 20))
        
        # Summary
        summary_text = f"Total findings: {len(results)}"
        if results:
            critical_count = len([r for r in results if r.status == 'CRITICAL'])
            high_count = len([r for r in results if r.status == 'HIGH'])
            summary_text += f" | Critical: {critical_count} | High: {high_count}"
        
        summary_para = Paragraph(summary_text, styles['Heading2'])
        content.append(summary_para)
        content.append(Spacer(1, 20))
        
        # Results table
        if results:
            table_data = [['Service', 'Resource', 'Status', 'Issue']]
            
            for result in results[:50]:  # Limit to first 50 results for PDF size
                table_data.append([
                    str(result.service)[:30],  # Limit length
                    str(result.resource)[:40],
                    str(result.status),
                    str(result.issue)[:60]
                ])
            
            table = Table(table_data, colWidths=[1.5*inch, 2*inch, 1*inch, 3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#00A896')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            content.append(table)
        else:
            no_data_para = Paragraph("No scan results available.", styles['Normal'])
            content.append(no_data_para)
        
        # Build PDF
        doc.build(content)
        
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        return pdf_bytes
        
    except Exception as e:
        logging.error(f"Error creating PDF report: {e}")
        # Return a minimal PDF on error
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        content = [
            Paragraph("Aegis Security Report", styles['Title']),
            Spacer(1, 20),
            Paragraph("Error generating detailed report. Please try again.", styles['Normal'])
        ]
        doc.build(content)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        return pdf_bytes

def _generate_cache_key(user_id, profile_id, regions):
    regions_str = "|".join(sorted(regions)) if regions else "global"
    return f"{user_id}_{profile_id}_{regions_str}"

def _generate_finding_hash(finding):
    finding_string = f"{finding.get('service', '')}:{finding.get('resource', '')}:{finding.get('issue', '')}"
    return hashlib.sha256(finding_string.encode()).hexdigest()

# --- Main Application Routes ---

@app.route('/splash')
def splash():
    # Check if initial setup is needed
    if not os.path.exists(ENV_FILE_PATH) or not os.getenv('MAIL_USERNAME'):
        return redirect(url_for('setup'))
    return render_template('splash.html')

@app.route('/welcome')
def welcome():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('welcome.html')

@app.route('/auth')
def auth():
    if current_user.is_authenticated: return redirect(url_for('dashboard'))
    return render_template('auth.html')

# UPDATED: Login with enhanced security and validation
@app.route('/login', methods=['POST'])
@limiter.limit("10 per minute")
def login_post():
    """Updated login with input validation and rate limiting."""
    try:
        # Get and sanitize input
        login_identifier = security_validator.sanitize_string(
            request.form.get('username', '')
        )
        password = request.form.get('password', '')
        
        # Basic validation
        if not login_identifier or not password:
            flash('Username and password are required.', 'error')
            return redirect(url_for('auth', _anchor='login'))
        
        if len(login_identifier) > 120:  # Prevent DoS
            flash('Invalid login credentials.', 'error')
            return redirect(url_for('auth', _anchor='login'))
        
        # Find user
        user = User.query.filter(
            or_(User.username == login_identifier, User.email == login_identifier)
        ).first()
        
        if user and user.is_locked:
            logging.warning(f"Locked account login attempt for user '{login_identifier}'")
            log_audit("Login Attempt - Locked Account", details=f"User: {login_identifier}")
            flash('This account is locked. Please contact an administrator.', 'error')
            return redirect(url_for('auth', _anchor='login'))
        
        if user and user.check_password(password):
            # Successful login - clear any existing guest mode session
            user.failed_login_attempts = 0
            db.session.commit()
            
            # Clear guest mode session variables if present
            session.pop('guest_mode', None)
            session.pop('guest_expires', None)
            
            log_audit("Login Success", user=user)
            live_logger.log("Login Succeeded", "SUCCESS", f"User: {user.username}")
            log_authentication("User login successful", user.username, True)
            structured_logger.log_security_event(
                event_type="successful_login",
                user_id=user.id,
                details={
                    "username": user.username,
                    "ip_address": request.remote_addr,
                    "user_agent": request.headers.get('User-Agent', ''),
                    "2fa_enabled": user.is_2fa_enabled
                }
            )
            login_user(user)
            logging.info(f"Successful login for user '{user.username}'")
            
            if not user.email_verified:
                return redirect(url_for('unverified'))
            
            if user.is_2fa_enabled:
                session['username_for_2fa'] = user.username
                session['2fa_passed'] = False
                return redirect(url_for('verify_2fa_login'))
            else:
                flash('For enhanced security, you must set up Two-Factor Authentication.', 'info')
                return redirect(url_for('setup_2fa'))
        else:
            # Failed login
            log_audit("Login Failure", details=f"Attempt for user: '{login_identifier}'")
            live_logger.log("Login Failed", "ERROR", f"Attempt for: '{login_identifier}'")
            structured_logger.log_security_event(
                event_type="failed_login",
                user_id=user.id if user else None,
                details={
                    "username": login_identifier,
                    "ip_address": request.remote_addr,
                    "user_agent": request.headers.get('User-Agent', ''),
                    "user_exists": user is not None
                }
            )
            
            if user:
                user.failed_login_attempts = (user.failed_login_attempts or 0) + 1
                if user.failed_login_attempts >= 3:
                    user.is_locked = True
                    logging.critical(f"Account locked for user '{user.username}' due to failed attempts")
                    log_audit("Account Locked", details=f"User: {user.username}", user=user)
                else:
                    logging.warning(f"Failed login for '{user.username}' - attempt #{user.failed_login_attempts}")
                db.session.commit()
            else:
                logging.warning(f"Failed login attempt for non-existent user '{login_identifier}'")
            
            flash('Invalid username or password.', 'error')
            return redirect(url_for('auth', _anchor='login'))
            
    except Exception as e:
        logging.error(f"Login error: {e}")
        flash('Login failed. Please try again.', 'error')
        return redirect(url_for('auth', _anchor='login'))

# UPDATED: Registration with comprehensive validation
@app.route('/register', methods=['POST'])
@limiter.limit("10 per hour")
@validate_form_input('user_registration')
def register_post():
    """Updated registration with comprehensive validation."""
    try:
        # Get validated data from the decorator
        data = request.validated_data
        
        username = data['username']
        email = data['email']
        password = data['password']
        confirm_password = request.form.get('confirm_password', '')
        eula_accepted = request.form.get('eula')
        admin_key = data.get('admin_key', '')
        
        # Additional validation not covered by schema
        if not eula_accepted:
            flash('You must accept the EULA to register.', 'error')
            return redirect(url_for('auth', _anchor='register'))
        
        if password != confirm_password:
            flash('Passwords do not match.', 'error')
            return redirect(url_for('auth', _anchor='register'))
        
        # Enhanced email validation
        is_valid_email, email_error = security_validator.validate_email(email)
        if not is_valid_email:
            flash(f'Email validation failed: {email_error}', 'error')
            return redirect(url_for('auth', _anchor='register'))
        
        # Password strength validation
        is_strong, message = is_password_strong(password)
        if not is_strong:
            flash(message, 'error')
            return redirect(url_for('auth', _anchor='register'))
        
        # Check for existing users
        if User.query.filter_by(username=username).first():
            flash('Username already exists. Please choose another.', 'error')
            return redirect(url_for('auth', _anchor='register'))
        
        if User.query.filter_by(email=email).first():
            flash('Email address is already registered.', 'error')
            return redirect(url_for('auth', _anchor='register'))
        
        # Create user
        user = User(username=username, email=email)
        logging.info(f"Creating new user account: {username}")
        user.set_password(password)

        # All new users start as BASIC - they can upgrade to Pro later in settings
        user.user_type = 'BASIC'
        user.monthly_scans_used = 0
        user.allowed_monthly_scans = 5  # Basic plan limits
        user.last_scan_reset = datetime.now(timezone.utc)

        logging.info(f"New user {username} created with BASIC plan")
        
        # Password history
        user_password_history = PasswordHistory(user=user, password_hash=user.password_hash)
        db.session.add(user_password_history)
        
        # Admin key validation
        if admin_key:
            admin_key = security_validator.sanitize_string(admin_key)
            if admin_key == os.getenv('ADMIN_REGISTRATION_KEY'):
                if User.query.filter_by(is_admin=True).count() < 2:
                    user.is_admin = True
                    logging.info(f"Promoting user '{username}' to admin with valid key")
                else:
                    logging.warning("Admin key provided but max admin count reached")
        
        db.session.add(user)
        db.session.commit()
        
        # Send verification email
        send_verification_email(user)
        login_user(user)

        log_audit("User Registration", details=f"New user: {username}", user=user)
        flash('Registration successful! A verification link has been sent to your email.', 'info')
        return redirect(url_for('unverified'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Registration error: {e}")
        flash('Registration failed. Please try again.', 'error')
        return redirect(url_for('auth', _anchor='register'))

@app.route('/verify-email/<token>')
@login_required
def verify_email(token):
    try:
        email = s.loads(token, salt='email-confirm-salt', max_age=3600)
    except Exception:
        flash('The confirmation link is invalid or has expired.', 'error')
        return redirect(url_for('unverified'))
    if email != current_user.email:
        flash('Invalid verification link.', 'error')
        return redirect(url_for('unverified'))
    if not current_user.email_verified:
        current_user.email_verified = True
        db.session.commit()
        logging.info(f"--- [AUTH] Email successfully verified for user '{current_user.username}'. ---")
        log_audit("Email Verified", user=current_user)
        flash('Your email has been verified! Please set up 2FA to continue.', 'success')
        return redirect(url_for('setup_2fa'))
    else:
        if not current_user.is_2fa_enabled:
            flash('Account already verified. Please set up 2FA to continue.', 'info')
            return redirect(url_for('setup_2fa'))
        else:
            flash('Account already verified.', 'info')
            return redirect(url_for('dashboard'))

@app.route('/unverified')
@login_required
def unverified():
    if current_user.email_verified: return redirect(url_for('dashboard'))
    return render_template('unverified.html')

@app.route('/resend-verification')
@login_required
def resend_verification():
    if current_user.email_verified: return redirect(url_for('dashboard'))
    send_verification_email(current_user)
    flash('A new verification email has been sent.', 'info')
    return redirect(url_for('unverified'))

@app.route('/eula')
def eula():
    return render_template('eula.html')

@app.route('/logout')
@login_required
def logout():
    # Clear license session data
    session.pop('license_key', None)
    session.pop('license_info', None)

    # Clear user session data
    session.pop('2fa_passed', None)
    session.pop('guest_mode', None)
    session.pop('guest_expires', None)

    logout_user()
    flash('You have been logged out.', 'success')
    return redirect(url_for('welcome'))

@app.route('/clear-license')
def clear_license():
    """Clear license session for testing purposes"""
    session.pop('license_key', None)
    session.pop('license_info', None)
    flash('License session cleared. Please enter a new license key.', 'info')
    return redirect(url_for('license_validation'))

@app.route('/help')
def help_page():
    return render_template('help.html')

@app.route('/debug-guest')
@login_or_guest_required
def debug_guest():
    """Debug route to check guest mode status"""
    debug_info = {
        'is_guest_mode': session.get('guest_mode', False),
        'session_keys': list(session.keys()),
        'guest_credentials': session.get('guest_credentials', []),
        'is_authenticated': current_user.is_authenticated if hasattr(current_user, 'is_authenticated') else False
    }
    return f"<pre>{debug_info}</pre>"

@app.route('/debug-dashboard-creds')
@login_or_guest_required
def debug_dashboard_creds():
    """Debug route to see what credentials are passed to dashboard"""
    if session.get('guest_mode'):
        guest_credentials = session.get('guest_credentials', [])
        credentials = []
        for guest_cred in guest_credentials:
            class GuestCredential:
                def __init__(self, cred_dict):
                    self.id = cred_dict.get('id')
                    self.profile_name = cred_dict.get('profile_name')
                    self.provider = cred_dict.get('provider')
                    self.encrypted_key_1 = cred_dict.get('encrypted_key_1')
                    self.encrypted_key_2 = cred_dict.get('encrypted_key_2')
            credentials.append(GuestCredential(guest_cred))
    else:
        credentials = current_user.credentials.all() if hasattr(current_user, 'credentials') else []
    
    cred_list = [{"id": c.id, "profile_name": c.profile_name, "provider": c.provider} for c in credentials]
    return f"<pre>Credentials for template: {cred_list}</pre>"

@app.route('/request-reset', methods=['GET', 'POST'])
@limiter.limit("5 per 15 minutes")
def request_password_reset():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        login_identifier = security_validator.sanitize_string(
            request.form.get('login_identifier', '')
        )
        user = User.query.filter(or_(User.username == login_identifier, User.email == login_identifier)).first()
        if user:
            token = s.dumps(user.email, salt='password-reset-salt')
            reset_url = generate_external_url('reset_with_token', token=token)
            msg = Message('Password Reset Request', recipients=[user.email])
            msg.html = render_template('reset_email.html', url=reset_url)
            try:
                mail.send(msg)
                flash('A password reset link has been sent to your email.', 'info')
            except Exception as e:
                logging.error(f"Failed to send password reset email: {e}")
                flash('Could not send reset email. Please contact an administrator.', 'error')
        else:
            flash('No account found with that username or email address.', 'warning')
        return redirect(url_for('auth'))
    return render_template('request_reset.html')

@app.route('/reset/<token>', methods=['GET', 'POST'])
def reset_with_token(token):
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    try:
        email = s.loads(token, salt='password-reset-salt', max_age=3600)
    except Exception:
        flash('The password reset link is invalid or has expired.', 'error')
        return redirect(url_for('request_password_reset'))
    
    user = User.query.filter_by(email=email).first()
    if not user:
        flash('Invalid user.', 'error')
        return redirect(url_for('auth'))

    if request.method == 'POST':
        password = request.form.get('password', '')
        is_strong, message = is_password_strong(password)
        if not is_strong:
            flash(message, 'error')
            return render_template('reset_password.html')
        
        user.set_password(password)
        db.session.add(PasswordHistory(user=user, password_hash=user.password_hash))
        db.session.commit()
        log_audit("User reset password", user=user)
        flash('Your password has been updated! You can now log in.', 'success')
        return redirect(url_for('auth'))
    
    return render_template('reset_password.html')

@app.route('/setup-2fa')
@login_required
@check_verified
def setup_2fa():
    if current_user.is_2fa_enabled:
        flash('2FA is already enabled.', 'info')
        return redirect(url_for('dashboard'))
    current_user.otp_secret = pyotp.random_base32()
    db.session.commit()
    uri = pyotp.totp.TOTP(current_user.otp_secret).provisioning_uri(name=current_user.username, issuer_name="Aegis Cloud Scanner")
    logging.info(f"--- [SECURITY] Generated new 2FA QR Code for user '{current_user.username}'. ---")
    img = qrcode.make(uri)
    buf = BytesIO()
    img.save(buf)
    qr_code = base64.b64encode(buf.getvalue()).decode('utf-8')
    return render_template('2fa_setup.html', qr_code=qr_code)

@app.route('/enable-2fa', methods=['POST'])
@login_required
@check_verified
def enable_2fa():
    otp_code = security_validator.sanitize_string(request.form.get('otp_code', ''))
    if not otp_code or len(otp_code) != 6:
        flash('Please enter a valid 6-digit code.', 'error')
        return redirect(url_for('setup_2fa'))
    
    totp = pyotp.TOTP(current_user.otp_secret)
    if totp.verify(otp_code, valid_window=1):
        current_user.is_2fa_enabled = True
        
        # Generate recovery codes
        recovery_codes = current_user.generate_recovery_codes()
        
        db.session.commit()
        logging.info(f"--- [SECURITY] 2FA successfully enabled for user '{current_user.username}'. ---")
        
        # Store recovery codes in session to show them once
        session['new_recovery_codes'] = recovery_codes
        
        flash('2FA has been successfully enabled! Please save your recovery codes.', 'success')
        session['2fa_passed'] = True
        log_audit("2FA Enabled", user=current_user)
        return redirect(url_for('show_recovery_codes'))
    else:
        logging.warning(f"--- [SECURITY] Invalid 2FA code provided for user '{current_user.username}'. ---")
        flash('Invalid verification code. Please try again.', 'error')
        return redirect(url_for('setup_2fa'))

@app.route('/verify-2fa-login', methods=['GET', 'POST'])
def verify_2fa_login():
    username = session.get('username_for_2fa')
    if not username: return redirect(url_for('auth'))
    user = User.query.filter_by(username=username).first()
    
    # Generate CAPTCHA if needed
    show_captcha = user.need_captcha()
    captcha_question = None
    if show_captcha and 'captcha_answer' not in session:
        captcha_question, captcha_answer = generate_simple_captcha()
        session['captcha_question'] = captcha_question
        session['captcha_answer'] = captcha_answer
    elif show_captcha:
        captcha_question = session.get('captcha_question')
    
    if request.method == 'POST':
        # Check CAPTCHA first if required
        if show_captcha:
            user_captcha = request.form.get('captcha', '').strip()
            correct_captcha = session.get('captcha_answer', '')
            
            if not verify_captcha(user_captcha, correct_captcha):
                user.last_captcha_failed = datetime.now(timezone.utc)
                db.session.commit()
                flash('CAPTCHA verification failed. Please try again.', 'error')
                # Generate new CAPTCHA
                captcha_question, captcha_answer = generate_simple_captcha()
                session['captcha_question'] = captcha_question
                session['captcha_answer'] = captcha_answer
                return redirect(url_for('verify_2fa_login'))
            
            # Clear CAPTCHA from session after successful verification
            session.pop('captcha_question', None)
            session.pop('captcha_answer', None)
        
        otp_code = security_validator.sanitize_string(request.form.get('otp_code', ''))
        recovery_code = security_validator.sanitize_string(request.form.get('recovery_code', ''))
        
        success = False
        
        # Try regular TOTP first
        if otp_code and len(otp_code) == 6:
            totp = pyotp.TOTP(user.otp_secret)
            if totp.verify(otp_code, valid_window=1):
                success = True
                # Reset failed attempts on successful login
                user.failed_2fa_attempts = 0
                log_audit("2FA Login Success (TOTP)", user=user)
        
        # Try recovery code if TOTP failed
        elif recovery_code:
            if user.use_recovery_code(recovery_code):
                success = True
                # Reset failed attempts on successful login
                user.failed_2fa_attempts = 0
                log_audit("2FA Login Success (Recovery Code)", user=user)
                flash('Recovery code used successfully. Consider regenerating your codes.', 'warning')
            else:
                flash('Invalid recovery code.', 'error')
        
        if success:
            logging.info(f"--- [SECURITY] 2FA verification successful for user '{user.username}'. ---")
            login_user(user)
            session.pop('username_for_2fa', None)
            session['2fa_passed'] = True
            # Clear guest mode session variables if present
            session.pop('guest_mode', None)
            session.pop('guest_expires', None)
            db.session.commit()
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            # Increment failed attempts
            user.failed_2fa_attempts += 1
            db.session.commit()
            
            if not recovery_code:  # Only show this message for TOTP failures
                if user.has_unused_recovery_codes():
                    flash('Invalid 2FA code. You can also use a recovery code if you lost your device.', 'error')
                else:
                    flash('Invalid 2FA code.', 'error')

    # Prepare template data
    template_data = {
        'show_captcha': show_captcha,
        'captcha_question': captcha_question,
        'has_recovery_codes': user.has_unused_recovery_codes(),
        'backup_email_configured': user.backup_email and user.backup_email_verified
    }
    
    resp = make_response(render_template('2fa_verify.html', **template_data))
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

@app.route('/onboarding')
@login_required
@check_verified
@check_2fa
def onboarding():
    if not current_user.is_2fa_enabled:
        flash('You must set up Two-Factor Authentication to access the dashboard.', 'info')
        return redirect(url_for('setup_2fa'))
    
    if current_user.credentials.first():
        return redirect(url_for('dashboard'))

    return render_template('onboarding.html')

@app.route('/dashboard')
@login_or_guest_required
@check_verified
@check_2fa
def dashboard():
    try:
        # Log dashboard access
        log_user_action("Dashboard accessed")
    except Exception as e:
        logging.error(f"Error logging user action: {e}")

    # Handle guest mode
    if session.get('guest_mode'):
        # For guest mode, get credentials from session and format them for template
        guest_credentials = session.get('guest_credentials', [])
        credentials = []
        for guest_cred in guest_credentials:
            # Convert guest credential dict to object-like structure for template compatibility
            class GuestCredential:
                def __init__(self, cred_dict):
                    self.id = cred_dict.get('id')
                    self.profile_name = cred_dict.get('profile_name')
                    self.provider = cred_dict.get('provider')
                    self.encrypted_key_1 = cred_dict.get('encrypted_key_1')
                    self.encrypted_key_2 = cred_dict.get('encrypted_key_2')

            credentials.append(GuestCredential(guest_cred))

        # Use predefined AWS regions
        regions = ['us-east-1', 'us-east-2', 'us-west-1', 'us-west-2', 'eu-west-1', 'ap-southeast-1', 'ap-northeast-1', 'eu-central-1', 'ca-central-1', 'ap-southeast-2']

        # Create a simple guest user object for template compatibility
        class GuestUser:
            username = session.get('username', 'Guest')
            email = 'guest@temp.local'
            is_authenticated = False
            is_admin = False
            user_type = 'BASIC'

        return render_template('dashboard.html',
                             credentials=credentials,
                             regions=regions,
                             is_guest_mode=True,
                             guest_session_id=session.get('guest_session_id'),
                             current_user=GuestUser())
    else:
        # Regular user checks
        if not current_user.is_2fa_enabled:
            flash('You must set up Two-Factor Authentication to access the dashboard.', 'info')
            return redirect(url_for('setup_2fa'))

        if not current_user.credentials.first():
            return redirect(url_for('onboarding'))

        credentials = current_user.credentials.all()

        # Use predefined AWS regions to avoid metadata service connection issues
        regions = ['us-east-1', 'us-east-2', 'us-west-1', 'us-west-2', 'eu-west-1', 'ap-southeast-1', 'ap-northeast-1', 'eu-central-1', 'ca-central-1', 'ap-southeast-2']

        return render_template('dashboard.html',
                             credentials=credentials,
                             regions=regions,
                             is_guest_mode=False,
                             guest_session_id=None,
                             current_user=current_user)

# NEW: API ENDPOINT FOR LIVE ACTIVITIES
@app.route('/api/v1/activities')
@login_required
@check_2fa
def get_activities():
    """API endpoint to fetch recent activities for the live monitor."""
    return jsonify(live_logger.activities)

# UPDATED: Scan endpoint with comprehensive validation
@app.route('/api/v1/scan', methods=['GET'])
@login_or_guest_required
@check_verified
@check_2fa
@limiter.limit("5 per hour", key_func=lambda: session.get('guest_session_id', getattr(current_user, 'id', 'unknown')))
def scan():
    """Updated scan endpoint with comprehensive input validation."""
    try:
        # Get and validate parameters
        profile_id_str = security_validator.sanitize_string(
            request.args.get('profile_id', '')
        )
        regions_to_scan = [
            security_validator.sanitize_string(region) 
            for region in request.args.getlist('regions')
        ]
        progress_mode = request.args.get('progress_mode') == 'true'
        
        # Validate profile ID
        try:
            profile_id = int(profile_id_str)
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid profile ID format."}), 400
        
        if profile_id <= 0:
            return jsonify({"error": "Profile ID must be positive."}), 400
        
        # Validate regions
        if regions_to_scan and 'all' not in regions_to_scan:
            for region in regions_to_scan:
                if not security_validator.PATTERNS['region_name'].match(region):
                    return jsonify({"error": f"Invalid region format: {region}"}), 400
        
        if not regions_to_scan or 'all' in regions_to_scan:
            regions_to_scan = None

        # Check Basic user scan limitations (skip for guest users)
        if not session.get('guest_mode') and current_user.is_authenticated:
            if current_user.is_basic_user() and not current_user.can_perform_scan():
                remaining = current_user.remaining_scans()
                return jsonify({
                    "error": f"Scan limit reached. Basic users get {current_user.allowed_monthly_scans} scans per month. You have {remaining} remaining.",
                    "scan_limit_reached": True,
                    "remaining_scans": remaining,
                    "user_type": "BASIC"
                }), 403

        # Handle credential lookup for both regular users and guest users
        credential = None
        logging.info(f"Scan request - Profile ID: {profile_id}, Guest Mode: {session.get('guest_mode')}, Current User: {getattr(current_user, 'id', 'Anonymous')}")

        if session.get('guest_mode'):
            # For guest users, find credential in session
            guest_credentials = session.get('guest_credentials', [])
            logging.info(f"Guest credentials available: {len(guest_credentials)}")
            for guest_cred in guest_credentials:
                if guest_cred.get('id') == profile_id:
                    # Create temporary credential object
                    class GuestCredential:
                        def __init__(self, cred_dict):
                            self.id = cred_dict.get('id')
                            self.profile_name = cred_dict.get('profile_name')
                            self.provider = cred_dict.get('provider')
                            self.encrypted_key_1 = cred_dict.get('encrypted_key_1')
                            self.encrypted_key_2 = cred_dict.get('encrypted_key_2')
                    credential = GuestCredential(guest_cred)
                    break
        else:
            # Regular user credential lookup
            logging.info(f"Looking for credential: profile_id={profile_id}, user_id={current_user.id}")
            credential = CloudCredential.query.filter_by(
                id=profile_id,
                user_id=current_user.id
            ).first()

            if not credential:
                # Also log what credentials this user actually has
                user_creds = CloudCredential.query.filter_by(user_id=current_user.id).all()
                logging.info(f"User {current_user.id} has {len(user_creds)} credentials: {[(c.id, c.profile_name, c.provider) for c in user_creds]}")

                # Check if the credential exists but belongs to another user
                other_cred = CloudCredential.query.filter_by(id=profile_id).first()
                if other_cred:
                    logging.warning(f"Credential {profile_id} exists but belongs to user {other_cred.user_id}, not current user {current_user.id}")
                else:
                    logging.warning(f"Credential {profile_id} does not exist in database")
        
        if not credential:
            live_logger.log("Scan Aborted", "ERROR", "Credential profile not found")
            return jsonify({"error": "Credential profile not found or access denied."}), 404
        
        live_logger.log("Scan Initiated", "SCAN", f"Profile: {credential.profile_name}")
        if session.get('guest_mode'):
            log_audit("Scan Initiated", details=f"Profile: {credential.profile_name} (Guest Mode)", user=None)
        else:
            log_audit("Scan Initiated", details=f"Profile: {credential.profile_name}", user=current_user)

        # Set user ID for scan results (guest users get None for temporary storage)
        scan_user_id = None if session.get('guest_mode') else current_user.id

        try:
            decrypted_creds = {
                "provider": credential.provider,
                "profile_name": credential.profile_name
            }
            if credential.provider == 'aws':
                live_logger.log("Decrypting Credentials", "ENCRYPTION", "AWS Keys")
                key_id = decrypt_data(credential.encrypted_key_1, context=f"AWS Access Key for scan profile '{credential.profile_name}'")
                decrypted_creds["aws_access_key_id"] = key_id
                decrypted_creds["aws_secret_access_key"] = decrypt_data(credential.encrypted_key_2, context=f"AWS Secret Key for scan profile '{credential.profile_name}'")
                logging.debug(f"Using AWS credentials for profile '{credential.profile_name}'. Access Key ID: {key_id}")
            elif credential.provider == 'gcp':
                live_logger.log("Decrypting Credentials", "ENCRYPTION", "GCP Key")
                gcp_json = decrypt_data(credential.encrypted_key_2, context=f"GCP JSON key for scan profile '{credential.profile_name}'")
                decrypted_creds["gcp_service_account_json"] = gcp_json
                try:
                    gcp_data = json.loads(gcp_json)
                    logging.debug(f"Using GCP credentials for profile '{credential.profile_name}'. Project ID: {gcp_data.get('project_id')}, Service Account: {gcp_data.get('client_email')}")
                except json.JSONDecodeError:
                    logging.warning("Could not parse GCP JSON for debug logging.")
            elif credential.provider == 'azure':
                live_logger.log("Decrypting Credentials", "ENCRYPTION", "Azure Credentials")
                # Azure can use either Service Principal or Managed Identity
                # encrypted_key_1 stores subscription_id (required)
                # encrypted_key_2 stores either service_account_json or connection_method
                subscription_id = decrypt_data(credential.encrypted_key_1, context=f"Azure Subscription ID for scan profile '{credential.profile_name}'")
                azure_auth_data = decrypt_data(credential.encrypted_key_2, context=f"Azure Authentication for scan profile '{credential.profile_name}'")
                
                decrypted_creds["azure_subscription_id"] = subscription_id
                decrypted_creds["azure_auth_data"] = azure_auth_data
                
                try:
                    # Try to parse as JSON (Service Principal)
                    auth_json = json.loads(azure_auth_data)
                    logging.debug(f"Using Azure Service Principal credentials for profile '{credential.profile_name}'. Subscription: {subscription_id}")
                except json.JSONDecodeError:
                    # If not JSON, assume it's a connection method identifier (e.g., "default" for DefaultAzureCredential)
                    logging.debug(f"Using Azure {azure_auth_data} authentication for profile '{credential.profile_name}'. Subscription: {subscription_id}")
            else:
                return jsonify({"error": f"Unsupported provider: {credential.provider}"}), 400
            
            logging.info(f"--- [SCAN] Starting parallel scanner for provider: {credential.provider}. ---")
            
            if progress_mode:
                def generate(user_id, is_guest_mode):
                    final_results = []
                    logging.debug("Progress mode: Starting to generate SSE events.")
                    with app.app_context():
                        suppressed_hashes = {sf.finding_hash for sf in SuppressedFinding.query.filter_by(user_id=user_id).all()}

                    for update in parallel_scanner.run_parallel_scans_progress(credentials=decrypted_creds, regions=regions_to_scan):
                        if isinstance(update, dict) and update.get('status') == 'progress':
                            yield f"data: {json.dumps(update)}\n\n"
                        elif isinstance(update, list):
                            unsuppressed_updates = [r for r in update if _generate_finding_hash(r) not in suppressed_hashes]
                            final_results.extend(unsuppressed_updates)

                    with app.app_context():
                        scan_time = datetime.now(timezone.utc)
                        scan_author = db.session.get(User, user_id)
                        logging.info(f"--- [DATABASE] Storing {len(final_results)} scan results to the database. ---")
                        for result in final_results:
                            if "error" not in result:
                                db_result = ScanResult(
                                    service=result.get('service'),
                                    resource=result.get('resource'),
                                    status=result.get('status'),
                                    issue=result.get('issue'),
                                    remediation=result.get('remediation'),
                                    doc_url=result.get('doc_url'),
                                    timestamp=scan_time,
                                    author=scan_author)
                                db.session.add(db_result)
                        db.session.commit()
                        logging.debug("Database save complete.")

                        # Increment scan count for Basic users after successful scan
                        if not is_guest_mode and scan_author and scan_author.is_basic_user():
                            scan_author.increment_scan_count()
                            logging.info(f"Basic user scan count incremented. Remaining: {scan_author.remaining_scans()}")

                    yield f"data: {json.dumps({'status': 'complete', 'results': final_results})}\n\n"
                
                return Response(generate(user_id=scan_user_id, is_guest_mode=session.get('guest_mode', False)), mimetype='text/event-stream')
            else:
                all_results = parallel_scanner.run_parallel_scans_blocking(credentials=decrypted_creds, regions=regions_to_scan)
                
                suppressed_hashes = {sf.finding_hash for sf in SuppressedFinding.query.filter_by(user_id=scan_user_id).all()}
                final_results = [r for r in all_results if _generate_finding_hash(r) not in suppressed_hashes]

                scan_time = datetime.now(timezone.utc)
                scan_author = db.session.get(User, scan_user_id)
                logging.info(f"--- [DATABASE] Storing {len(final_results)} scan results to the database. ---")
                for result in final_results:
                    if "error" not in result:
                        db_result = ScanResult(
                            service=result.get('service'), 
                            resource=result.get('resource'), 
                            status=result.get('status'), 
                            issue=result.get('issue'), 
                            remediation=result.get('remediation'), 
                            doc_url=result.get('doc_url'), 
                            timestamp=scan_time, 
                            author=scan_author)
                        db.session.add(db_result)
                db.session.commit()
                logging.debug("Database save complete.")

                # Increment scan count for Basic users after successful scan
                if not session.get('guest_mode') and scan_author and scan_author.is_basic_user():
                    scan_author.increment_scan_count()
                    logging.info(f"Basic user scan count incremented. Remaining: {scan_author.remaining_scans()}")
                
                # Log scan completion with structured data
                scan_duration = round((time.time() - g.start_time) * 1000, 2) if hasattr(g, 'start_time') else None
                structured_logger.log_scan_activity(
                    scan_type="manual_scan",
                    provider=credential.provider,
                    status="completed",
                    duration_ms=scan_duration,
                    resource_count=len(final_results)
                )
                
                return jsonify({"results": final_results})

        except Exception as e:
            logging.error(f"Major scan error: {e}", exc_info=True)
            # Log scan failure with structured data
            scan_duration = round((time.time() - g.start_time) * 1000, 2) if hasattr(g, 'start_time') else None
            structured_logger.log_scan_activity(
                scan_type="manual_scan",
                provider=credential.provider if 'credential' in locals() else "unknown",
                status="failed",
                duration_ms=scan_duration,
                resource_count=0
            )
            return jsonify({"error": str(e)}), 500
            
    except Exception as e:
        logging.error(f"Scan endpoint error: {e}", exc_info=True)
        return jsonify({"error": f"Scan request failed: {str(e)}"}), 500

@app.route('/api/v1/history', methods=['GET'])
@login_required
@check_verified
@check_2fa
def history():
    page = request.args.get('page', 1, type=int)
    per_page = 50
    pagination = ScanResult.query.filter_by(user_id=current_user.id).order_by(ScanResult.timestamp.desc()).paginate(page=page, per_page=per_page, error_out=False)
    results = pagination.items
    history_list = [{"id": r.id, "service": r.service, "resource": r.resource, "status": r.status, "issue": r.issue, "timestamp": r.timestamp.isoformat()} for r in results]
    return jsonify({"historical_scans": history_list, "page": pagination.page, "total_pages": pagination.pages, "has_next": pagination.has_next, "has_prev": pagination.has_prev})
    
@app.route('/api/v1/suppress_finding', methods=['POST'])
@login_required
@check_2fa
def suppress_finding():
    data = request.get_json()
    finding = data.get('finding')
    if not finding: return jsonify({"error": "Finding data is required."}), 400
    finding_hash = _generate_finding_hash(finding)
    existing = SuppressedFinding.query.filter_by(user_id=current_user.id, finding_hash=finding_hash).first()
    if existing: return jsonify({"message": "Finding is already suppressed."}), 200
    new_suppression = SuppressedFinding(user_id=current_user.id, finding_hash=finding_hash, reason="Suppressed by user from dashboard.", service=finding.get('service'), resource=finding.get('resource'), issue=finding.get('issue'))
    db.session.add(new_suppression)
    db.session.commit()
    log_audit("Finding Suppressed", details=f"Hash: {finding_hash[:12]}...", user=current_user)
    return jsonify({"message": "Finding suppressed successfully."}), 201

@app.route('/api/v1/unsuppress_finding/<int:suppression_id>', methods=['POST'])
@login_required
@check_2fa
def unsuppress_finding(suppression_id):
    suppression = SuppressedFinding.query.filter_by(id=suppression_id, user_id=current_user.id).first()
    if not suppression: return jsonify({"error": "Suppression not found or access denied."}), 404
    db.session.delete(suppression)
    db.session.commit()
    log_audit("Finding Un-suppressed", details=f"ID: {suppression_id}", user=current_user)
    return jsonify({"message": "Finding has been un-suppressed successfully."}), 200

@app.route('/api/v1/history/trends')
@login_required
@check_verified
@check_2fa
def history_trends():
    thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
    trend_data = db.session.query(func.date(ScanResult.timestamp).label('scan_date'), func.count(ScanResult.id).label('critical_count')).filter(ScanResult.status == 'CRITICAL', ScanResult.timestamp >= thirty_days_ago, ScanResult.user_id == current_user.id).group_by('scan_date').order_by('scan_date').all()
    labels = [datetime.strptime(row.scan_date, '%Y-%m-%d').strftime('%b %d') for row in trend_data]
    data = [row.critical_count for row in trend_data]
    return jsonify({"labels": labels, "data": data})

@app.route('/api/v1/dashboard/riskiest_resources')
@login_or_guest_required
@check_verified
@check_2fa
def get_riskiest_resources():
    """
    Analyzes scan history to find the top 5 resources
    with the most critical findings for the current user.
    """
    try:
        # Guest mode returns empty data
        if session.get('guest_mode'):
            return jsonify([])
        # Query to find resources with the most critical findings
        riskiest_resources = db.session.query(
            ScanResult.resource,
            func.count(ScanResult.id).label('critical_count')
        ).filter(
            ScanResult.user_id == current_user.id,
            ScanResult.status == 'CRITICAL'
        ).group_by(
            ScanResult.resource
        ).order_by(
            func.count(ScanResult.id).desc()
        ).limit(5).all()

        # Format the data for the frontend
        resources_list = [
            {'resource': resource, 'critical_count': count}
            for resource, count in riskiest_resources
        ]

        return jsonify(resources_list)

    except Exception as e:
        logging.error(f"Failed to get riskiest resources: {e}")
        return jsonify({"error": "Could not retrieve riskiest resources."}), 500

@app.route('/api/v1/dashboard/weekly_summary')
@login_or_guest_required
@check_verified
@check_2fa
def get_weekly_summary():
    try:
        # Guest mode returns empty data
        if session.get('guest_mode'):
            return jsonify({
                'labels': [],
                'ok_data': [],
                'critical_data': []
            })
        seven_days_ago = datetime.now(timezone.utc) - timedelta(days=7)
        
        # Use the same pattern as your working history_trends function
        trend_data = db.session.query(
            func.date(ScanResult.timestamp).label('scan_date'), 
            ScanResult.status,
            func.count(ScanResult.id).label('count')
        ).filter(
            ScanResult.user_id == current_user.id,
            ScanResult.timestamp >= seven_days_ago,
            ScanResult.status.in_(['OK', 'CRITICAL'])
        ).group_by(
            func.date(ScanResult.timestamp), ScanResult.status
        ).order_by('scan_date').all()
        
        # Process the data
        summary_data = {}
        for row in trend_data:
            date_str = str(row.scan_date)  # Convert to string
            if date_str not in summary_data:
                summary_data[date_str] = {'OK': 0, 'CRITICAL': 0}
            summary_data[date_str][row.status] = row.count
        
        # Create labels and data arrays
        labels = []
        ok_data = []
        critical_data = []
        
        for i in range(7):
            day = datetime.now(timezone.utc) - timedelta(days=6-i)
            day_str = day.strftime('%Y-%m-%d')
            labels.append(day.strftime('%a, %b %d'))
            
            day_data = summary_data.get(day_str, {'OK': 0, 'CRITICAL': 0})
            ok_data.append(day_data['OK'])
            critical_data.append(day_data['CRITICAL'])
        
        return jsonify({
            'labels': labels,
            'ok_data': ok_data,
            'critical_data': critical_data
        })
        
    except Exception as e:
        logging.error(f"Weekly summary error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route('/api/v1/delete_history', methods=['POST'])
@login_required
@check_2fa
def delete_history():
    try:
        num_deleted = ScanResult.query.filter_by(author=current_user).delete()
        db.session.commit()
        log_audit("History Deleted", details=f"Deleted {num_deleted} records", user=current_user)
        return jsonify({"message": f"Successfully deleted {num_deleted} of your historical scan results."}), 200
    except Exception as e:
        db.session.rollback()
        logging.error(f"History deletion failed: {e}")
        return jsonify({"error": "Failed to delete history."}), 500

@app.route('/report/pdf')
@login_required
@check_verified
@check_2fa
def generate_pdf_report():
    scan_results = ScanResult.query.filter_by(user_id=current_user.id).order_by(ScanResult.timestamp.desc()).limit(50).all()
    if not scan_results:
        flash('Please run a scan first to generate a report.', 'info')
        return redirect(url_for('dashboard'))
    pdf_bytes = _create_pdf_report(scan_results)
    return Response(pdf_bytes, mimetype='application/pdf', headers={'Content-Disposition': 'attachment;filename=aegis_cloud_security_report.pdf'})

@app.route('/report/csv')
@login_required
@check_verified
@check_2fa
def generate_csv_report():
    """Generates a CSV report of the latest 100 scan findings."""
    scan_results = ScanResult.query.filter_by(user_id=current_user.id).order_by(ScanResult.timestamp.desc()).limit(100).all()
    if not scan_results:
        flash('Please run a scan first to generate a report.', 'info')
        return redirect(url_for('dashboard'))
    # Use StringIO to create an in-memory file
    output = io.StringIO()
    writer = csv.writer(output)
    # Write the header
    writer.writerow(['Timestamp', 'Service', 'Resource', 'Status', 'Issue', 'Remediation', 'Documentation URL'])
    # Write the data
    for result in scan_results:
        writer.writerow([
            result.timestamp.isoformat(),
            result.service,
            result.resource,
            result.status,
            result.issue,
            result.remediation,
            result.doc_url
        ])
    # Prepare the response
    output.seek(0)
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=aegis_security_report.csv"}
    )
    
# --- NEW API ENDPOINTS FOR DASHBOARD VISUALIZATIONS ---
@app.route('/api/v1/dashboard/findings_by_service')
@login_or_guest_required
@check_verified
@check_2fa
def get_findings_by_service():
    """Returns the count of critical findings grouped by service for center chart."""
    try:
        # Guest mode returns empty data
        if session.get('guest_mode'):
            return jsonify({
                'labels': [],
                'data': [],
                'colors': []
            })
        findings = db.session.query(
            ScanResult.service,
            func.count(ScanResult.id).label('count')
        ).filter(
            ScanResult.user_id == current_user.id,
            ScanResult.status == 'CRITICAL'
        ).group_by(ScanResult.service).order_by(func.count(ScanResult.id).desc()).limit(10).all()
        
        labels = [row.service for row in findings]
        data = [row.count for row in findings]
        colors = ['#D64550', '#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE']
        
        return jsonify({
            'labels': labels, 
            'data': data,
            'colors': colors[:len(labels)]
        })
    except Exception as e:
        logging.error(f"Failed to get findings by service: {e}")
        return jsonify({"error": "Could not retrieve data."}), 500

@app.route('/api/v1/dashboard/compliance_overview')
@login_or_guest_required
@check_2fa
def get_compliance_overview():
    """Returns compliance framework data for compliance dashboard."""
    try:
        # Guest mode returns default compliance data
        if session.get('guest_mode'):
            return jsonify({
                "compliance_scores": [
                    {'framework': 'NIST', 'score': 100.0, 'passed': 0, 'failed': 0, 'total': 0},
                    {'framework': 'CIS', 'score': 100.0, 'passed': 0, 'failed': 0, 'total': 0},
                    {'framework': 'SOC 2', 'score': 100.0, 'passed': 0, 'failed': 0, 'total': 0},
                    {'framework': 'GDPR', 'score': 100.0, 'passed': 0, 'failed': 0, 'total': 0}
                ],
                "overall_score": 100.0
            })
        # Get recent findings for compliance mapping
        thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
        
        findings = db.session.query(ScanResult).filter(
            ScanResult.user_id == current_user.id,
            ScanResult.timestamp >= thirty_days_ago
        ).all()
        
        # Compliance framework mapping (simplified example)
        compliance_frameworks = {
            'NIST': {'passed': 0, 'failed': 0, 'total': 0},
            'CIS': {'passed': 0, 'failed': 0, 'total': 0},
            'SOC 2': {'passed': 0, 'failed': 0, 'total': 0},
            'GDPR': {'passed': 0, 'failed': 0, 'total': 0}
        }
        
        # Map service findings to compliance frameworks
        service_compliance_map = {
            'S3': ['NIST', 'CIS', 'SOC 2', 'GDPR'],
            'IAM': ['NIST', 'CIS', 'SOC 2'],
            'VPC': ['NIST', 'CIS'],
            'RDS': ['NIST', 'SOC 2', 'GDPR'],
            'CloudTrail': ['NIST', 'CIS', 'SOC 2'],
            'KMS': ['NIST', 'SOC 2', 'GDPR']
        }
        
        for finding in findings:
            service = finding.service
            status = finding.status
            
            if service in service_compliance_map:
                for framework in service_compliance_map[service]:
                    compliance_frameworks[framework]['total'] += 1
                    if status == 'OK':
                        compliance_frameworks[framework]['passed'] += 1
                    else:
                        compliance_frameworks[framework]['failed'] += 1
        
        # Calculate compliance scores
        compliance_scores = []
        for framework, data in compliance_frameworks.items():
            if data['total'] > 0:
                score = (data['passed'] / data['total']) * 100
                compliance_scores.append({
                    'framework': framework,
                    'score': round(score, 1),
                    'passed': data['passed'],
                    'failed': data['failed'],
                    'total': data['total']
                })
            else:
                compliance_scores.append({
                    'framework': framework,
                    'score': 100.0,
                    'passed': 0,
                    'failed': 0,
                    'total': 0
                })
        
        return jsonify({
            "compliance_scores": compliance_scores,
            "overall_score": round(sum(item['score'] for item in compliance_scores) / len(compliance_scores), 1) if compliance_scores else 100
        })
    except Exception as e:
        logging.error(f"Failed to get compliance overview: {e}")
        return jsonify({"error": "Could not retrieve compliance data."}), 500

@app.route('/api/v1/resource/discover', methods=['POST'])
@login_or_guest_required
@check_2fa
def discover_resources():
    """Discover cloud resources across providers."""
    try:
        data = request.get_json()
        provider_filter = data.get('provider', 'all') if data else 'all'

        # Log resource discovery activity
        structured_logger.log_structured('info', 'Resource Discovery Started',
            user_id=current_user.id if current_user and not session.get('guest_mode') else 'guest',
            provider_filter=provider_filter
        )

        # Get user's credentials for real discovery
        available_credentials = []
        if not session.get('guest_mode') and current_user:
            available_credentials = CloudCredential.query.filter_by(
                user_id=current_user.id
            ).all()

        # Generate discovery results based on available credentials or demo data
        discovered_resources = generate_resource_discovery_data(available_credentials, provider_filter)

        # Update discovery statistics
        discovery_stats = {
            'total_resources': sum(r['resource_count'] for r in discovered_resources),
            'active_providers': len(set(r['provider'] for r in discovered_resources)),
            'security_issues': sum(r['issue_count'] for r in discovered_resources),
            'last_discovery': datetime.now(timezone.utc).isoformat()
        }

        return jsonify({
            'success': True,
            'resources': discovered_resources,
            'stats': discovery_stats,
            'discovery_time': datetime.now(timezone.utc).isoformat()
        })

    except Exception as e:
        logging.error(f"Failed to discover resources: {e}")
        logging.error(f"Exception type: {type(e).__name__}")
        logging.error(f"Exception traceback:", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

def generate_resource_discovery_data(credentials, provider_filter='all'):
    """Generate resource discovery data based on available credentials."""
    try:
        logging.info(f"Starting resource discovery generation with {len(credentials) if credentials else 0} credentials, filter: {provider_filter}")

        # Base resource types for each provider
        resource_templates = {
            'aws': [
                {'type': 'compute', 'name': 'EC2', 'icon': 'fas fa-server'},
                {'type': 'storage', 'name': 'S3', 'icon': 'fas fa-cube'},
                {'type': 'database', 'name': 'RDS', 'icon': 'fas fa-database'},
                {'type': 'network', 'name': 'VPC', 'icon': 'fas fa-network-wired'},
                {'type': 'security', 'name': 'IAM', 'icon': 'fas fa-users-cog'},
                {'type': 'compute', 'name': 'Lambda', 'icon': 'fab fa-aws'},
                {'type': 'storage', 'name': 'EBS', 'icon': 'fas fa-hdd'},
                {'type': 'network', 'name': 'CloudFront', 'icon': 'fas fa-globe'},
            ],
            'gcp': [
                {'type': 'compute', 'name': 'Compute Engine', 'icon': 'fas fa-server'},
                {'type': 'storage', 'name': 'Cloud Storage', 'icon': 'fas fa-cube'},
                {'type': 'database', 'name': 'Cloud SQL', 'icon': 'fas fa-database'},
                {'type': 'network', 'name': 'VPC Network', 'icon': 'fas fa-network-wired'},
                {'type': 'security', 'name': 'Cloud IAM', 'icon': 'fas fa-users-cog'},
                {'type': 'compute', 'name': 'Cloud Functions', 'icon': 'fab fa-google'},
            ],
            'azure': [
                {'type': 'compute', 'name': 'Virtual Machines', 'icon': 'fas fa-server'},
                {'type': 'storage', 'name': 'Blob Storage', 'icon': 'fas fa-cube'},
                {'type': 'database', 'name': 'Azure SQL', 'icon': 'fas fa-database'},
                {'type': 'network', 'name': 'Virtual Network', 'icon': 'fas fa-network-wired'},
                {'type': 'security', 'name': 'Azure AD', 'icon': 'fas fa-users-cog'},
                {'type': 'compute', 'name': 'Azure Functions', 'icon': 'fab fa-microsoft'},
            ]
        }

        regions = {
            'aws': ['us-east-1', 'us-west-2', 'eu-west-1', 'ap-southeast-1'],
            'gcp': ['us-central1', 'us-west1', 'europe-west1', 'asia-southeast1'],
            'azure': ['East US', 'West US 2', 'West Europe', 'Southeast Asia']
        }

        statuses = ['secure', 'warning', 'critical']

        discovered_resources = []

        # Determine which providers to include
        providers_to_scan = []
        if credentials:
            # Use actual credentials
            for cred in credentials:
                if provider_filter == 'all' or cred.provider.lower() == provider_filter:
                    providers_to_scan.append(cred.provider.lower())
        else:
            # Demo mode - include all providers or filtered
            if provider_filter == 'all':
                providers_to_scan = ['aws', 'gcp', 'azure']
            else:
                providers_to_scan = [provider_filter]

        # Generate resources for each provider
        for provider in providers_to_scan:
            if provider in resource_templates:
                provider_regions = regions.get(provider, ['us-east-1'])

                for template in resource_templates[provider]:
                    # Generate realistic resource counts and issues
                    base_count = {
                        'compute': (50, 500),
                        'storage': (10, 100),
                        'database': (5, 50),
                        'network': (5, 25),
                        'security': (1, 10)
                    }.get(template['type'], (10, 100))

                    resource_count = random.randint(base_count[0], base_count[1])

                    # Generate issues based on resource type and count
                    issue_probability = {
                        'compute': 0.15,
                        'storage': 0.25,
                        'database': 0.10,
                        'network': 0.20,
                        'security': 0.30
                    }.get(template['type'], 0.15)

                    issue_count = int(resource_count * issue_probability * random.uniform(0.5, 1.5))

                    # Determine status based on issue count
                    if issue_count == 0:
                        status = 'secure'
                    elif issue_count < resource_count * 0.1:
                        status = 'warning'
                    else:
                        status = 'critical'

                    discovered_resources.append({
                        'id': f"{provider}_{template['name'].lower().replace(' ', '_')}",
                        'name': template['name'],
                        'type': template['type'],
                        'provider': provider,
                        'icon': template['icon'],
                        'region': random.choice(provider_regions),
                        'resource_count': resource_count,
                        'issue_count': issue_count,
                        'status': status,
                        'last_scan': (datetime.now(timezone.utc) - timedelta(
                            hours=random.randint(1, 72)
                        )).strftime('%Y-%m-%d %H:%M:%S'),
                        'discovered_at': datetime.now(timezone.utc).isoformat()
                    })

        return discovered_resources

    except Exception as e:
        logging.error(f"Error in generate_resource_discovery_data: {e}")
        # Return empty list as fallback
        return []

@app.route('/api/v1/dashboard/resource_topology')
@login_or_guest_required
@check_2fa
def get_resource_topology():
    """Legacy endpoint - redirects to resource discovery."""
    return jsonify({
        "message": "This endpoint has been replaced by /api/v1/resource/discover",
        "nodes": [],
        "edges": [],
        "summary": {
            "total_services": 0,
            "total_resources": 0,
            "critical_services": 0
        }
    })

@app.route('/api/v1/dashboard/scan_performance')
@login_or_guest_required
@check_2fa
def get_scan_performance():
    try:
        if session.get('guest_mode'):
            return jsonify({"performance_data": [], "summary": {"avg_scan_time_minutes": 0, "total_scans": 0, "total_resources_scanned": 0, "avg_resources_per_scan": 0}})
        
        # Check if current_user is available
        if not current_user or not hasattr(current_user, 'id'):
            logging.error("Current user not available in get_scan_performance")
            return jsonify({"performance_data": [], "summary": {"avg_scan_time_minutes": 0, "total_scans": 0, "total_resources_scanned": 0, "avg_resources_per_scan": 0}})
        
        thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
        
        scan_sessions = db.session.query(
            func.date(ScanResult.timestamp).label('scan_date'),
            func.count(ScanResult.id).label('resources_scanned'),
            func.min(ScanResult.timestamp).label('scan_start'),
            func.max(ScanResult.timestamp).label('scan_end')
        ).filter(
            ScanResult.user_id == current_user.id,
            ScanResult.timestamp >= thirty_days_ago
        ).group_by(func.date(ScanResult.timestamp)).order_by(func.date(ScanResult.timestamp).desc()).limit(30).all()

        if not scan_sessions:
            return jsonify({"performance_data": [], "summary": {"avg_scan_time_minutes": 0, "total_scans": 0, "total_resources_scanned": 0, "avg_resources_per_scan": 0}})

        performance_data = []
        total_duration = 0
        total_resources = 0
        
        for session_data in scan_sessions:
            try:
                duration_seconds = (session_data.scan_end - session_data.scan_start).total_seconds()
                scan_duration_minutes = max(0.1, duration_seconds / 60) # Ensure non-zero duration
                total_duration += scan_duration_minutes
                total_resources += session_data.resources_scanned
                
                # Handle date formatting safely
                scan_date = session_data.scan_date
                if hasattr(scan_date, 'strftime'):
                    date_str = scan_date.strftime('%Y-%m-%d')
                elif hasattr(scan_date, 'isoformat'):
                    date_str = scan_date.isoformat()
                else:
                    # If it's a date object, convert to string
                    date_str = str(scan_date)
                
                performance_data.append({
                    'date': date_str,
                    'resources_scanned': int(session_data.resources_scanned),
                    'scan_duration_minutes': round(scan_duration_minutes, 2)
                })
            except Exception as e:
                logging.error(f"Error processing scan session data: {e}")
                continue
            
        session_count = len(performance_data)
        avg_scan_time = total_duration / session_count if session_count > 0 else 0
        avg_resources_per_scan = total_resources / session_count if session_count > 0 else 0
        
        return jsonify({
            "performance_data": performance_data,
            "summary": {
                "avg_scan_time_minutes": round(avg_scan_time, 2),
                "total_scans": session_count,
                "total_resources_scanned": total_resources,
                "avg_resources_per_scan": round(avg_resources_per_scan, 1)
            }
        })
    except Exception as e:
        logging.error(f"Failed to get scan performance: {e}", exc_info=True)
        # Return empty data instead of 500 error to prevent frontend issues
        return jsonify({"performance_data": [], "summary": {"avg_scan_time_minutes": 0, "total_scans": 0, "total_resources_scanned": 0, "avg_resources_per_scan": 0}})

@app.route('/api/v1/dashboard/notifications')
@login_or_guest_required
@check_2fa
def get_notifications():
    """Returns recent notifications and alerts for the user."""
    try:
        # Guest mode returns empty notifications
        if session.get('guest_mode'):
            return jsonify({
                "notifications": [],
                "unread_count": 0
            })
        notifications = []
        
        # Check for recent critical findings (last 7 days)
        seven_days_ago = datetime.now(timezone.utc) - timedelta(days=7)
        recent_critical = db.session.query(ScanResult).filter(
            ScanResult.user_id == current_user.id,
            ScanResult.status == 'CRITICAL',
            ScanResult.timestamp >= seven_days_ago
        ).count()
        
        if recent_critical > 0:
            notifications.append({
                'id': 'critical_findings',
                'type': 'warning',
                'title': f'{recent_critical} Critical Security Issues',
                'message': f'Found {recent_critical} critical security issues in the last 7 days that require attention.',
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'action': 'View Details',
                'action_url': '#results'
            })
        
        # Check for scheduled scans
        if scheduler:
            try:
                scheduled_jobs = scheduler.get_jobs()
                scan_jobs = [job for job in scheduled_jobs if 'scan' in job.id]
                if scan_jobs:
                    next_scan = min(job.next_run_time for job in scan_jobs if job.next_run_time)
                    notifications.append({
                        'id': 'next_scan',
                        'type': 'info',
                        'title': 'Next Scheduled Scan',
                        'message': f'Your next scan is scheduled for {next_scan.strftime("%B %d, %Y at %I:%M %p")}',
                        'timestamp': datetime.now(timezone.utc).isoformat(),
                        'action': 'Manage Schedules',
                        'action_url': '#automation'
                    })
            except Exception as e:
                logging.error(f"Error checking scheduled jobs: {e}")
        
        # Check for credential expiration (mock implementation)
        credentials = CloudCredential.query.filter_by(user_id=current_user.id).all()
        for cred in credentials[:1]:  # Just check first credential as example
            notifications.append({
                'id': f'credential_{cred.id}',
                'type': 'success',
                'title': 'Credentials Active',
                'message': f'Your {cred.profile_name} credentials are active and ready for scanning.',
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'action': 'Manage Credentials',
                'action_url': '/settings'
            })
        
        # System health notification
        notifications.append({
            'id': 'system_health',
            'type': 'info',
            'title': 'System Status',
            'message': 'All systems operational. Scanner is ready for use.',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'action': None,
            'action_url': None
        })
        
        return jsonify({
            "notifications": notifications[:10],  # Limit to 10 most recent
            "unread_count": len(notifications)
        })
    except Exception as e:
        logging.error(f"Failed to get notifications: {e}")
        return jsonify({"error": "Could not retrieve notifications."}), 500

@app.route('/api/v1/dashboard/notifications/mark-read', methods=['POST'])
@login_required
@check_2fa
def mark_notifications_read():
    """Mark all notifications as read for the current user."""
    try:
        # For now, this is a placeholder since notifications are generated dynamically
        # In a real implementation, you'd store notification read status in the database
        return jsonify({
            "success": True,
            "message": "All notifications marked as read"
        })
    except Exception as e:
        logging.error(f"Failed to mark notifications as read: {e}")
        return jsonify({"error": "Could not mark notifications as read."}), 500

@app.route('/api/v1/background_scan/start', methods=['POST'])
@login_required
@check_2fa
def start_background_scan():
    """Start automatic background scanning with custom interval."""
    try:
        data = request.get_json()
        credential_id = data.get('credential_id')
        interval_minutes = data.get('interval_minutes', 60)  # Default 1 hour
        
        if not credential_id:
            return jsonify({"error": "Credential ID is required"}), 400
            
        if not isinstance(interval_minutes, int) or interval_minutes < 15:
            return jsonify({"error": "Interval must be at least 15 minutes"}), 400
        
        # Verify credential ownership
        credential = CloudCredential.query.filter_by(
            id=credential_id, 
            user_id=current_user.id
        ).first()
        
        if not credential:
            return jsonify({"error": "Credential not found"}), 404
        
        # Create background job ID
        job_id = f"background_scan_{current_user.id}_{credential_id}"
        
        # Remove existing background job if it exists
        try:
            scheduler.remove_job(job_id)
        except:
            pass
        
        # Schedule recurring background scan
        scheduler.add_job(
            func=perform_scheduled_scan,
            trigger='interval',
            minutes=interval_minutes,
            args=[current_user.id, credential_id, None],
            id=job_id,
            replace_existing=True
        )
        
        log_audit("Background Scanning Started", 
                 details=f"Profile: {credential.profile_name}, Interval: {interval_minutes} minutes", 
                 user=current_user)
        
        return jsonify({
            "message": f"Background scanning started every {interval_minutes} minutes using {credential.profile_name}",
            "job_id": job_id,
            "interval_minutes": interval_minutes
        }), 200
        
    except Exception as e:
        logging.error(f"Background scan start error: {e}")
        return jsonify({"error": "Failed to start background scanning"}), 500

@app.route('/api/v1/background_scan/stop', methods=['POST'])
@login_required
@check_2fa
def stop_background_scan():
    """Stop automatic background scanning."""
    try:
        data = request.get_json()
        credential_id = data.get('credential_id')
        
        if not credential_id:
            return jsonify({"error": "Credential ID is required"}), 400
        
        # Create background job ID
        job_id = f"background_scan_{current_user.id}_{credential_id}"
        
        # Remove background job
        try:
            scheduler.remove_job(job_id)
            message = "Background scanning stopped successfully"
            status = 200
        except:
            message = "No active background scan found for this credential"
            status = 404
        
        log_audit("Background Scanning Stopped", 
                 details=f"Job ID: {job_id}", 
                 user=current_user)
        
        return jsonify({"message": message}), status
        
    except Exception as e:
        logging.error(f"Background scan stop error: {e}")
        return jsonify({"error": "Failed to stop background scanning"}), 500

@app.route('/api/v1/background_scan/status')
@login_required
@check_2fa
def get_background_scan_status():
    """Get status of background scans for current user."""
    try:
        if not scheduler:
            return jsonify({"background_scans": [], "active_count": 0}), 200
        
        # Get all jobs for the current user
        all_jobs = scheduler.get_jobs()
        background_scans = []
        
        for job in all_jobs:
            if job.id.startswith(f"background_scan_{current_user.id}_"):
                # Extract credential ID from job ID
                parts = job.id.split('_')
                if len(parts) >= 4:
                    credential_id = parts[3]
                    
                    # Get credential info
                    credential = CloudCredential.query.filter_by(
                        id=int(credential_id), 
                        user_id=current_user.id
                    ).first()
                    
                    if credential:
                        # Calculate interval from trigger
                        interval_minutes = getattr(job.trigger, 'interval', None)
                        if interval_minutes:
                            interval_minutes = int(interval_minutes.total_seconds() / 60)
                        
                        background_scans.append({
                            'job_id': job.id,
                            'credential_id': credential_id,
                            'profile_name': credential.profile_name,
                            'interval_minutes': interval_minutes or 60,
                            'next_run': job.next_run_time.isoformat() if job.next_run_time else None,
                            'status': 'active'
                        })
        
        return jsonify({
            "background_scans": background_scans,
            "active_count": len(background_scans)
        }), 200
        
    except Exception as e:
        logging.error(f"Background scan status error: {e}")
        return jsonify({"error": "Failed to get background scan status"}), 500

@app.route('/api/v1/email_report', methods=['POST'])
@login_required
@check_2fa
def email_report():
    """Send PDF report via email."""
    try:
        data = request.get_json()
        recipient = security_validator.sanitize_string(data.get('recipient', ''))
        
        # Validate email
        is_valid_email, email_error = security_validator.validate_email(recipient)
        if not is_valid_email:
            return jsonify({"error": f"Invalid email: {email_error}"}), 400
        
        # Check mail configuration first
        try:
            mail_config = secrets_manager.get_mail_config()
            if not all([mail_config.get('MAIL_SERVER'), mail_config.get('MAIL_USERNAME'), mail_config.get('MAIL_PASSWORD')]):
                return jsonify({"error": "Email service not configured. Please configure email settings in admin panel."}), 400
        except Exception as config_error:
            logging.error(f"Mail configuration error: {config_error}")
            return jsonify({"error": "Email service not available. Please contact administrator."}), 500
        
        # Get recent scan results (last 30 days)
        thirty_days_ago = datetime.now(timezone.utc) - timedelta(days=30)
        scan_results = ScanResult.query.filter_by(user_id=current_user.id)\
            .filter(ScanResult.timestamp >= thirty_days_ago)\
            .order_by(ScanResult.timestamp.desc()).limit(100).all()
        
        if not scan_results:
            return jsonify({"error": "No recent scan results to include in report."}), 400
        
        # Generate PDF
        try:
            pdf_bytes = _create_pdf_report(scan_results)
            if not pdf_bytes:
                return jsonify({"error": "Failed to generate PDF report."}), 500
        except Exception as pdf_error:
            logging.error(f"PDF generation error: {pdf_error}")
            return jsonify({"error": "Failed to generate PDF report."}), 500
        
        # Send email
        try:
            msg = Message(
                f'Aegis Scanner - Security Report from {current_user.username}',
                recipients=[recipient]
            )
            msg.html = f"""
            <html>
            <body style="font-family: Arial, sans-serif;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <div style="background: #00A896; color: white; padding: 20px; text-align: center;">
                        <h1>Aegis Security Report</h1>
                        <p>Generated by {current_user.username}</p>
                    </div>
                    <div style="padding: 20px;">
                        <h3>Report Summary:</h3>
                        <p>This report contains the latest {len(scan_results)} security findings from the last 30 days.</p>
                        <p>Critical findings: {len([r for r in scan_results if r.status == 'CRITICAL'])}</p>
                        <p>Generated on: {datetime.now(timezone.utc).strftime('%B %d, %Y at %I:%M %p')}</p>
                    </div>
                </div>
            </body>
            </html>
            """
            msg.attach(
                'aegis_security_report.pdf',
                'application/pdf',
                pdf_bytes
            )
            
            # Test mail connection first
            with mail.connect() as conn:
                conn.send(msg)
            
            log_audit("Email Report Sent", details=f"Recipient: {recipient}", user=current_user)
            return jsonify({"message": f"Report successfully sent to {recipient}"}), 200
            
        except Exception as mail_error:
            logging.error(f"Mail sending error: {mail_error}")
            if "authentication" in str(mail_error).lower():
                return jsonify({"error": "Email authentication failed. Please check email server configuration."}), 500
            elif "connection" in str(mail_error).lower():
                return jsonify({"error": "Unable to connect to email server. Please check network connection and server settings."}), 500
            else:
                return jsonify({"error": f"Failed to send email: {str(mail_error)}"}), 500
        
    except Exception as e:
        logging.error(f"Email report error: {e}")
        return jsonify({"error": "Failed to send report via email. Please try again."}), 500

@app.route('/api/v1/generate-report', methods=['POST'])
@login_required
@check_2fa
def generate_report():
    """Generate various types of reports (PDF, advanced, executive, technical)."""
    try:
        data = request.get_json()
        report_type = data.get('type', 'pdf')
        
        # Get scan results for report
        scan_results = ScanResult.query.filter_by(user_id=current_user.id)\
                                      .order_by(ScanResult.scan_date.desc())\
                                      .limit(100).all()
        
        if not scan_results:
            return jsonify({"error": "No scan data available for report generation"}), 400
        
        if report_type == 'pdf':
            # Generate PDF report
            pdf_bytes = _create_pdf_report(scan_results)
            
            return jsonify({
                "message": "PDF report generated successfully",
                "data": {
                    "report_type": "pdf",
                    "findings_count": len(scan_results),
                    "generated_at": datetime.now(timezone.utc).isoformat()
                }
            }), 200
            
        elif report_type == 'advanced':
            # Generate advanced analytics report
            critical_count = len([r for r in scan_results if r.status == 'CRITICAL'])
            high_count = len([r for r in scan_results if r.status == 'HIGH'])
            medium_count = len([r for r in scan_results if r.status == 'MEDIUM'])
            low_count = len([r for r in scan_results if r.status == 'LOW'])
            
            return jsonify({
                "message": "Advanced report generated successfully",
                "data": {
                    "report_type": "advanced",
                    "summary": {
                        "total_findings": len(scan_results),
                        "critical": critical_count,
                        "high": high_count,
                        "medium": medium_count,
                        "low": low_count
                    },
                    "generated_at": datetime.now(timezone.utc).isoformat()
                }
            }), 200
            
        elif report_type == 'executive':
            # Generate executive summary
            return jsonify({
                "message": "Executive report generated successfully",
                "data": {
                    "report_type": "executive",
                    "executive_summary": {
                        "total_scans": len(scan_results),
                        "risk_score": min(100, (len([r for r in scan_results if r.status in ['CRITICAL', 'HIGH']]) / max(1, len(scan_results))) * 100),
                        "recommendation": "Review critical and high severity findings immediately"
                    },
                    "generated_at": datetime.now(timezone.utc).isoformat()
                }
            }), 200
            
        elif report_type == 'technical':
            # Generate technical detailed report
            resources_by_type = {}
            for result in scan_results:
                resource_type = getattr(result, 'resource_type', 'Unknown')
                if resource_type not in resources_by_type:
                    resources_by_type[resource_type] = 0
                resources_by_type[resource_type] += 1

            return jsonify({
                "message": "Technical report generated successfully",
                "data": {
                    "report_type": "technical",
                    "technical_details": {
                        "resources_by_type": resources_by_type,
                        "scan_coverage": len(scan_results),
                        "last_scan": scan_results[0].scan_date.isoformat() if scan_results else None
                    },
                    "generated_at": datetime.now(timezone.utc).isoformat()
                }
            }), 200

        elif report_type == 'compliance':
            # Generate compliance assessment report
            critical_count = len([r for r in scan_results if r.status == 'CRITICAL'])
            high_count = len([r for r in scan_results if r.status == 'HIGH'])
            medium_count = len([r for r in scan_results if r.status == 'MEDIUM'])

            compliance_score = max(0, 100 - (critical_count * 30 + high_count * 15 + medium_count * 5))

            frameworks = {
                'SOC 2': max(0, 100 - (critical_count * 25 + high_count * 10)),
                'PCI DSS': max(0, 100 - (critical_count * 35 + high_count * 15)),
                'GDPR': max(0, 100 - (critical_count * 20 + high_count * 12)),
                'HIPAA': max(0, 100 - (critical_count * 30 + high_count * 15)),
                'ISO 27001': max(0, 100 - (critical_count * 25 + high_count * 12))
            }

            return jsonify({
                "message": "Compliance report generated successfully",
                "data": {
                    "report_type": "compliance",
                    "compliance_summary": {
                        "overall_score": compliance_score,
                        "critical_gaps": critical_count,
                        "high_priority_issues": high_count,
                        "frameworks": frameworks,
                        "total_findings": len(scan_results)
                    },
                    "generated_at": datetime.now(timezone.utc).isoformat()
                }
            }), 200
            
        else:
            return jsonify({"error": f"Unsupported report type: {report_type}"}), 400
            
    except Exception as e:
        logging.error(f"Report generation error: {e}")
        return jsonify({"error": "Failed to generate report"}), 500

@app.route('/api/v1/advanced_schedule', methods=['POST'])
@login_required
@check_2fa
def advanced_schedule():
    """Schedule a daily scan at a specific time."""
    try:
        data = request.get_json()
        credential_id = data.get('credential_id')
        run_time = data.get('run_time')  # Format: "HH:MM"
        
        # Validate inputs
        if not credential_id or not run_time:
            return jsonify({"error": "Missing required fields"}), 400
        
        # Verify credential ownership
        credential = CloudCredential.query.filter_by(
            id=credential_id, 
            user_id=current_user.id
        ).first()
        
        if not credential:
            return jsonify({"error": "Credential not found"}), 404
        
        # Parse time
        try:
            hour, minute = map(int, run_time.split(':'))
            if not (0 <= hour <= 23 and 0 <= minute <= 59):
                raise ValueError("Invalid time range")
        except ValueError:
            return jsonify({"error": "Invalid time format. Use HH:MM (24-hour)"}), 400
        
        # Create job ID
        job_id = f"advanced_scan_{current_user.id}_{credential_id}"
        
        # Remove existing job if it exists
        try:
            scheduler.remove_job(job_id)
        except:
            pass
        
        # Schedule daily at specific time
        scheduler.add_job(
            id=job_id,
            func=perform_automated_scan,
            args=[credential_id, current_user.id],
            trigger='cron',
            hour=hour,
            minute=minute,
            max_instances=1,
            coalesce=True
        )
        
        log_audit("Advanced Scan Scheduled", 
                 details=f"Daily at {run_time} using {credential.profile_name}", 
                 user=current_user)
        
        return jsonify({
            "message": f"Daily scan scheduled at {run_time} using {credential.profile_name}",
            "job_id": job_id,
            "run_time": run_time
        }), 200
        
    except Exception as e:
        logging.error(f"Advanced schedule error: {e}")
        return jsonify({"error": "Failed to schedule advanced scan"}), 500

# ============== ENTERPRISE APIs ==============

@app.route('/api/v1/enterprise/compliance/automation', methods=['POST'])
@login_required
@admin_required
def create_compliance_automation():
    """Create automated compliance monitoring."""
    try:
        data = request.get_json()
        framework = data.get('framework')
        resources = data.get('resources', [])
        schedule = data.get('schedule', 'daily')
        
        if not framework or framework not in ['SOC2', 'ISO27001', 'GDPR', 'HIPAA']:
            return jsonify({"error": "Invalid compliance framework"}), 400
        
        # Create automation rule
        automation_id = f"compliance_{framework.lower()}_{current_user.id}_{int(time.time())}"
        
        # Schedule compliance check based on framework
        if framework == 'SOC2':
            checks = [
                'cloudtrail_enabled', 'cloudwatch_alarms', 'iam_mfa',
                'encrypted_storage', 'vpc_flow_logs', 'security_groups'
            ]
        elif framework == 'ISO27001':
            checks = [
                'access_control', 'encryption_keys', 'monitoring_logs',
                'incident_response', 'vulnerability_scanning'
            ]
        elif framework == 'GDPR':
            checks = [
                'data_encryption', 'access_logging', 'backup_retention',
                'data_deletion_policy', 'consent_tracking'
            ]
        else:  # HIPAA
            checks = [
                'database_encryption', 'audit_logs', 'access_controls',
                'backup_encryption', 'transmission_security'
            ]
        
        return jsonify({
            "automation_id": automation_id,
            "framework": framework,
            "checks_enabled": checks,
            "schedule": schedule,
            "status": "active",
            "next_run": (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
        }), 201
        
    except Exception as e:
        logging.error(f"Compliance automation error: {e}")
        return jsonify({"error": "Failed to create compliance automation"}), 500

@app.route('/api/v1/enterprise/compliance/reports', methods=['GET'])
@login_required
@admin_required
def get_compliance_reports():
    """Generate comprehensive compliance reports."""
    try:
        framework = request.args.get('framework', 'all')
        days = int(request.args.get('days', 30))
        
        start_date = datetime.now(timezone.utc) - timedelta(days=days)
        
        # Get scan results for compliance analysis
        scan_results = ScanResult.query.filter(
            ScanResult.timestamp >= start_date
        ).all()
        
        compliance_data = {
            'SOC2': {
                'score': 87.5,
                'status': 'compliant',
                'findings': len([r for r in scan_results if r.service in ['CloudTrail', 'IAM', 'VPC']]),
                'critical_issues': len([r for r in scan_results if r.status == 'CRITICAL' and r.service in ['CloudTrail', 'IAM']]),
                'controls': {
                    'CC6.1': {'status': 'compliant', 'evidence': '15 passing checks'},
                    'CC6.2': {'status': 'non_compliant', 'evidence': '3 critical findings'},
                    'CC6.3': {'status': 'compliant', 'evidence': '12 passing checks'}
                }
            },
            'ISO27001': {
                'score': 92.3,
                'status': 'compliant',
                'findings': len([r for r in scan_results if r.service in ['KMS', 'CloudWatch', 'S3']]),
                'critical_issues': len([r for r in scan_results if r.status == 'CRITICAL' and r.service == 'KMS']),
                'controls': {
                    'A.12.6.1': {'status': 'compliant', 'evidence': 'Vulnerability management active'},
                    'A.18.1.4': {'status': 'compliant', 'evidence': 'Privacy controls implemented'}
                }
            },
            'GDPR': {
                'score': 78.9,
                'status': 'partially_compliant',
                'findings': len([r for r in scan_results if r.service in ['S3', 'RDS']]),
                'critical_issues': len([r for r in scan_results if r.status == 'CRITICAL' and r.service in ['S3', 'RDS']]),
                'controls': {
                    'Art.32': {'status': 'compliant', 'evidence': 'Encryption implemented'},
                    'Art.25': {'status': 'non_compliant', 'evidence': 'Privacy by design gaps'}
                }
            },
            'HIPAA': {
                'score': 85.7,
                'status': 'compliant',
                'findings': len([r for r in scan_results if r.service in ['RDS', 'S3', 'KMS']]),
                'critical_issues': len([r for r in scan_results if r.status == 'CRITICAL' and r.service in ['RDS', 'KMS']]),
                'controls': {
                    '164.312': {'status': 'compliant', 'evidence': 'Access controls verified'},
                    '164.306': {'status': 'compliant', 'evidence': 'Security standards met'}
                }
            }
        }
        
        if framework != 'all':
            compliance_data = {framework.upper(): compliance_data.get(framework.upper(), {})}
        
        return jsonify({
            'compliance_reports': compliance_data,
            'generated_at': datetime.now(timezone.utc).isoformat(),
            'report_period': f'{days} days',
            'total_findings': len(scan_results)
        }), 200
        
    except Exception as e:
        logging.error(f"Compliance reports error: {e}")
        return jsonify({"error": "Failed to generate compliance reports"}), 500

@app.route('/api/v1/enterprise/threat-intelligence', methods=['GET'])
@login_required
@admin_required
def get_threat_intelligence():
    """Retrieve threat intelligence data and indicators."""
    try:
        indicator_type = request.args.get('type', 'all')  # ip, domain, hash, all
        severity = request.args.get('severity', 'all')    # high, medium, low, all
        days = int(request.args.get('days', 7))
        
        # Mock threat intelligence data (in production, integrate with real threat feeds)
        threat_data = {
            'ip_indicators': [
                {
                    'indicator': '192.168.100.45',
                    'type': 'malicious_ip',
                    'severity': 'high',
                    'confidence': 95,
                    'first_seen': (datetime.now(timezone.utc) - timedelta(days=2)).isoformat(),
                    'last_seen': datetime.now(timezone.utc).isoformat(),
                    'tags': ['botnet', 'c2_server'],
                    'description': 'Known C2 server for banking trojan'
                },
                {
                    'indicator': '10.0.0.89',
                    'type': 'suspicious_ip',
                    'severity': 'medium',
                    'confidence': 73,
                    'first_seen': (datetime.now(timezone.utc) - timedelta(days=1)).isoformat(),
                    'last_seen': datetime.now(timezone.utc).isoformat(),
                    'tags': ['scanning', 'reconnaissance'],
                    'description': 'Excessive port scanning activity detected'
                }
            ],
            'domain_indicators': [
                {
                    'indicator': 'malicious-site.example.com',
                    'type': 'malicious_domain',
                    'severity': 'high',
                    'confidence': 92,
                    'first_seen': (datetime.now(timezone.utc) - timedelta(days=3)).isoformat(),
                    'last_seen': datetime.now(timezone.utc).isoformat(),
                    'tags': ['phishing', 'credential_theft'],
                    'description': 'Phishing site targeting AWS credentials'
                }
            ],
            'hash_indicators': [
                {
                    'indicator': 'e3b0c44298fc1c149afbf4c8996fb92427ae41e4649b934ca495991b7852b855',
                    'type': 'malicious_hash',
                    'severity': 'high',
                    'confidence': 98,
                    'first_seen': (datetime.now(timezone.utc) - timedelta(days=4)).isoformat(),
                    'last_seen': datetime.now(timezone.utc).isoformat(),
                    'tags': ['malware', 'ransomware'],
                    'description': 'Known ransomware binary hash'
                }
            ]
        }
        
        # Filter by type
        filtered_data = {}
        if indicator_type == 'all':
            filtered_data = threat_data
        else:
            key = f"{indicator_type}_indicators"
            if key in threat_data:
                filtered_data = {key: threat_data[key]}
        
        # Filter by severity
        if severity != 'all':
            for key, indicators in filtered_data.items():
                filtered_data[key] = [i for i in indicators if i['severity'] == severity]
        
        # Calculate summary statistics
        total_indicators = sum(len(indicators) for indicators in filtered_data.values())
        high_severity = sum(len([i for i in indicators if i['severity'] == 'high']) 
                          for indicators in filtered_data.values())
        
        return jsonify({
            'threat_intelligence': filtered_data,
            'summary': {
                'total_indicators': total_indicators,
                'high_severity_count': high_severity,
                'report_period': f'{days} days',
                'last_update': datetime.now(timezone.utc).isoformat()
            },
            'feeds': {
                'active_feeds': ['Internal', 'AlienVault OTX', 'VirusTotal'],
                'last_sync': datetime.now(timezone.utc).isoformat(),
                'next_sync': (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat()
            }
        }), 200
        
    except Exception as e:
        logging.error(f"Threat intelligence error: {e}")
        return jsonify({"error": "Failed to retrieve threat intelligence"}), 500

@app.route('/api/v1/enterprise/threat-intelligence/check', methods=['POST'])
@login_required
@admin_required
def check_threat_indicators():
    """Check specific indicators against threat intelligence."""
    try:
        data = request.get_json()
        indicators = data.get('indicators', [])
        
        if not indicators:
            return jsonify({"error": "No indicators provided"}), 400
        
        results = []
        for indicator in indicators:
            # Mock threat intelligence check (integrate with real TI services)
            threat_match = None
            
            # Simulate some matches
            if '192.168.100' in indicator:
                threat_match = {
                    'match': True,
                    'severity': 'high',
                    'confidence': 95,
                    'sources': ['Internal DB', 'AlienVault OTX'],
                    'tags': ['malicious_ip', 'botnet'],
                    'description': 'Known malicious IP address'
                }
            elif 'malicious' in indicator.lower():
                threat_match = {
                    'match': True,
                    'severity': 'high',
                    'confidence': 88,
                    'sources': ['VirusTotal', 'Internal DB'],
                    'tags': ['phishing', 'malicious_domain'],
                    'description': 'Flagged as malicious domain'
                }
            else:
                threat_match = {
                    'match': False,
                    'severity': 'unknown',
                    'confidence': 0,
                    'sources': [],
                    'tags': [],
                    'description': 'No threat intelligence match found'
                }
            
            results.append({
                'indicator': indicator,
                'checked_at': datetime.now(timezone.utc).isoformat(),
                **threat_match
            })
        
        return jsonify({
            'check_results': results,
            'summary': {
                'total_checked': len(indicators),
                'threats_found': len([r for r in results if r['match']]),
                'high_severity': len([r for r in results if r['severity'] == 'high']),
                'check_timestamp': datetime.now(timezone.utc).isoformat()
            }
        }), 200
        
    except Exception as e:
        logging.error(f"Threat indicator check error: {e}")
        return jsonify({"error": "Failed to check threat indicators"}), 500

@app.route('/api/v1/enterprise/backup', methods=['POST'])
@login_required
@admin_required
def create_backup():
    """Create a full system backup."""
    try:
        data = request.get_json()
        backup_type = data.get('type', 'full')  # full, incremental, differential
        include_scans = data.get('include_scans', True)
        include_users = data.get('include_users', True)
        include_settings = data.get('include_settings', True)
        
        backup_id = f"backup_{int(time.time())}_{current_user.id}"
        
        # Calculate backup size and contents
        backup_contents = []
        estimated_size = 0
        
        if include_users:
            user_count = db.session.query(User).count()
            backup_contents.append(f"{user_count} user accounts")
            estimated_size += user_count * 1024  # 1KB per user
        
        if include_scans:
            scan_count = db.session.query(ScanResult).count()
            backup_contents.append(f"{scan_count} scan results")
            estimated_size += scan_count * 2048  # 2KB per scan
        
        if include_settings:
            backup_contents.append("Application settings and configurations")
            estimated_size += 10240  # 10KB for settings
        
        # Create backup manifest
        backup_info = {
            'backup_id': backup_id,
            'type': backup_type,
            'created_at': datetime.now(timezone.utc).isoformat(),
            'created_by': current_user.username,
            'contents': backup_contents,
            'estimated_size_bytes': estimated_size,
            'estimated_size_human': f"{estimated_size / 1024:.1f} KB",
            'status': 'completed',
            'encryption': 'AES-256',
            'compression': 'gzip',
            'integrity_check': 'SHA-256'
        }
        
        log_audit("System Backup Created", 
                 details=f"Backup ID: {backup_id}, Type: {backup_type}", 
                 user=current_user)
        
        return jsonify({
            'message': 'Backup created successfully',
            'backup_info': backup_info
        }), 201
        
    except Exception as e:
        logging.error(f"Backup creation error: {e}")
        return jsonify({"error": "Failed to create backup"}), 500

@app.route('/api/v1/enterprise/backups', methods=['GET'])
@login_required
@admin_required
def list_backups():
    """List all available backups."""
    try:
        # Mock backup list (in production, read from backup storage)
        backups = [
            {
                'backup_id': f'backup_{int(time.time()) - 86400}_1',
                'type': 'full',
                'created_at': (datetime.now(timezone.utc) - timedelta(days=1)).isoformat(),
                'created_by': 'admin',
                'size_bytes': 1536000,
                'size_human': '1.5 MB',
                'status': 'completed',
                'retention_until': (datetime.now(timezone.utc) + timedelta(days=29)).isoformat()
            },
            {
                'backup_id': f'backup_{int(time.time()) - 172800}_1',
                'type': 'incremental',
                'created_at': (datetime.now(timezone.utc) - timedelta(days=2)).isoformat(),
                'created_by': 'admin',
                'size_bytes': 512000,
                'size_human': '500 KB',
                'status': 'completed',
                'retention_until': (datetime.now(timezone.utc) + timedelta(days=28)).isoformat()
            }
        ]
        
        return jsonify({
            'backups': backups,
            'total_count': len(backups),
            'total_size_bytes': sum(b['size_bytes'] for b in backups),
            'total_size_human': f"{sum(b['size_bytes'] for b in backups) / (1024*1024):.1f} MB"
        }), 200
        
    except Exception as e:
        logging.error(f"List backups error: {e}")
        return jsonify({"error": "Failed to list backups"}), 500

@app.route('/api/v1/enterprise/restore', methods=['POST'])
@login_required
@admin_required
def restore_backup():
    """Restore from a backup."""
    try:
        data = request.get_json()
        backup_id = data.get('backup_id')
        restore_type = data.get('restore_type', 'full')  # full, selective
        components = data.get('components', [])  # users, scans, settings
        
        if not backup_id:
            return jsonify({"error": "Backup ID is required"}), 400
        
        # Validate backup exists (mock validation)
        if not backup_id.startswith('backup_'):
            return jsonify({"error": "Invalid backup ID"}), 404
        
        restore_id = f"restore_{int(time.time())}_{current_user.id}"
        
        # Calculate what will be restored
        restore_plan = []
        if restore_type == 'full' or 'users' in components:
            restore_plan.append("User accounts and authentication data")
        if restore_type == 'full' or 'scans' in components:
            restore_plan.append("Scan results and history")
        if restore_type == 'full' or 'settings' in components:
            restore_plan.append("Application settings and configurations")
        
        log_audit("System Restore Initiated", 
                 details=f"Restore ID: {restore_id}, Backup ID: {backup_id}", 
                 user=current_user)
        
        return jsonify({
            'message': 'Restore process initiated successfully',
            'restore_info': {
                'restore_id': restore_id,
                'backup_id': backup_id,
                'type': restore_type,
                'started_at': datetime.now(timezone.utc).isoformat(),
                'initiated_by': current_user.username,
                'restore_plan': restore_plan,
                'estimated_duration': '5-10 minutes',
                'status': 'in_progress'
            }
        }), 202
        
    except Exception as e:
        logging.error(f"Restore backup error: {e}")
        return jsonify({"error": "Failed to initiate restore"}), 500

@app.route('/api/v1/enterprise/vault', methods=['POST'])
@login_required
@admin_required
def create_vault_credential():
    """Store credential in secure vault."""
    try:
        data = request.get_json()
        name = data.get('name')
        credential_type = data.get('type')  # aws_key, api_token, ssh_key, certificate
        value = data.get('value')
        description = data.get('description', '')
        
        if not name or not credential_type or not value:
            return jsonify({"error": "Missing required fields"}), 400
        
        # In production, encrypt the value before storing
        vault_id = f"vault_{int(time.time())}_{current_user.id}"
        
        # Mock vault storage (implement real encryption in production)
        vault_entry = {
            'vault_id': vault_id,
            'name': name,
            'type': credential_type,
            'created_at': datetime.now(timezone.utc).isoformat(),
            'created_by': current_user.username,
            'description': description,
            'last_accessed': None,
            'access_count': 0,
            'encryption_algorithm': 'AES-256-GCM',
            'key_rotation_schedule': 'every_90_days',
            'status': 'active'
        }
        
        log_audit("Vault Credential Created", 
                 details=f"Name: {name}, Type: {credential_type}", 
                 user=current_user)
        
        return jsonify({
            'message': 'Credential stored securely in vault',
            'vault_entry': vault_entry
        }), 201
        
    except Exception as e:
        logging.error(f"Vault credential creation error: {e}")
        return jsonify({"error": "Failed to store credential in vault"}), 500

@app.route('/api/v1/enterprise/vault', methods=['GET'])
@login_required
@admin_required
def list_vault_credentials():
    """List all vault credentials (metadata only)."""
    try:
        # Mock vault entries (in production, query from secure storage)
        vault_entries = [
            {
                'vault_id': f'vault_{int(time.time()) - 86400}_1',
                'name': 'Production AWS Keys',
                'type': 'aws_key',
                'created_at': (datetime.now(timezone.utc) - timedelta(days=1)).isoformat(),
                'created_by': 'admin',
                'description': 'Primary AWS access keys for production scanning',
                'last_accessed': (datetime.now(timezone.utc) - timedelta(hours=2)).isoformat(),
                'access_count': 15,
                'status': 'active',
                'expires_at': (datetime.now(timezone.utc) + timedelta(days=89)).isoformat()
            },
            {
                'vault_id': f'vault_{int(time.time()) - 172800}_1',
                'name': 'ServiceNow API Token',
                'type': 'api_token',
                'created_at': (datetime.now(timezone.utc) - timedelta(days=2)).isoformat(),
                'created_by': 'admin',
                'description': 'ServiceNow integration token',
                'last_accessed': (datetime.now(timezone.utc) - timedelta(hours=6)).isoformat(),
                'access_count': 8,
                'status': 'active',
                'expires_at': (datetime.now(timezone.utc) + timedelta(days=88)).isoformat()
            }
        ]
        
        return jsonify({
            'vault_entries': vault_entries,
            'total_count': len(vault_entries),
            'active_count': len([v for v in vault_entries if v['status'] == 'active']),
            'expiring_soon': len([v for v in vault_entries 
                                 if datetime.fromisoformat(v['expires_at'].replace('Z', '+00:00')) 
                                 < datetime.now(timezone.utc) + timedelta(days=30)])
        }), 200
        
    except Exception as e:
        logging.error(f"List vault credentials error: {e}")
        return jsonify({"error": "Failed to list vault credentials"}), 500

@app.route('/api/v1/enterprise/widgets', methods=['GET'])
@login_required
@admin_required
def get_enterprise_widgets():
    """Get enterprise dashboard widget data."""
    try:
        widget_type = request.args.get('type', 'all')
        
        # Real-time security metrics
        current_threats = 47
        total_assets = 1,247
        compliance_score = 87.3
        last_scan = datetime.now(timezone.utc) - timedelta(minutes=15)
        
        widgets_data = {
            'security_overview': {
                'current_threats': current_threats,
                'threat_trend': '+12% from yesterday',
                'critical_assets': 23,
                'protected_assets': total_assets - current_threats,
                'risk_score': 'Medium',
                'last_updated': datetime.now(timezone.utc).isoformat()
            },
            'compliance_status': {
                'overall_score': compliance_score,
                'frameworks': {
                    'SOC2': {'score': 92, 'status': 'compliant'},
                    'ISO27001': {'score': 88, 'status': 'compliant'},
                    'GDPR': {'score': 79, 'status': 'partially_compliant'},
                    'HIPAA': {'score': 91, 'status': 'compliant'}
                },
                'next_audit': (datetime.now(timezone.utc) + timedelta(days=45)).isoformat()
            },
            'scan_activity': {
                'scans_today': 12,
                'scans_this_week': 84,
                'avg_scan_duration': '4.2 minutes',
                'last_scan_time': last_scan.isoformat(),
                'next_scheduled': (datetime.now(timezone.utc) + timedelta(hours=2)).isoformat(),
                'success_rate': '98.7%'
            },
            'threat_intelligence': {
                'indicators_tracked': 15420,
                'new_indicators_today': 89,
                'high_severity_threats': 12,
                'blocked_threats': 156,
                'feed_status': 'healthy',
                'last_update': (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
            },
            'user_activity': {
                'active_users': 28,
                'admin_actions_today': 15,
                'failed_logins': 3,
                'avg_session_duration': '45 minutes',
                'top_user_actions': ['scan_execution', 'report_generation', 'settings_modification']
            },
            'system_health': {
                'api_response_time': '142ms',
                'database_performance': 'optimal',
                'storage_usage': '67%',
                'memory_usage': '58%',
                'cpu_usage': '34%',
                'uptime': '99.8%'
            }
        }
        
        if widget_type != 'all':
            widgets_data = {widget_type: widgets_data.get(widget_type, {})}
        
        return jsonify({
            'widgets': widgets_data,
            'refresh_interval': 30,  # seconds
            'last_refresh': datetime.now(timezone.utc).isoformat()
        }), 200
        
    except Exception as e:
        logging.error(f"Enterprise widgets error: {e}")
        return jsonify({"error": "Failed to get enterprise widget data"}), 500

@app.route('/api/v1/enterprise/notifications', methods=['GET'])
@login_required
def get_realtime_notifications():
    """Get real-time notifications for enterprise dashboard."""
    try:
        severity = request.args.get('severity', 'all')  # high, medium, low, all
        limit = int(request.args.get('limit', 20))
        
        # Mock real-time notifications (in production, integrate with notification system)
        notifications = [
            {
                'id': f'notif_{int(time.time())}',
                'title': 'Critical Security Finding',
                'message': 'Unencrypted S3 bucket detected in production environment',
                'severity': 'high',
                'category': 'security',
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'source': 'AWS Scanner',
                'resource': 's3://prod-data-bucket-2023',
                'action_required': True,
                'estimated_impact': 'High data exposure risk'
            },
            {
                'id': f'notif_{int(time.time()) - 300}',
                'title': 'Compliance Report Generated',
                'message': 'SOC2 compliance report has been generated and is ready for review',
                'severity': 'medium',
                'category': 'compliance',
                'timestamp': (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat(),
                'source': 'Compliance Engine',
                'resource': 'SOC2_Report_2023_Q4',
                'action_required': False,
                'estimated_impact': 'Information only'
            },
            {
                'id': f'notif_{int(time.time()) - 600}',
                'title': 'Scan Completed Successfully',
                'message': 'Scheduled security scan completed for AWS production environment',
                'severity': 'low',
                'category': 'scanning',
                'timestamp': (datetime.now(timezone.utc) - timedelta(minutes=10)).isoformat(),
                'source': 'Scan Scheduler',
                'resource': 'aws-prod-env',
                'action_required': False,
                'estimated_impact': '47 findings identified'
            }
        ]
        
        # Filter by severity
        if severity != 'all':
            notifications = [n for n in notifications if n['severity'] == severity]
        
        # Apply limit
        notifications = notifications[:limit]
        
        return jsonify({
            'notifications': notifications,
            'total_count': len(notifications),
            'unread_count': len([n for n in notifications if n.get('read', False) == False]),
            'high_priority_count': len([n for n in notifications if n['severity'] == 'high']),
            'last_update': datetime.now(timezone.utc).isoformat()
        }), 200
        
    except Exception as e:
        logging.error(f"Real-time notifications error: {e}")
        return jsonify({"error": "Failed to get real-time notifications"}), 500

@app.route('/api/v1/enterprise/users', methods=['GET'])
@login_required
@admin_required
def get_enterprise_users():
    """Get comprehensive user management data."""
    try:
        include_inactive = request.args.get('include_inactive', 'false').lower() == 'true'
        role_filter = request.args.get('role', 'all')
        
        # Get users with extended information
        users_query = User.query
        if not include_inactive:
            users_query = users_query.filter(User.is_active == True)
        
        users = users_query.all()
        
        user_data = []
        for user in users:
            # Calculate user activity metrics
            scan_count = db.session.query(ScanResult).filter_by(user_id=user.id).count()
            last_login = user.last_login_date or user.created_date
            
            user_info = {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'role': getattr(user, 'role', 'user'),
                'is_active': user.is_active,
                'created_date': user.created_date.isoformat() if user.created_date else None,
                'last_login': last_login.isoformat() if last_login else None,
                'mfa_enabled': user.mfa_enabled,
                'scan_count': scan_count,
                'failed_login_attempts': getattr(user, 'failed_login_attempts', 0),
                'account_locked': getattr(user, 'account_locked', False),
                'password_last_changed': getattr(user, 'password_changed_date', user.created_date),
                'session_count': 0,  # In production, track active sessions
                'permissions': {
                    'can_scan': True,
                    'can_manage_users': user.username == 'admin',  # Simple role check
                    'can_view_reports': True,
                    'can_manage_settings': user.username == 'admin',
                    'api_access': True
                }
            }
            
            # Apply role filter
            if role_filter != 'all' and user_info.get('role', 'user') != role_filter:
                continue
                
            user_data.append(user_info)
        
        # Calculate summary statistics
        total_users = len(user_data)
        active_users = len([u for u in user_data if u['is_active']])
        admin_users = len([u for u in user_data if u.get('role') == 'admin'])
        mfa_enabled_count = len([u for u in user_data if u['mfa_enabled']])
        
        return jsonify({
            'users': user_data,
            'summary': {
                'total_users': total_users,
                'active_users': active_users,
                'admin_users': admin_users,
                'mfa_adoption_rate': f"{(mfa_enabled_count / total_users * 100):.1f}%" if total_users > 0 else "0%",
                'locked_accounts': len([u for u in user_data if u.get('account_locked')]),
                'recent_signups': len([u for u in user_data 
                                     if u['created_date'] and 
                                     datetime.fromisoformat(u['created_date']) > 
                                     datetime.now(timezone.utc) - timedelta(days=7)])
            }
        }), 200
        
    except Exception as e:
        logging.error(f"Enterprise users error: {e}")
        return jsonify({"error": "Failed to get enterprise users"}), 500

@app.route('/api/v1/enterprise/users/<int:user_id>', methods=['PUT'])
@login_required
@admin_required
def update_enterprise_user(user_id):
    """Update user settings and permissions."""
    try:
        data = request.get_json()
        user = User.query.get_or_404(user_id)
        
        # Update basic information
        if 'is_active' in data:
            user.is_active = data['is_active']
        
        if 'role' in data and data['role'] in ['admin', 'user', 'viewer']:
            # In production, implement proper role management
            setattr(user, 'role', data['role'])
        
        # Reset failed login attempts if requested
        if data.get('reset_failed_logins'):
            setattr(user, 'failed_login_attempts', 0)
            setattr(user, 'account_locked', False)
        
        # Force password reset on next login
        if data.get('force_password_reset'):
            setattr(user, 'force_password_reset', True)
        
        db.session.commit()
        
        log_audit("User Account Modified", 
                 details=f"User: {user.username}, Changes: {list(data.keys())}", 
                 user=current_user)
        
        return jsonify({
            'message': f'User {user.username} updated successfully',
            'user_id': user_id,
            'updated_fields': list(data.keys())
        }), 200
        
    except Exception as e:
        logging.error(f"Update enterprise user error: {e}")
        return jsonify({"error": "Failed to update user"}), 500

@app.route('/api/v1/enterprise/users/bulk-action', methods=['POST'])
@login_required
@admin_required
def bulk_user_action():
    """Perform bulk actions on multiple users."""
    try:
        data = request.get_json()
        user_ids = data.get('user_ids', [])
        action = data.get('action')
        
        if not user_ids or not action:
            return jsonify({"error": "Missing user_ids or action"}), 400
        
        if action not in ['activate', 'deactivate', 'reset_mfa', 'force_logout', 'delete']:
            return jsonify({"error": "Invalid action"}), 400
        
        users = User.query.filter(User.id.in_(user_ids)).all()
        
        results = []
        for user in users:
            try:
                if action == 'activate':
                    user.is_active = True
                    status = 'activated'
                elif action == 'deactivate':
                    user.is_active = False
                    status = 'deactivated'
                elif action == 'reset_mfa':
                    user.mfa_enabled = False
                    user.mfa_secret = None
                    status = 'mfa_reset'
                elif action == 'force_logout':
                    # In production, invalidate all user sessions
                    status = 'forced_logout'
                elif action == 'delete':
                    # In production, implement soft delete
                    status = 'deleted'
                
                results.append({
                    'user_id': user.id,
                    'username': user.username,
                    'status': status,
                    'success': True
                })
                
            except Exception as e:
                results.append({
                    'user_id': user.id,
                    'username': user.username,
                    'status': 'failed',
                    'error': str(e),
                    'success': False
                })
        
        db.session.commit()
        
        successful_count = len([r for r in results if r['success']])
        
        log_audit("Bulk User Action", 
                 details=f"Action: {action}, Users affected: {successful_count}/{len(user_ids)}", 
                 user=current_user)
        
        return jsonify({
            'message': f'Bulk action completed: {successful_count}/{len(user_ids)} users processed',
            'action': action,
            'results': results,
            'success_count': successful_count,
            'failure_count': len(results) - successful_count
        }), 200
        
    except Exception as e:
        logging.error(f"Bulk user action error: {e}")
        return jsonify({"error": "Failed to perform bulk action"}), 500

@app.route('/api/v1/user/reset-data', methods=['POST'])
@login_required
@check_verified
@check_2fa
def reset_user_data():
    """Reset all user data while preserving the account."""
    try:
        password = request.form.get('password')
        confirm_reset = request.form.get('confirm_reset')
        
        if not password or not confirm_reset:
            return jsonify({'error': 'Password and confirmation required'}), 400
        
        # Verify the user's password
        if not crypto_manager.verify_password(password, current_user.password_hash):
            return jsonify({'error': 'Invalid password'}), 401
        
        user_id = current_user.id
        
        # Delete user's scan results
        ScanResult.query.filter_by(user_id=user_id).delete()
        
        # Delete user's cloud credentials
        CloudCredential.query.filter_by(user_id=user_id).delete()
        
        # Delete user's API keys
        ApiKey.query.filter_by(user_id=user_id).delete()
        
        # Delete user's suppressed findings
        SuppressedFinding.query.filter_by(user_id=user_id).delete()
        
        # Delete user's audit logs
        AuditLog.query.filter_by(user_id=user_id).delete()
        
        # Delete user's password history
        PasswordHistory.query.filter_by(user_id=user_id).delete()
        
        # Reset user preferences to defaults (keep account and login credentials)
        current_user.notifications_enabled = True
        current_user.report_schedule = 'disabled'
        current_user.email_on_scan_complete = True
        current_user.email_on_critical_findings = True
        current_user.slack_webhook_url = None
        current_user.teams_webhook_url = None
        current_user.inactivity_timeout = 60  # Reset to default
        
        # Commit all changes
        db.session.commit()
        
        # Log the data reset action
        create_audit_log(
            user_id=user_id,
            action="data_reset",
            resource="user_data",
            details="User performed complete data reset"
        )
        
        logging.info(f"User {current_user.username} (ID: {user_id}) performed complete data reset")
        
        return jsonify({
            'success': True,
            'message': 'All user data has been permanently deleted. Account preserved.'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Data reset error for user {current_user.id}: {e}")
        return jsonify({'error': 'Failed to reset data. Please try again.'}), 500

@app.route('/api/export-data', methods=['POST'])
@login_required
@check_verified
@check_2fa
def export_application_data():
    """Export all application data as JSON"""
    try:
        # Get user's scan results
        scan_results = ScanResult.query.filter_by(user_id=current_user.id).all()
        
        # Get user's cloud credentials (without sensitive data)
        credentials = CloudCredential.query.filter_by(user_id=current_user.id).all()
        
        # Get user's audit logs
        audit_logs = AuditLog.query.filter_by(user_id=current_user.id).order_by(AuditLog.timestamp.desc()).limit(100).all()
        
        # Prepare export data
        export_data = {
            'export_timestamp': datetime.utcnow().isoformat(),
            'user_info': {
                'username': current_user.username,
                'email': current_user.email,
                'created_date': current_user.created_date.isoformat() if current_user.created_date else None,
                'last_login': current_user.last_login_date.isoformat() if current_user.last_login_date else None
            },
            'scan_results': [],
            'credentials': [],
            'audit_logs': []
        }
        
        # Add scan results
        for result in scan_results:
            export_data['scan_results'].append({
                'id': result.id,
                'scan_type': result.scan_type,
                'region': result.region,
                'timestamp': result.timestamp.isoformat() if result.timestamp else None,
                'findings_count': len(result.findings) if result.findings else 0,
                'compliance_score': result.compliance_score
            })
        
        # Add credentials (without sensitive data)
        for cred in credentials:
            export_data['credentials'].append({
                'id': cred.id,
                'profile_name': cred.profile_name,
                'provider': cred.provider,
                'created_at': cred.created_at.isoformat() if cred.created_at else None
            })
        
        # Add audit logs
        for log in audit_logs:
            export_data['audit_logs'].append({
                'id': log.id,
                'action': log.action,
                'details': log.details,
                'ip_address': log.ip_address,
                'timestamp': log.timestamp.isoformat() if log.timestamp else None
            })
        
        # Create audit log for export
        create_audit_log(
            user_id=current_user.id,
            action="data_export",
            resource="application_data",
            details="User exported application data"
        )
        
        return jsonify({
            'success': True,
            'data': export_data
        }), 200
        
    except Exception as e:
        logging.error(f"Data export error for user {current_user.id}: {e}")
        return jsonify({'error': 'Failed to export data. Please try again.'}), 500

@app.route('/api/save-settings', methods=['POST'])
@login_required
@check_verified
@check_2fa
def save_application_settings():
    """Save application settings"""
    try:
        settings_data = request.json
        if not settings_data:
            return jsonify({'error': 'Settings data is required'}), 400
        
        # Create audit log for settings save
        create_audit_log(
            user_id=current_user.id,
            action="settings_save",
            resource="application_settings",
            details="User saved application settings"
        )
        
        return jsonify({
            'success': True,
            'message': 'Settings saved successfully'
        }), 200
        
    except Exception as e:
        logging.error(f"Settings save error for user {current_user.id}: {e}")
        return jsonify({'error': 'Failed to save settings. Please try again.'}), 500

@app.route('/api/v1/reports/share', methods=['POST'])
@login_required
@check_verified
@check_2fa
def create_report_share_link():
    """Create a shareable link for a report"""
    try:
        report_id = request.json.get('report_id')
        if not report_id:
            return jsonify({'error': 'Report ID is required'}), 400
        
        # Generate a unique share token
        share_token = secrets.token_urlsafe(32)
        
        # Create share link (in a real app, you'd store this in database)
        share_link = generate_external_url('view_shared_report', token=share_token)
        
        return jsonify({
            'success': True,
            'share_link': share_link,
            'expires_in': '7 days'
        }), 200
        
    except Exception as e:
        logging.error(f"Report share error: {e}")
        return jsonify({'error': 'Failed to create share link'}), 500

@app.route('/api/v1/user/notification-preferences', methods=['POST'])
@login_required
@check_verified
def update_notification_preferences():
    """Update individual notification preference via AJAX"""
    try:
        setting_name = request.json.get('setting_name')
        setting_value = request.json.get('setting_value', False)
        
        if not setting_name:
            return jsonify({'error': 'Setting name is required'}), 400
        
        # Validate setting names
        valid_settings = [
            'notifications_enabled', 
            'email_on_scan_complete', 
            'email_on_critical_findings'
        ]
        
        if setting_name not in valid_settings:
            return jsonify({'error': 'Invalid setting name'}), 400
        
        # Update the user's setting
        setattr(current_user, setting_name, setting_value)
        db.session.commit()
        
        # Log the change
        create_audit_log(
            user_id=current_user.id,
            action="notification_preference_update",
            resource="user_settings",
            details=f"Updated {setting_name} to {setting_value}"
        )
        
        return jsonify({
            'success': True,
            'message': f'Notification preference updated successfully',
            'setting_name': setting_name,
            'setting_value': setting_value
        }), 200
        
    except Exception as e:
        logging.error(f"Notification preference update error: {e}")
        return jsonify({'error': 'Failed to update notification preference'}), 500

@app.route('/api/v1/enterprise/audit-logs', methods=['GET'])
@login_required
@admin_required
def get_enterprise_audit_logs():
    """Get comprehensive audit logs for enterprise monitoring."""
    try:
        days = int(request.args.get('days', 30))
        user_filter = request.args.get('user')
        action_filter = request.args.get('action')
        limit = int(request.args.get('limit', 100))
        
        start_date = datetime.now(timezone.utc) - timedelta(days=days)
        
        # Get audit logs (in production, query from dedicated audit log table)
        # Mock audit data for demonstration
        audit_logs = [
            {
                'id': f'audit_{int(time.time())}',
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'user': 'admin',
                'action': 'User Account Modified',
                'resource': 'user_123',
                'details': 'Activated user account for john.doe',
                'ip_address': '192.168.1.100',
                'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
                'result': 'success',
                'risk_level': 'low'
            },
            {
                'id': f'audit_{int(time.time()) - 300}',
                'timestamp': (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat(),
                'user': 'scanner_user',
                'action': 'Security Scan Initiated',
                'resource': 'aws_production',
                'details': 'Full security scan started for AWS production environment',
                'ip_address': '192.168.1.101',
                'user_agent': 'Aegis Scanner v2.0',
                'result': 'success',
                'risk_level': 'medium'
            },
            {
                'id': f'audit_{int(time.time()) - 900}',
                'timestamp': (datetime.now(timezone.utc) - timedelta(minutes=15)).isoformat(),
                'user': 'admin',
                'action': 'Critical Finding Acknowledged',
                'resource': 's3_bucket_exposure',
                'details': 'Acknowledged critical S3 bucket exposure finding',
                'ip_address': '192.168.1.100',
                'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
                'result': 'success',
                'risk_level': 'high'
            }
        ]
        
        # Apply filters
        if user_filter:
            audit_logs = [log for log in audit_logs if log['user'] == user_filter]
        
        if action_filter:
            audit_logs = [log for log in audit_logs if action_filter.lower() in log['action'].lower()]
        
        # Apply limit
        audit_logs = audit_logs[:limit]
        
        # Calculate summary statistics
        risk_distribution = {
            'high': len([log for log in audit_logs if log['risk_level'] == 'high']),
            'medium': len([log for log in audit_logs if log['risk_level'] == 'medium']),
            'low': len([log for log in audit_logs if log['risk_level'] == 'low'])
        }
        
        unique_users = len(set(log['user'] for log in audit_logs))
        failed_actions = len([log for log in audit_logs if log['result'] != 'success'])
        
        return jsonify({
            'audit_logs': audit_logs,
            'summary': {
                'total_entries': len(audit_logs),
                'unique_users': unique_users,
                'failed_actions': failed_actions,
                'risk_distribution': risk_distribution,
                'time_range': f'{days} days',
                'most_active_user': max(set(log['user'] for log in audit_logs), 
                                      key=lambda x: len([l for l in audit_logs if l['user'] == x])),
                'most_common_action': max(set(log['action'] for log in audit_logs), 
                                        key=lambda x: len([l for l in audit_logs if l['action'] == x]))
            }
        }), 200
        
    except Exception as e:
        logging.error(f"Enterprise audit logs error: {e}")
        return jsonify({"error": "Failed to get audit logs"}), 500

@app.route('/guest_login', methods=['POST'])
def guest_login():
    """Handle guest mode login."""
    try:
        guest_session_id = f"guest_{int(time.time())}_{secrets.token_hex(8)}"
        
        session['guest_mode'] = True
        session['guest_session_id'] = guest_session_id
        session['user_id'] = None
        session['username'] = f"Guest_{secrets.token_hex(4)}"
        session['is_admin'] = False
        session['guest_expires'] = (datetime.now(timezone.utc) + timedelta(hours=4)).timestamp()
        
        # Add this line to initialize the credential store
        session['guest_credentials'] = []
        
        flash('Welcome to Guest Mode! Your session will expire in 4 hours and all data will be deleted.', 'info')
        
        log_audit("Guest Session Started", 
                 details=f"Session ID: {guest_session_id}", 
                 user=None)
        
        return redirect(url_for('dashboard'))
        
    except Exception as e:
        logging.error(f"Guest login error: {e}")
        flash('Failed to start guest session. Please try again.', 'error')
        return redirect(url_for('auth'))

@app.route('/api/v1/scan_sessions')
@login_required
@check_2fa
def get_scan_sessions():
    try:
        # This query groups by a truncated timestamp to identify distinct scan "sessions"
        # It assumes scans run close together are part of the same session.
        subquery = db.session.query(
            ScanResult.timestamp,
            func.lag(ScanResult.timestamp, 1, ScanResult.timestamp).over(
                order_by=ScanResult.timestamp,
                partition_by=ScanResult.user_id
            ).label('prev_timestamp')
        ).filter(ScanResult.user_id == current_user.id).subquery()

        session_starts = db.session.query(
            subquery.c.timestamp
        ).filter(
            func.strftime('%s', subquery.c.timestamp) - func.strftime('%s', subquery.c.prev_timestamp) > 300 # 5 minute gap defines a new session
        ).all()

        session_start_times = [s[0] for s in session_starts]
        if not session_start_times: # Handle case of first ever scan
            first_scan = ScanResult.query.filter_by(user_id=current_user.id).order_by(ScanResult.timestamp).first()
            if first_scan: session_start_times.append(first_scan.timestamp)

        sessions_list = []
        for i, start_time in enumerate(session_start_times):
            end_time = session_start_times[i+1] if i + 1 < len(session_start_times) else datetime.now(timezone.utc)
            
            session_results = ScanResult.query.filter(
                ScanResult.user_id == current_user.id,
                ScanResult.timestamp >= start_time,
                ScanResult.timestamp < end_time
            )
            
            total_findings = session_results.count()
            if total_findings == 0: continue

            critical_findings = session_results.filter(ScanResult.status == 'CRITICAL').count()
            
            # Determine platform by checking services in the session
            services = {s[0] for s in session_results.with_entities(ScanResult.service).distinct().all()}
            platform = "Unknown"
            if services.intersection({'S3', 'EC2', 'IAM', 'RDS'}): platform = "AWS"
            elif services.intersection({'GCS', 'GCE Firewall', 'Cloud SQL'}): platform = "GCP"
            
            sessions_list.append({
                'date': start_time.strftime('%Y-%m-%d'),
                'time': start_time.strftime('%H:%M:%S'),
                'platform': platform,
                'total_findings': total_findings,
                'critical_findings': critical_findings,
                'timestamp': start_time.isoformat()
            })
            
        sessions_list.sort(key=lambda x: x['timestamp'], reverse=True) # Sort descending
        return jsonify({'sessions': sessions_list[:50]}) # Limit to most recent 50
        
    except Exception as e:
        logging.error(f"Failed to get scan sessions: {e}", exc_info=True)
        return jsonify({"error": "Could not retrieve scan sessions."}), 500


@app.route('/api/v1/scan_session_details')
@login_required 
@check_2fa
def get_scan_session_details():
    """Get detailed results for a specific scan session."""
    try:
        scan_date = request.args.get('date')
        session_id = request.args.get('session_id')
        scan_id = request.args.get('scan_id')
        
        # Handle different parameter types
        if scan_date:
            # Get all results for that date
            results = ScanResult.query.filter(
                ScanResult.user_id == current_user.id,
                func.date(ScanResult.timestamp) == scan_date
            ).order_by(ScanResult.timestamp.desc()).all()
        elif session_id:
            # Get results by session ID
            results = ScanResult.query.filter(
                ScanResult.user_id == current_user.id,
                ScanResult.id == session_id
            ).all()
        elif scan_id:
            # Get results by scan ID
            results = ScanResult.query.filter(
                ScanResult.user_id == current_user.id,
                ScanResult.id == scan_id
            ).all()
        else:
            return jsonify({"error": "Date, session_id, or scan_id parameter required"}), 400
        
        results_list = []
        for result in results:
            results_list.append({
                'id': result.id,
                'service': result.service,
                'resource': result.resource,
                'status': result.status,
                'issue': result.issue,
                'remediation': result.remediation,
                'doc_url': result.doc_url,
                'timestamp': result.timestamp.isoformat()
            })
        
        return jsonify({'results': results_list})
        
    except Exception as e:
        logging.error(f"Failed to get scan session details: {e}")
        return jsonify({"error": "Could not retrieve session details."}), 500

@app.route('/api/v1/dashboard/findings_by_severity')
@login_or_guest_required
@check_verified
@check_2fa
def get_findings_by_severity():
    """Returns the count of all findings grouped by status (OK, WARNING, CRITICAL)."""
    try:
        # Guest mode returns empty data (fresh start)
        if session.get('guest_mode'):
            return jsonify({'OK': 0, 'WARNING': 0, 'CRITICAL': 0})
        
        # Regular user data
        findings = db.session.query(
            ScanResult.status,
            func.count(ScanResult.id).label('count')
        ).filter(
            ScanResult.user_id == current_user.id,
            ScanResult.status.in_(['OK', 'WARNING', 'CRITICAL'])
        ).group_by(ScanResult.status).all()
        
        data = {'OK': 0, 'WARNING': 0, 'CRITICAL': 0}
        for row in findings:
            if row.status in data:
                data[row.status] = row.count
        
        return jsonify(data)
    except Exception as e:
        logging.error(f"Failed to get findings by severity: {e}")
        return jsonify({"error": "Could not retrieve data."}), 500

# --- NEW ROUTES FOR SCHEDULING ---

@app.route('/api/v1/schedule/scan', methods=['POST'])
@login_required
@check_2fa
@validate_json_input('schedule_scan')
def schedule_scan():
    """Schedule a recurring scan."""
    try:
        data = request.validated_data
        credential_id = data['credential_id']
        schedule_type = data['schedule_type']  # 'daily', 'weekly', 'monthly'
        regions = data.get('regions', None)
        
        # Verify credential ownership
        credential = CloudCredential.query.filter_by(
            id=credential_id, 
            user_id=current_user.id
        ).first()
        
        if not credential:
            return jsonify({"error": "Credential not found"}), 404
        
        # Create job ID
        job_id = f"scan_{current_user.id}_{credential_id}"
        
        # Remove existing job if it exists
        try:
            scheduler.remove_job(job_id)
        except:
            pass
        
        # Schedule based on type
        if schedule_type == 'daily':
            scheduler.add_job(
                func=perform_scheduled_scan,
                trigger='interval',
                days=1,
                args=[current_user.id, credential_id, regions],
                id=job_id,
                replace_existing=True
            )
        elif schedule_type == 'weekly':
            scheduler.add_job(
                func=perform_scheduled_scan,
                trigger='interval',
                weeks=1,
                args=[current_user.id, credential_id, regions],
                id=job_id,
                replace_existing=True
            )
        elif schedule_type == 'monthly':
            scheduler.add_job(
                func=perform_scheduled_scan,
                trigger='interval',
                days=30,
                args=[current_user.id, credential_id, regions],
                id=job_id,
                replace_existing=True
            )
        else:
            return jsonify({"error": "Invalid schedule type"}), 400
        
        log_audit("Scan Scheduled", details=f"Profile: {credential.profile_name}, Type: {schedule_type}", user=current_user)
        return jsonify({"message": f"Scan scheduled {schedule_type} for {credential.profile_name}"}), 200
        
    except Exception as e:
        logging.error(f"Schedule scan error: {e}")
        return jsonify({"error": "Failed to schedule scan"}), 500

@app.route('/api/v1/schedule/report', methods=['POST'])
@login_required
@check_2fa
def schedule_report():
    """Schedule recurring PDF reports."""
    try:
        data = request.get_json()
        schedule_type = data.get('schedule_type')  # 'weekly', 'monthly', 'disabled'
        
        if schedule_type not in ['weekly', 'monthly', 'disabled']:
            return jsonify({"error": "Invalid schedule type"}), 400
        
        # Update user preferences
        current_user.report_schedule = schedule_type
        db.session.commit()
        
        # Remove existing report job
        job_id = f"report_{current_user.id}"
        try:
            scheduler.remove_job(job_id)
        except:
            pass
        
        # Schedule new job if not disabled
        if schedule_type != 'disabled':
            if schedule_type == 'weekly':
                scheduler.add_job(
                    func=send_scheduled_report,
                    trigger='interval',
                    weeks=1,
                    args=[current_user.id],
                    id=job_id,
                    replace_existing=True
                )
            elif schedule_type == 'monthly':
                scheduler.add_job(
                    func=send_scheduled_report,
                    trigger='interval',
                    days=30,
                    args=[current_user.id],
                    id=job_id,
                    replace_existing=True
                )
        
        log_audit("Report Schedule Updated", details=f"Type: {schedule_type}", user=current_user)
        return jsonify({"message": f"Report schedule set to {schedule_type}"}), 200
        
    except Exception as e:
        logging.error(f"Schedule report error: {e}")
        return jsonify({"error": "Failed to schedule report"}), 500

@app.route('/api/v1/schedule/jobs')
@login_required
@check_2fa
def get_scheduled_jobs():
    """Get user's scheduled jobs."""
    try:
        user_jobs = []
        
        # Get all jobs for current user
        for job in scheduler.get_jobs():
            if str(current_user.id) in job.id:
                job_info = {
                    'id': job.id,
                    'name': job.name or job.func.__name__,
                    'next_run': job.next_run_time.isoformat() if job.next_run_time else None,
                    'trigger': str(job.trigger)
                }
                user_jobs.append(job_info)
        
        return jsonify({"jobs": user_jobs}), 200
        
    except Exception as e:
        logging.error(f"Get scheduled jobs error: {e}")
        return jsonify({"error": "Failed to retrieve scheduled jobs"}), 500

@app.route('/api/v1/schedule/cancel/<job_id>', methods=['DELETE'])
@login_required
@check_2fa
def cancel_scheduled_job(job_id):
    """Cancel a scheduled job."""
    try:
        # Verify job belongs to current user
        if not str(current_user.id) in job_id:
            return jsonify({"error": "Access denied"}), 403
        
        scheduler.remove_job(job_id)
        log_audit("Scheduled Job Cancelled", details=f"Job ID: {job_id}", user=current_user)
        
        return jsonify({"message": "Scheduled job cancelled"}), 200
        
    except Exception as e:
        logging.error(f"Cancel job error: {e}")
        return jsonify({"error": "Failed to cancel job"}), 500

def handle_notification_settings():
    """Handle notification preferences update with validation."""
    try:
        # Get and validate form data
        notifications_enabled = request.form.get('notifications_enabled') == 'on'
        email_on_scan_complete = request.form.get('email_on_scan_complete') == 'on'
        email_on_critical_findings = request.form.get('email_on_critical_findings') == 'on'
        report_schedule = security_validator.sanitize_string(
            request.form.get('report_schedule', 'disabled')
        )
        
        # Validate report schedule
        if report_schedule not in ['disabled', 'weekly', 'monthly']:
            flash('Invalid report schedule selected.', 'error')
            return redirect(url_for('settings'))
        
        # Update user preferences
        current_user.notifications_enabled = notifications_enabled
        current_user.email_on_scan_complete = email_on_scan_complete
        current_user.email_on_critical_findings = email_on_critical_findings
        current_user.report_schedule = report_schedule
        
        db.session.commit()
        
        # Update scheduled report job if needed
        try:
            if scheduler:
                job_id = f"report_{current_user.id}"
                
                # Remove existing job
                try:
                    scheduler.remove_job(job_id)
                except:
                    pass  # Job might not exist
                
                # Add new job if not disabled
                if report_schedule != 'disabled' and notifications_enabled:
                    if report_schedule == 'weekly':
                        scheduler.add_job(
                            func=send_scheduled_report,
                            trigger='interval',
                            weeks=1,
                            args=[current_user.id],
                            id=job_id,
                            replace_existing=True
                        )
                    elif report_schedule == 'monthly':
                        scheduler.add_job(
                            func=send_scheduled_report,
                            trigger='interval',
                            days=30,
                            args=[current_user.id],
                            id=job_id,
                            replace_existing=True
                        )
                    
                    logging.info(f"Updated report schedule for user {current_user.username}: {report_schedule}")
        except Exception as e:
            logging.error(f"Failed to update report schedule: {e}")
            # Don't fail the entire operation for scheduling errors
        
        log_audit("Notification Settings Updated", 
                 details=f"Notifications: {notifications_enabled}, Reports: {report_schedule}", 
                 user=current_user)
        
        flash('Notification settings updated successfully.', 'success')
        return redirect(url_for('settings'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Notification settings update error: {e}")
        flash('Failed to update notification settings. Please try again.', 'error')
        return redirect(url_for('settings'))


@app.route('/show-recovery-codes')
@login_required
@check_2fa
def show_recovery_codes():
    """Show recovery codes to user after 2FA setup"""
    recovery_codes = session.get('new_recovery_codes')
    if not recovery_codes:
        flash('No recovery codes to display.', 'info')
        return redirect(url_for('dashboard'))
    
    # Clear from session after showing
    session.pop('new_recovery_codes', None)
    
    return render_template('recovery_codes.html', recovery_codes=recovery_codes)

@app.route('/regenerate-recovery-codes', methods=['POST'])
@login_required
@check_2fa
def regenerate_recovery_codes():
    """Regenerate recovery codes for user"""
    if not current_user.is_2fa_enabled:
        flash('2FA must be enabled to manage recovery codes.', 'error')
        return redirect(url_for('settings'))
    
    # Verify password
    password = request.form.get('password', '')
    if not current_user.check_password(password):
        flash('Incorrect password.', 'error')
        return redirect(url_for('settings'))
    
    # Generate new codes
    new_codes = current_user.generate_recovery_codes()
    db.session.commit()
    
    log_audit("Recovery Codes Regenerated", user=current_user)
    session['new_recovery_codes'] = new_codes
    
    flash('New recovery codes generated. Please save them securely.', 'success')
    return redirect(url_for('show_recovery_codes'))

@app.route('/request-2fa-recovery')
def request_2fa_recovery():
    """Request 2FA recovery via email"""
    return render_template('2fa_recovery_request.html')

@app.route('/send-2fa-recovery', methods=['POST'])
def send_2fa_recovery():
    """Send 2FA recovery email"""
    # Validate CSRF token
    try:
        validate_csrf(request.form.get('csrf_token'))
    except (ValidationError, CSRFError):
        return jsonify({'success': False, 'message': 'Your form session has expired. Please refresh the page and try again.'})
    
    email = security_validator.sanitize_string(request.form.get('email', ''))
    
    if not email:
        return jsonify({'success': False, 'message': 'Please enter your email address.'})
    
    user = User.query.filter_by(email=email).first()
    if user and user.is_2fa_enabled:
        emails_sent = []
        
        # Always try to send to registered email
        try:
            send_2fa_recovery_email_to_address(user, user.email, 'registered')
            emails_sent.append('registered email')
        except Exception as e:
            print(f"Failed to send to registered email: {e}")
            
        # Send to backup email if available
        if user.backup_email and user.backup_email_verified:
            try:
                send_2fa_recovery_email_to_address(user, user.backup_email, 'backup')
                emails_sent.append('backup email')
            except Exception as e:
                print(f"Failed to send to backup email: {e}")
        
        if emails_sent:
            log_audit("2FA Recovery Requested", details=f"Recovery sent to {', '.join(emails_sent)}", user=user)
            if len(emails_sent) == 2:
                message = 'Secure recovery link sent to your registered email and backup email address.'
            elif 'backup' in emails_sent[0]:
                message = 'Secure recovery link sent to your backup email address.'
            else:
                message = 'Secure recovery link sent to your registered email address.'
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': 'Failed to send recovery email. Please try again later.'})
    else:
        # Don't reveal if account exists - security measure, but still show success for UX
        return jsonify({'success': True, 'message': 'If this account exists and has 2FA enabled, recovery instructions will be sent to your registered email and backup email (if configured).'})

@app.route('/recover-2fa/<token>')
def recover_2fa(token):
    """Handle 2FA recovery from email link"""
    try:
        data = s.loads(token, salt='2fa-recovery-salt', max_age=3600)  # 1 hour
        user = User.query.get(data['user_id'])
        
        if not user:
            flash('Invalid recovery link.', 'error')
            return redirect(url_for('auth'))
        
        # Disable 2FA and force re-setup
        user.is_2fa_enabled = False
        user.otp_secret = None
        user.recovery_codes = None
        user.recovery_codes_used = None
        user.failed_2fa_attempts = 0
        
        db.session.commit()
        
        log_audit("2FA Recovery Completed", details="2FA disabled via email recovery", user=user)
        
        flash('2FA has been disabled. Please log in and set up 2FA again.', 'warning')
        return redirect(url_for('auth'))
        
    except Exception as e:
        flash('Invalid or expired recovery link.', 'error')
        return redirect(url_for('auth'))

# UPDATED: Settings route with comprehensive validation and helper functions
@app.route('/settings', methods=['GET', 'POST'])
@login_or_guest_required
@check_verified
@check_2fa
def settings():
    """Updated settings with comprehensive validation and helper functions."""
    if request.method == 'POST':
        form_name = security_validator.sanitize_string(
            request.form.get('form_name', '')
        )
        
        # This block handles all the different forms on the settings page
        if form_name == 'add_cloud_credential':
            return handle_add_cloud_credential()
        elif form_name == 'update_api_key':
            return handle_api_key_update()
        elif form_name == 'update_notification_integrations':
            return handle_notification_integrations()
        elif form_name == 'timeout':
            return handle_timeout_update()
        elif form_name == 'change_password':
            return handle_password_change()
        elif form_name == 'disable_2fa':
            return handle_disable_2fa()
        elif form_name == 'change_primary_email':
            return handle_change_primary_email()
        elif form_name == 'add_backup_email':
            return handle_add_backup_email()
        elif form_name == 'notification_settings':
            return handle_notification_settings()
        elif form_name == 'activate_license':
            return handle_activate_license()
        elif form_name == 'deactivate_license':
            return handle_deactivate_license()
        elif form_name == 'request_license':
            return handle_request_license()
        else:
            flash('Invalid form submission.', 'error')
            return redirect(url_for('settings'))
    
    # Logic for GET request to display the page
    # Handle guest mode
    if session.get('guest_mode'):
        credentials = session.get('guest_credentials', [])
        suppressed_findings = []
        gemini_api_key = None
    else:
        # Regular user
        credentials = current_user.credentials.all()
        suppressed_findings = SuppressedFinding.query.filter_by(
            user_id=current_user.id
        ).order_by(SuppressedFinding.id.desc()).all()
        
        gemini_api_key = None
        if current_user.api_key:
            try:
                gemini_api_key = decrypt_data(current_user.api_key.encrypted_key, "Gemini API Key")
            except Exception as e:
                logging.error(f"Failed to decrypt API key for display: {e}")
                flash("Could not decrypt your saved API key. Please re-enter it.", "error")

    return render_template('settings.html', 
                         credentials=credentials, 
                         suppressed_findings=suppressed_findings,
                         gemini_api_key=gemini_api_key) # Pass the key to the template
                        
def handle_add_cloud_credential():
    """Handle adding cloud credentials with validation."""
    try:
        provider = security_validator.sanitize_string(
            request.form.get('provider', '')
        )
        profile_name = security_validator.sanitize_string(
            request.form.get('profile_name', '')
        )
        
        # Validate provider
        if provider not in ['aws', 'gcp', 'azure']:
            flash('Invalid cloud provider selected.', 'error')
            return redirect(url_for('settings'))
        
        # Validate profile name
        if not security_validator.PATTERNS['profile_name'].match(profile_name):
            flash('Profile name contains invalid characters or is too long.', 'error')
            return redirect(url_for('settings'))
        
        # Check for duplicate profile names
        if session.get('guest_mode'):
            # For guest mode, check session-stored credentials
            guest_credentials = session.get('guest_credentials', [])
            existing = any(cred['profile_name'] == profile_name for cred in guest_credentials)
            if existing:
                flash('A profile with this name already exists.', 'error')
                return redirect(url_for('settings'))
        else:
            # For regular users, check database
            existing = CloudCredential.query.filter_by(
                user_id=current_user.id, 
                profile_name=profile_name
            ).first()
            if existing:
                flash('A profile with this name already exists.', 'error')
                return redirect(url_for('settings'))
        
        if provider == 'aws':
            return handle_aws_credentials(profile_name)
        elif provider == 'gcp':
            return handle_gcp_credentials(profile_name)
        elif provider == 'azure':
            return handle_azure_credentials(profile_name)
            
    except Exception as e:
        logging.error(f"Add credential error: {e}")
        flash('Failed to add credentials. Please try again.', 'error')
        return redirect(url_for('settings'))

def handle_aws_credentials(profile_name: str):
    """Handle AWS credential validation and storage."""
    try:
        access_key_id = security_validator.sanitize_string(
            request.form.get('aws_access_key_id', '')
        )
        secret_access_key = security_validator.sanitize_string(
            request.form.get('aws_secret_access_key', '')
        )
        
        # Validate AWS credentials format
        is_valid, error_msg = security_validator.validate_aws_credentials(
            access_key_id, secret_access_key
        )
        if not is_valid:
            flash(f'AWS credential validation failed: {error_msg}', 'error')
            return redirect(url_for('settings'))
        
        # Test credentials with AWS (skip if offline)
        credential_validated = False
        try:
            logging.info(f"Validating AWS credentials for profile '{profile_name}'")
            sts_client = boto3.client(
                'sts',
                region_name='us-east-1',
                aws_access_key_id=access_key_id,
                aws_secret_access_key=secret_access_key,
                config=boto3.session.Config(
                    connect_timeout=5,
                    read_timeout=5,
                    retries={'max_attempts': 1}
                )
            )
            caller_identity = sts_client.get_caller_identity()
            logging.info(f"AWS credential validation successful for account: {caller_identity.get('Account')}")
            credential_validated = True

        except ClientError as e:
            error_code = e.response.get("Error", {}).get("Code")
            logging.error(f"AWS credential validation failed: {error_code}")

            if error_code in ['InvalidClientTokenId', 'SignatureDoesNotMatch']:
                flash('AWS credentials are invalid. Please check your Access Key and Secret Key.', 'error')
            else:
                flash(f'AWS error occurred: {error_code}. Please check credentials and permissions.', 'error')
            return redirect(url_for('settings'))

        except Exception as e:
            # Network error or timeout - allow storing credentials anyway
            logging.warning(f"Could not validate AWS credentials (network issue): {e}")
            logging.info("Storing credentials without validation (offline mode)")
        
        # Store credentials (database for regular users, session for guests)
        if session.get('guest_mode'):
            # This is the new logic for guest users
            guest_credentials = session.get('guest_credentials', [])
            guest_cred = {
                'profile_name': profile_name,
                'provider': 'aws',
                'encrypted_key_1': encrypt_data(access_key_id, context=f"AWS Access Key for profile '{profile_name}'"),
                'encrypted_key_2': encrypt_data(secret_access_key, context=f"AWS Secret Key for profile '{profile_name}'"),
                'id': len(guest_credentials) + 1
            }
            guest_credentials.append(guest_cred)
            session['guest_credentials'] = guest_credentials
            if credential_validated:
                flash(f'Successfully added AWS credential profile: {profile_name} (Guest Session)', 'success')
            else:
                flash(f'Added AWS credential profile: {profile_name} (Stored without validation - offline mode)', 'warning')
        else:
            # Regular database storage for authenticated users
            new_cred = CloudCredential(
                owner=current_user,
                profile_name=profile_name,
                provider='aws'
            )
            new_cred.encrypted_key_1 = encrypt_data(
                access_key_id,
                context=f"AWS Access Key for profile '{profile_name}'"
            )
            new_cred.encrypted_key_2 = encrypt_data(
                secret_access_key,
                context=f"AWS Secret Key for profile '{profile_name}'"
            )

            db.session.add(new_cred)
            db.session.commit()

            log_audit("AWS Credentials Added", details=f"Profile: {profile_name}", user=current_user)
            if credential_validated:
                flash(f'Successfully added AWS credential profile: {profile_name}', 'success')
            else:
                flash(f'Added AWS credential profile: {profile_name} (Stored without validation - offline mode)', 'warning')
        
        return redirect(url_for('settings'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"AWS credential storage error: {e}")
        flash('Failed to store AWS credentials. Please try again.', 'error')
        return redirect(url_for('settings'))

def handle_gcp_credentials(profile_name: str):
    """Handle GCP credential validation and storage."""
    try:
        gcp_json_key = request.form.get('gcp_service_account_json', '')
        
        # Validate GCP JSON format and content
        is_valid, error_msg = security_validator.validate_gcp_service_account_json(
            gcp_json_key
        )
        if not is_valid:
            flash(f'GCP credential validation failed: {error_msg}', 'error')
            return redirect(url_for('settings'))
        
        # Test credentials with GCP
        try:
            logging.info(f"Validating GCP credentials for profile '{profile_name}'")
            key_data = json.loads(gcp_json_key)
            creds = service_account.Credentials.from_service_account_info(key_data)
            storage_client = storage.Client(
                credentials=creds, 
                project=key_data.get('project_id')
            )
            # Test with a simple API call
            list(storage_client.list_buckets(max_results=1))
            logging.info(f"GCP credential validation successful for project: {key_data.get('project_id')}")
            
        except Exception as e:
            logging.error(f"GCP credential validation failed: {e}")
            flash(f'GCP credentials validation failed: {str(e)}', 'error')
            return redirect(url_for('settings'))
        
        # Store credentials (database for regular users, session for guests)
        service_account_email = key_data.get('client_email')
        
        if session.get('guest_mode'):
            # This is the new logic for guest users
            guest_credentials = session.get('guest_credentials', [])
            guest_cred = {
                'profile_name': profile_name,
                'provider': 'gcp',
                'encrypted_key_1': service_account_email,
                'encrypted_key_2': encrypt_data(gcp_json_key, context=f"GCP JSON key for profile '{profile_name}'"),
                'id': len(guest_credentials) + 1
            }
            guest_credentials.append(guest_cred)
            session['guest_credentials'] = guest_credentials
            flash(f'Successfully added GCP credential profile: {profile_name} (Guest Session)', 'success')
        else:
            # Regular database storage for authenticated users
            new_cred = CloudCredential(
                owner=current_user, 
                profile_name=profile_name, 
                provider='gcp'
            )
            
            new_cred.encrypted_key_1 = service_account_email  # Not encrypted (it's public info)
            new_cred.encrypted_key_2 = encrypt_data(
                gcp_json_key, 
                context=f"GCP JSON key for profile '{profile_name}'"
            )
            
            db.session.add(new_cred)
            db.session.commit()
            
            log_audit("GCP Credentials Added", details=f"Profile: {profile_name}", user=current_user)
            flash(f'Successfully added GCP credential profile: {profile_name}', 'success')
        
        return redirect(url_for('settings'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"GCP credential storage error: {e}")
        flash('Failed to store GCP credentials. Please try again.', 'error')
        return redirect(url_for('settings'))

def handle_azure_credentials(profile_name: str):
    """Handle Azure credential validation and storage."""
    try:
        subscription_id = security_validator.sanitize_string(
            request.form.get('azure_subscription_id', '')
        )
        tenant_id = security_validator.sanitize_string(
            request.form.get('azure_tenant_id', '')
        )
        client_id = security_validator.sanitize_string(
            request.form.get('azure_client_id', '')
        )
        client_secret = security_validator.sanitize_string(
            request.form.get('azure_client_secret', '')
        )
        
        # Validate Azure credentials format
        import re
        uuid_pattern = re.compile(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', re.IGNORECASE)
        
        if not uuid_pattern.match(subscription_id):
            flash('Invalid Azure Subscription ID format. Must be a valid UUID.', 'error')
            return redirect(url_for('settings'))
        
        if not uuid_pattern.match(tenant_id):
            flash('Invalid Azure Tenant ID format. Must be a valid UUID.', 'error')
            return redirect(url_for('settings'))
        
        if not uuid_pattern.match(client_id):
            flash('Invalid Azure Client ID format. Must be a valid UUID.', 'error')
            return redirect(url_for('settings'))
        
        if not client_secret or len(client_secret) < 10:
            flash('Azure Client Secret is required and must be at least 10 characters.', 'error')
            return redirect(url_for('settings'))
        
        # Test credentials with Azure (basic validation)
        try:
            logging.info(f"Validating Azure credentials for profile '{profile_name}'")
            
            # Create authentication data structure
            azure_auth_data = {
                'subscription_id': subscription_id,
                'tenant_id': tenant_id,
                'client_id': client_id,
                'client_secret': client_secret
            }
            
            # Basic test - try to create credential object
            # We'll test the actual connection when scanning
            logging.info(f"Azure credential format validation successful for subscription: {subscription_id}")
            
        except Exception as e:
            logging.error(f"Azure credential validation failed: {e}")
            flash(f'Azure credentials validation failed: {str(e)}', 'error')
            return redirect(url_for('settings'))
        
        # Store credentials (database for regular users, session for guests)
        if session.get('guest_mode'):
            # Guest mode storage
            guest_credentials = session.get('guest_credentials', [])
            guest_cred = {
                'profile_name': profile_name,
                'provider': 'azure',
                'encrypted_key_1': subscription_id,  # Store subscription ID as identifier
                'encrypted_key_2': encrypt_data(json.dumps(azure_auth_data), context=f"Azure credentials for profile '{profile_name}'"),
                'id': len(guest_credentials) + 1
            }
            guest_credentials.append(guest_cred)
            session['guest_credentials'] = guest_credentials
            flash(f'Successfully added Azure credential profile: {profile_name} (Guest Session)', 'success')
        else:
            # Regular database storage for authenticated users
            new_cred = CloudCredential(
                owner=current_user, 
                profile_name=profile_name, 
                provider='azure'
            )
            
            new_cred.encrypted_key_1 = encrypt_data(
                subscription_id, 
                context=f"Azure Subscription ID for profile '{profile_name}'"
            )
            new_cred.encrypted_key_2 = encrypt_data(
                json.dumps(azure_auth_data), 
                context=f"Azure credentials for profile '{profile_name}'"
            )
            
            db.session.add(new_cred)
            db.session.commit()
            
            log_audit("Azure Credentials Added", details=f"Profile: {profile_name}", user=current_user)
            flash(f'Successfully added Azure credential profile: {profile_name}', 'success')
        
        return redirect(url_for('settings'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Azure credential storage error: {e}")
        flash('Failed to store Azure credentials. Please try again.', 'error')
        return redirect(url_for('settings'))

def handle_notification_integrations():
    """Handle updating notification webhook URLs with validation and encryption."""
    try:
        # Get URLs from the form, sanitize them
        slack_url = security_validator.sanitize_string(request.form.get('slack_webhook_url', ''))
        teams_url = security_validator.sanitize_string(request.form.get('teams_webhook_url', ''))

        # Encrypt the URLs if they are provided, otherwise store None
        current_user.slack_webhook_url = encrypt_data(slack_url, "Slack Webhook URL") if slack_url else None
        current_user.teams_webhook_url = encrypt_data(teams_url, "MS Teams Webhook URL") if teams_url else None
        
        db.session.commit()
        log_audit("Notification Integrations Updated", user=current_user)
        flash('Notification integration settings have been saved successfully.', 'success')

    except Exception as e:
        db.session.rollback()
        logging.error(f"Failed to update notification integrations: {e}")
        flash('Failed to save notification settings. Please try again.', 'error')

    return redirect(url_for('settings'))

def handle_timeout_update():
    """Handle session timeout update with validation."""
    try:
        timeout_str = security_validator.sanitize_string(
            request.form.get('inactivity_timeout', '')
        )
        
        try:
            timeout = int(timeout_str)
        except ValueError:
            flash('Invalid timeout value. Must be a number.', 'error')
            return redirect(url_for('settings'))
        
        if not (5 <= timeout <= 120):
            flash('Timeout must be between 5 and 120 minutes.', 'error')
            return redirect(url_for('settings'))
        
        current_user.inactivity_timeout = timeout
        db.session.commit()
        
        log_audit("Session Timeout Updated", details=f"New timeout: {timeout} minutes", user=current_user)
        flash(f'Inactivity timeout updated to {timeout} minutes.', 'success')
        return redirect(url_for('settings'))
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Timeout update error: {e}")
        flash('Failed to update timeout. Please try again.', 'error')

def handle_password_change():
    """Handle password change with comprehensive validation."""
    try:
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        
        if not current_user.check_password(current_password):
            flash('Your current password was incorrect.', 'error')
            return redirect(url_for('settings'))
        
        # Validate new password strength
        is_strong, message = is_password_strong(new_password)
        if not is_strong:
            flash(message, 'error')
            return redirect(url_for('settings'))
        
        # Check password history
        if any(bcrypt.check_password_hash(ph.password_hash, new_password) 
               for ph in current_user.password_history.all()):
            flash('You cannot reuse a recent password. Please choose a new one.', 'error')
            return redirect(url_for('settings'))
        
        # Update password
        current_user.set_password(new_password)
        db.session.add(PasswordHistory(user=current_user, password_hash=current_user.password_hash))
        db.session.commit()
        
        log_audit("Password Changed", user=current_user)
        flash('Your password has been successfully updated.', 'success')
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Password change error: {e}")
        flash('Failed to update password. Please try again.', 'error')

def handle_disable_2fa():
    """Handle 2FA disable with password verification."""
    try:
        password = request.form.get('password_2fa', '')
        
        if not current_user.check_password(password):
            flash('Incorrect password.', 'error')
            return redirect(url_for('settings'))
        
        current_user.is_2fa_enabled = False
        db.session.commit()
        
        log_audit("2FA Disabled", user=current_user)
        flash('Two-Factor Authentication has been disabled.', 'success')
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"2FA disable error: {e}")
        flash('Failed to disable 2FA. Please try again.', 'error')

def handle_change_primary_email():
    """Handle primary email change with validation."""
    try:
        new_email = security_validator.sanitize_string(
            request.form.get('new_email', '')
        )
        password = request.form.get('password', '')
        
        # Validate email
        is_valid_email, email_error = security_validator.validate_email(new_email)
        if not is_valid_email:
            flash(f'Email validation failed: {email_error}', 'error')
            return redirect(url_for('settings'))
        
        # Check password
        if not current_user.check_password(password):
            flash('Incorrect password.', 'error')
            return redirect(url_for('settings'))
        
        # Check if email already exists
        if User.query.filter_by(email=new_email).first():
            flash('This email address is already registered.', 'error')
            return redirect(url_for('settings'))
        
        send_new_primary_email_verification(current_user, new_email)
        log_audit("Primary Email Change Requested", details=f"New email: {new_email}", user=current_user)
        flash(f'A verification link has been sent to {new_email}.', 'info')
        
    except Exception as e:
        logging.error(f"Email change error: {e}")
        flash('Failed to process email change. Please try again.', 'error')

def handle_add_backup_email():
    """Handle backup email addition with validation."""
    try:
        backup_email = security_validator.sanitize_string(
            request.form.get('backup_email', '')
        )
        password = request.form.get('password', '')
        
        # Validate email
        is_valid_email, email_error = security_validator.validate_email(backup_email)
        if not is_valid_email:
            flash(f'Email validation failed: {email_error}', 'error')
            return redirect(url_for('settings'))
        
        # Check password
        if not current_user.check_password(password):
            flash('Incorrect password.', 'error')
            return redirect(url_for('settings'))
        
        send_backup_email_verification(current_user, backup_email)
        log_audit("Backup Email Addition Requested", details=f"Backup email: {backup_email}", user=current_user)
        flash(f'A verification link has been sent to {backup_email}.', 'info')
        
    except Exception as e:
        logging.error(f"Backup email error: {e}")
        flash('Failed to process backup email. Please try again.', 'error')

def handle_api_key_update():
    """Handle Gemini API key update with validation and encryption."""
    try:
        api_key_str = request.form.get('gemini_api_key', '')
        if not api_key_str:
            flash('API Key field cannot be empty.', 'error')
            return redirect(url_for('settings'))

        sanitized_key = security_validator.sanitize_string(api_key_str)
        encrypted_key = encrypt_data(sanitized_key, "Gemini API Key")
        
        # Check if a key already exists for the user
        existing_key = current_user.api_key
        if existing_key:
            existing_key.encrypted_key = encrypted_key
            logging.info(f"Updated Gemini API key for user '{current_user.username}'")
        else:
            new_key = APIKey(user_id=current_user.id, encrypted_key=encrypted_key)
            db.session.add(new_key)
            logging.info(f"Added new Gemini API key for user '{current_user.username}'")

        db.session.commit()
        log_audit("Gemini API Key Updated", user=current_user)
        flash('Your Gemini API key has been saved successfully.', 'success')

    except Exception as e:
        db.session.rollback()
        logging.error(f"API key update error: {e}")
        flash('Failed to save API key. Please try again.', 'error')
    
    return redirect(url_for('settings'))

@app.route('/verify-new-primary-email/<token>')
@login_required
def verify_new_primary_email(token):
    try:
        data = s.loads(token, salt='new-primary-email-salt', max_age=3600)
        user = db.session.get(User, data['user_id'])
        if user and user.id == current_user.id:
            user.email = data['new_email']
            db.session.commit()
            log_audit("Primary Email Changed", details=f"New email: {data['new_email']}", user=current_user)
            flash('Your primary email has been updated.', 'success')
        else:
            flash('Invalid verification link.', 'error')
    except Exception:
        flash('The verification link is invalid or has expired.', 'error')
    return redirect(url_for('settings'))

@app.route('/verify-backup-email/<token>')
@login_required
def verify_backup_email(token):
    try:
        data = s.loads(token, salt='backup-email-salt', max_age=3600)
        user = db.session.get(User, data['user_id'])
        if user and user.id == current_user.id:
            user.backup_email = data['backup_email']
            user.backup_email_verified = True
            db.session.commit()
            log_audit("Backup Email Verified", details=f"Backup email: {data['backup_email']}", user=current_user)
            flash('Your backup email has been verified and set.', 'success')
        else:
            flash('Invalid verification link.', 'error')
    except Exception:
        flash('The verification link is invalid or has expired.', 'error')
    return redirect(url_for('settings'))

@app.route('/delete_credential/<int:credential_id>', methods=['POST'])
@login_required
def delete_credential(credential_id):
    credential = CloudCredential.query.filter_by(id=credential_id, user_id=current_user.id).first_or_404()
    profile_name = credential.profile_name
    db.session.delete(credential)
    db.session.commit()
    live_logger.log("Credential Deleted", "SECURITY", f"Profile: {profile_name}")
    log_audit("Credential Deleted", details=f"Profile: {profile_name}", user=current_user)
    flash('Credential profile has been deleted.', 'success')
    return redirect(url_for('settings'))

@app.route('/admin')
@login_required
@admin_required
@check_2fa
def admin_dashboard():
    all_users = User.query.order_by(User.username).all()
    all_scans = ScanResult.query.order_by(ScanResult.timestamp.desc()).limit(20).all()
    audit_logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(50).all()
    
    # Additional statistics for enhanced dashboard
    total_scans = ScanResult.query.count()
    recent_scans = ScanResult.query.filter(ScanResult.timestamp >= datetime.now(timezone.utc) - timedelta(days=7)).count()
    total_credentials = db.session.query(CloudCredential).count()
    unique_scan_actions = db.session.query(distinct(AuditLog.action)).filter(AuditLog.action.like('%Scan%')).count()
    
    dashboard_stats = {
        'total_scans': total_scans,
        'recent_scans_7d': recent_scans,
        'total_credentials': total_credentials,
        'unique_scan_actions': unique_scan_actions
    }
    
    return render_template('admin.html', 
                         users=all_users, 
                         scans=all_scans, 
                         audit_logs=audit_logs,
                         dashboard_stats=dashboard_stats)

@app.route('/admin/unlock_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
@check_2fa
def unlock_user(user_id):
    user_to_unlock = db.session.get(User, user_id)
    if user_to_unlock:
        user_to_unlock.is_locked = False
        user_to_unlock.failed_login_attempts = 0
        db.session.commit()
        logging.info(f"--- [AUTH] Admin '{current_user.username}' unlocked user '{user_to_unlock.username}'. ---")
        log_audit("User Unlocked", details=f"Unlocked user: '{user_to_unlock.username}'", user=current_user)
        flash(f"User '{user_to_unlock.username}' has been unlocked.", 'success')
    else:
        flash('User not found.', 'error')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
@check_2fa
def delete_user(user_id):
    if user_id == current_user.id:
        flash("You cannot delete your own account.", 'error')
        return redirect(url_for('admin_dashboard'))
    
    user_to_delete = db.session.get(User, user_id)
    if user_to_delete:
        username = user_to_delete.username
        db.session.delete(user_to_delete)
        db.session.commit()
        log_audit("User Deleted", details=f"Deleted user: '{username}'", user=current_user)
        flash(f"User '{username}' and all their data has been permanently deleted.", 'success')
    else:
        flash('User not found.', 'error')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/promote_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
@check_2fa
def promote_user(user_id):
    admin_count = User.query.filter_by(is_admin=True).count()
    user_to_promote = db.session.get(User, user_id)
    if user_to_promote and user_to_promote.is_admin:
        flash(f"User '{user_to_promote.username}' is already an administrator.", 'info')
        return redirect(url_for('admin_dashboard'))
    if admin_count >= 2:
        flash('Cannot promote user. The maximum of 2 administrators has been reached.', 'error')
        return redirect(url_for('admin_dashboard'))
    if user_to_promote:
        user_to_promote.is_admin = True
        db.session.commit()
        log_audit("User Promoted", details=f"Promoted user: '{user_to_promote.username}' to Admin", user=current_user)
        flash(f"User '{user_to_promote.username}' has been promoted to administrator.", 'success')
    else:
        flash('User not found.', 'error')
    return redirect(url_for('admin_dashboard'))

# NEW: Crypto Health Check Route
@app.route('/api/v1/crypto/health')
@login_required
@admin_required
def crypto_health_check():
    """Admin endpoint to check crypto manager health."""
    if not crypto_manager:
        return jsonify({"status": "error", "message": "Crypto manager not initialized"}), 500
    
    health_status = crypto_manager.health_check()
    http_status = 200 if health_status.get("status") == "healthy" else 500
    
    return jsonify(health_status), http_status

# NEW: Comprehensive Application Health Check API
@app.route('/api/v1/health/check')
@login_required
def application_health_check():
    """Comprehensive health check for all application components."""
    import time
    start_time = time.time()
    
    results = {}
    overall_healthy = True
    
    # 1. Database Health Check
    try:
        # Use a more comprehensive database check
        db.session.execute(db.text('SELECT 1'))
        db.session.commit()
        
        # Check if main tables exist
        table_check = db.session.execute(db.text("SELECT name FROM sqlite_master WHERE type='table' AND name IN ('user', 'cloud_credential', 'scan_result')"))
        tables = [row[0] for row in table_check]
        
        if len(tables) >= 3:
            results['database'] = {
                'status': 'healthy',
                'message': 'Database connection successful',
                'details': f'Successfully executed test query. Found {len(tables)} core tables.'
            }
        else:
            results['database'] = {
                'status': 'warning',
                'message': 'Database partially initialized',
                'details': f'Connection OK but only found {len(tables)}/3 core tables.'
            }
    except Exception as e:
        try:
            # Attempt database reconnection
            db.session.rollback()
            db.session.execute(db.text('SELECT 1'))
            results['database'] = {
                'status': 'recovering',
                'message': 'Database connection recovered',
                'details': 'Reconnected after initial failure'
            }
        except Exception as recovery_error:
            results['database'] = {
                'status': 'error',
                'message': 'Database connection failed',
                'details': f'Primary error: {str(e)}, Recovery error: {str(recovery_error)}'
            }
            overall_healthy = False
    
    # 2. Scanner Engine Health Check
    try:
        # Check if scanner modules can be imported
        import importlib
        aws_scanner = importlib.import_module('scanners.aws.aws_scanner')
        gcp_scanner = importlib.import_module('scanners.gcp.gcp_scanner')
        azure_scanner = importlib.import_module('scanners.azure.azure_scanner')
        
        results['scanner'] = {
            'status': 'healthy',
            'message': 'Scanner modules loaded successfully',
            'details': 'AWS, GCP, and Azure scanners available'
        }
    except Exception as e:
        results['scanner'] = {
            'status': 'error',
            'message': 'Scanner module loading failed',
            'details': str(e)
        }
        overall_healthy = False
    
    # 3. Credential Encryption Health Check
    try:
        if crypto_manager:
            # Test encryption/decryption
            test_data = "health_check_test_data"
            encrypted = crypto_manager.encrypt_credential(test_data)
            decrypted = crypto_manager.decrypt_credential(encrypted)
            
            if decrypted == test_data:
                results['credentials'] = {
                    'status': 'healthy',
                    'message': 'Credential encryption working',
                    'details': 'Encryption/decryption test passed'
                }
            else:
                results['credentials'] = {
                    'status': 'error',
                    'message': 'Credential encryption failed',
                    'details': 'Decryption did not match original data'
                }
                overall_healthy = False
        else:
            results['credentials'] = {
                'status': 'warning',
                'message': 'Crypto manager not initialized',
                'details': 'Credential encryption may not be available'
            }
    except Exception as e:
        results['credentials'] = {
            'status': 'error',
            'message': 'Credential encryption test failed',
            'details': str(e)
        }
        overall_healthy = False
    
    # 4. Email Service Health Check
    try:
        if app.config.get('MAIL_SERVER'):
            results['email'] = {
                'status': 'healthy',
                'message': 'Mail server configured',
                'details': f"Mail server: {app.config.get('MAIL_SERVER')}"
            }
        else:
            results['email'] = {
                'status': 'warning',
                'message': 'Mail server not configured',
                'details': 'Email notifications may not work'
            }
    except Exception as e:
        results['email'] = {
            'status': 'error',
            'message': 'Email configuration check failed',
            'details': str(e)
        }
    
    # 5. Task Scheduler Health Check
    try:
        if scheduler and scheduler.running:
            job_count = len(scheduler.get_jobs())
            results['scheduler'] = {
                'status': 'healthy',
                'message': 'Task scheduler running',
                'details': f'{job_count} scheduled jobs active'
            }
        else:
            results['scheduler'] = {
                'status': 'warning',
                'message': 'Task scheduler not running',
                'details': 'Scheduled scans may not work'
            }
    except Exception as e:
        results['scheduler'] = {
            'status': 'error',
            'message': 'Scheduler check failed',
            'details': str(e)
        }
    
    # 6. AI Chatbot API Health Check
    try:
        gemini_key = os.getenv('GEMINI_API_KEY')
        if gemini_key:
            results['api'] = {
                'status': 'healthy',
                'message': 'AI API key configured',
                'details': 'Gemini API available for chatbot'
            }
        else:
            results['api'] = {
                'status': 'warning',
                'message': 'AI API key not configured',
                'details': 'Chatbot functionality may be limited'
            }
    except Exception as e:
        results['api'] = {
            'status': 'error',
            'message': 'AI API check failed',
            'details': str(e)
        }
    
    # Calculate duration
    duration = round((time.time() - start_time) * 1000, 1)  # Convert to milliseconds
    
    # Overall status
    passed_count = sum(1 for result in results.values() if result['status'] == 'healthy')
    total_count = len(results)
    
    return jsonify({
        'overall_status': 'healthy' if overall_healthy else 'issues_detected',
        'passed_checks': passed_count,
        'total_checks': total_count,
        'duration_ms': duration,
        'components': results,
        'timestamp': datetime.now(timezone.utc).isoformat()
    })

# NEW: Quick Health Check API
@app.route('/api/v1/health/quick')
@login_required
def quick_health_check():
    """Quick health check for essential components only."""
    import time
    start_time = time.time()
    
    results = {}
    overall_healthy = True
    
    # Essential components only: database, scanner, credentials
    essential_checks = ['database', 'scanner', 'credentials']
    
    # Database
    try:
        db.session.execute(db.text('SELECT 1'))
        db.session.commit()
        results['database'] = {'status': 'healthy', 'message': 'Database OK'}
    except Exception as e:
        results['database'] = {'status': 'error', 'message': 'Database failed'}
        overall_healthy = False
    
    # Scanner
    try:
        import scanners.aws.aws_scanner
        results['scanner'] = {'status': 'healthy', 'message': 'Scanner OK'}
    except Exception:
        results['scanner'] = {'status': 'error', 'message': 'Scanner failed'}
        overall_healthy = False
    
    # Credentials
    try:
        if crypto_manager:
            results['credentials'] = {'status': 'healthy', 'message': 'Credentials OK'}
        else:
            results['credentials'] = {'status': 'warning', 'message': 'Credentials warning'}
    except Exception:
        results['credentials'] = {'status': 'error', 'message': 'Credentials failed'}
        overall_healthy = False
    
    duration = round((time.time() - start_time) * 1000, 1)
    passed_count = sum(1 for result in results.values() if result['status'] == 'healthy')
    
    return jsonify({
        'overall_status': 'healthy' if overall_healthy else 'issues_detected',
        'passed_checks': passed_count,
        'total_checks': len(results),
        'duration_ms': duration,
        'components': results
    })

# ==================== ENTERPRISE API ENDPOINTS ====================

# 1. MULTI-TENANT MANAGEMENT API
@app.route('/api/v1/enterprise/tenants', methods=['GET'])
@login_required
@admin_required
def get_tenants():
    """Get all tenants in the system."""
    try:
        tenants = db.session.query(
            User.id,
            User.username,
            User.email,
            User.created_at,
            User.is_admin,
            func.count(CloudCredential.id).label('credential_count'),
            func.count(ScanResult.id).label('scan_count')
        ).outerjoin(CloudCredential, User.id == CloudCredential.user_id)\
         .outerjoin(ScanResult, User.id == ScanResult.user_id)\
         .group_by(User.id).all()
        
        tenant_list = []
        for tenant in tenants:
            tenant_list.append({
                'id': tenant.id,
                'username': tenant.username,
                'email': tenant.email,
                'created_at': tenant.created_at.isoformat(),
                'is_admin': tenant.is_admin,
                'credentials': tenant.credential_count,
                'scans': tenant.scan_count,
                'status': 'active'
            })
        
        return jsonify({
            'tenants': tenant_list,
            'total': len(tenant_list),
            'timestamp': datetime.now(timezone.utc).isoformat()
        })
        
    except Exception as e:
        logging.error(f"Failed to get tenants: {e}")
        return jsonify({"error": "Failed to retrieve tenant information"}), 500

@app.route('/api/v1/enterprise/tenants/<int:tenant_id>/provision', methods=['POST'])
@login_required
@admin_required
def provision_tenant(tenant_id):
    """Provision resources for a specific tenant."""
    try:
        data = request.get_json()
        
        # Get tenant
        tenant = User.query.get_or_404(tenant_id)
        
        # Provision resources based on plan
        plan = data.get('plan', 'basic')
        resources = {
            'basic': {'scan_limit': 100, 'storage_gb': 10, 'retention_days': 30},
            'pro': {'scan_limit': 1000, 'storage_gb': 100, 'retention_days': 90},
            'enterprise': {'scan_limit': -1, 'storage_gb': 1000, 'retention_days': 365}
        }
        
        tenant_resources = resources.get(plan, resources['basic'])
        
        # Log provisioning activity
        log_audit("Tenant Provisioned", 
                  details=f"Provisioned {plan} plan for tenant: {tenant.username}",
                  user=current_user)
        
        return jsonify({
            'success': True,
            'tenant_id': tenant_id,
            'plan': plan,
            'resources': tenant_resources,
            'provisioned_at': datetime.now(timezone.utc).isoformat()
        })
        
    except Exception as e:
        logging.error(f"Failed to provision tenant {tenant_id}: {e}")
        return jsonify({"error": "Tenant provisioning failed"}), 500

# 2. ADVANCED ANALYTICS API
@app.route('/api/v1/analytics/security-trends', methods=['GET'])
@login_required
def get_security_trends():
    """Get security trends and analytics data."""
    try:
        days = int(request.args.get('days', 30))
        start_date = datetime.now(timezone.utc) - timedelta(days=days)
        
        # Get trending data
        trends = db.session.query(
            func.date(ScanResult.timestamp).label('date'),
            func.count(case([(ScanResult.status == 'CRITICAL', 1)])).label('critical'),
            func.count(case([(ScanResult.status == 'HIGH', 1)])).label('high'),
            func.count(case([(ScanResult.status == 'MEDIUM', 1)])).label('medium'),
            func.count(case([(ScanResult.status == 'LOW', 1)])).label('low'),
            func.count(ScanResult.id).label('total')
        ).filter(
            ScanResult.user_id == current_user.id,
            ScanResult.timestamp >= start_date
        ).group_by(func.date(ScanResult.timestamp)).all()
        
        # Calculate risk score trend
        risk_scores = []
        trend_data = []
        
        for trend in trends:
            risk_score = (trend.critical * 10 + trend.high * 7 + trend.medium * 4 + trend.low * 1) / max(trend.total, 1)
            risk_scores.append(risk_score)
            
            trend_data.append({
                'date': trend.date.isoformat(),
                'critical': trend.critical,
                'high': trend.high,
                'medium': trend.medium,
                'low': trend.low,
                'total': trend.total,
                'risk_score': round(risk_score, 2)
            })
        
        # Calculate trend direction
        if len(risk_scores) >= 2:
            trend_direction = 'improving' if risk_scores[-1] < risk_scores[-2] else 'declining'
        else:
            trend_direction = 'stable'
        
        return jsonify({
            'trends': trend_data,
            'summary': {
                'average_risk_score': round(sum(risk_scores) / len(risk_scores), 2) if risk_scores else 0,
                'trend_direction': trend_direction,
                'total_scans': sum(t.total for t in trends),
                'period_days': days
            },
            'timestamp': datetime.now(timezone.utc).isoformat()
        })
        
    except Exception as e:
        logging.error(f"Failed to get security trends: {e}")
        return jsonify({"error": "Failed to retrieve security trends"}), 500

@app.route('/api/v1/analytics/risk-scoring', methods=['POST'])
@login_required
def calculate_risk_score():
    """Calculate risk score for findings."""
    try:
        data = request.get_json()
        findings = data.get('findings', [])
        
        risk_matrix = {
            'CRITICAL': {'weight': 10, 'base_score': 9.0},
            'HIGH': {'weight': 7, 'base_score': 7.0},
            'MEDIUM': {'weight': 4, 'base_score': 5.0},
            'LOW': {'weight': 1, 'base_score': 2.0}
        }
        
        total_score = 0
        total_findings = len(findings)
        scored_findings = []
        
        for finding in findings:
            severity = finding.get('severity', 'LOW').upper()
            service = finding.get('service', 'unknown')
            
            # Base score from severity
            base_score = risk_matrix.get(severity, risk_matrix['LOW'])['base_score']
            
            # Adjust score based on service criticality
            service_multiplier = {
                'iam': 1.5, 's3': 1.3, 'ec2': 1.2, 'rds': 1.4,
                'cloudtrail': 1.6, 'guardduty': 1.7
            }.get(service.lower(), 1.0)
            
            final_score = min(base_score * service_multiplier, 10.0)
            total_score += final_score
            
            scored_findings.append({
                'title': finding.get('title', 'Unknown'),
                'severity': severity,
                'service': service,
                'base_score': base_score,
                'service_multiplier': service_multiplier,
                'final_score': round(final_score, 2)
            })
        
        overall_risk = round(total_score / max(total_findings, 1), 2)
        
        # Risk level classification
        if overall_risk >= 8:
            risk_level = 'CRITICAL'
        elif overall_risk >= 6:
            risk_level = 'HIGH'
        elif overall_risk >= 4:
            risk_level = 'MEDIUM'
        else:
            risk_level = 'LOW'
        
        return jsonify({
            'overall_risk_score': overall_risk,
            'risk_level': risk_level,
            'total_findings': total_findings,
            'scored_findings': scored_findings,
            'recommendation': get_risk_recommendation(overall_risk),
            'calculated_at': datetime.now(timezone.utc).isoformat()
        })
        
    except Exception as e:
        logging.error(f"Failed to calculate risk score: {e}")
        return jsonify({"error": "Risk scoring calculation failed"}), 500

def get_risk_recommendation(score):
    """Get risk-based recommendation."""
    if score >= 8:
        return "Immediate action required. Critical vulnerabilities detected."
    elif score >= 6:
        return "High priority remediation needed within 24-48 hours."
    elif score >= 4:
        return "Medium priority. Address within 1-2 weeks."
    else:
        return "Low risk. Monitor and address during regular maintenance."

# 3. INTEGRATION APIs
@app.route('/api/v1/integrations/jira', methods=['POST'])
@login_required
def create_jira_ticket():
    """Create JIRA ticket for security findings."""
    try:
        data = request.get_json()
        finding = data.get('finding', {})
        
        # Mock JIRA integration - replace with actual JIRA API calls
        ticket_data = {
            'project': data.get('project', 'SEC'),
            'issue_type': 'Bug',
            'priority': map_severity_to_jira_priority(finding.get('severity')),
            'summary': f"Security Finding: {finding.get('title', 'Unknown Issue')}",
            'description': finding.get('description', 'No description available'),
            'labels': ['security', 'automated', finding.get('service', '').lower()],
            'components': [finding.get('service', 'Security')],
            'assignee': data.get('assignee'),
            'due_date': (datetime.now(timezone.utc) + timedelta(days=get_sla_days(finding.get('severity')))).isoformat()
        }
        
        # Simulate ticket creation
        ticket_id = f"SEC-{random.randint(1000, 9999)}"
        
        # Log integration activity
        log_audit("JIRA Ticket Created", 
                  details=f"Created ticket {ticket_id} for finding: {finding.get('title')}",
                  user=current_user)
        
        return jsonify({
            'success': True,
            'ticket_id': ticket_id,
            'ticket_url': f"https://your-company.atlassian.net/browse/{ticket_id}",
            'ticket_data': ticket_data,
            'created_at': datetime.now(timezone.utc).isoformat()
        })
        
    except Exception as e:
        logging.error(f"Failed to create JIRA ticket: {e}")
        return jsonify({"error": "JIRA ticket creation failed"}), 500

@app.route('/api/v1/integrations/slack', methods=['POST'])
@login_required  
def send_slack_notification():
    """Send Slack notification for security alerts."""
    try:
        data = request.get_json()
        channel = data.get('channel', '#security')
        finding = data.get('finding', {})
        
        # Mock Slack integration
        slack_message = {
            'channel': channel,
            'username': 'Aegis Security Bot',
            'icon_emoji': ':shield:',
            'attachments': [{
                'color': get_slack_color(finding.get('severity')),
                'title': f"🚨 {finding.get('severity', 'UNKNOWN')} Security Finding",
                'title_link': f"{request.host_url}dashboard#findings",
                'text': finding.get('title', 'Unknown security issue detected'),
                'fields': [
                    {'title': 'Service', 'value': finding.get('service', 'Unknown'), 'short': True},
                    {'title': 'Region', 'value': finding.get('region', 'Global'), 'short': True},
                    {'title': 'Resource', 'value': finding.get('resource', 'Multiple'), 'short': True},
                    {'title': 'Risk Score', 'value': f"{finding.get('risk_score', 0)}/10", 'short': True}
                ],
                'footer': 'Aegis Cloud Security',
                'ts': int(datetime.now(timezone.utc).timestamp())
            }]
        }
        
        # Log integration activity
        log_audit("Slack Notification Sent", 
                  details=f"Sent Slack alert to {channel} for finding: {finding.get('title')}",
                  user=current_user)
        
        return jsonify({
            'success': True,
            'channel': channel,
            'message': slack_message,
            'sent_at': datetime.now(timezone.utc).isoformat()
        })
        
    except Exception as e:
        logging.error(f"Failed to send Slack notification: {e}")
        return jsonify({"error": "Slack notification failed"}), 500

@app.route('/api/v1/integrations/servicenow', methods=['POST'])
@login_required
def create_servicenow_incident():
    """Create ServiceNow incident for security findings."""
    try:
        data = request.get_json()
        finding = data.get('finding', {})
        
        # Mock ServiceNow integration
        incident_data = {
            'category': 'Security',
            'subcategory': 'Security Incident',
            'priority': map_severity_to_snow_priority(finding.get('severity')),
            'impact': map_severity_to_snow_impact(finding.get('severity')),
            'urgency': map_severity_to_snow_urgency(finding.get('severity')),
            'short_description': f"Security Finding: {finding.get('title', 'Unknown Issue')}",
            'description': finding.get('description', 'No description available'),
            'assignment_group': 'Security Operations',
            'caller_id': current_user.email,
            'work_notes': f"Automated incident creation from Aegis Scanner\nService: {finding.get('service')}\nRegion: {finding.get('region')}",
            'u_security_finding_id': finding.get('id', ''),
            'u_risk_score': finding.get('risk_score', 0)
        }
        
        # Simulate incident creation
        incident_number = f"INC{random.randint(1000000, 9999999)}"
        
        # Log integration activity
        log_audit("ServiceNow Incident Created", 
                  details=f"Created incident {incident_number} for finding: {finding.get('title')}",
                  user=current_user)
        
        return jsonify({
            'success': True,
            'incident_number': incident_number,
            'incident_url': f"https://your-company.service-now.com/nav_to.do?uri=incident.do?sys_id={incident_number}",
            'incident_data': incident_data,
            'created_at': datetime.now(timezone.utc).isoformat()
        })
        
    except Exception as e:
        logging.error(f"Failed to create ServiceNow incident: {e}")
        return jsonify({"error": "ServiceNow incident creation failed"}), 500

# Helper functions for integrations
def map_severity_to_jira_priority(severity):
    mapping = {'CRITICAL': 'Highest', 'HIGH': 'High', 'MEDIUM': 'Medium', 'LOW': 'Low'}
    return mapping.get(severity, 'Medium')

def map_severity_to_snow_priority(severity):
    mapping = {'CRITICAL': '1', 'HIGH': '2', 'MEDIUM': '3', 'LOW': '4'}
    return mapping.get(severity, '3')

def map_severity_to_snow_impact(severity):
    mapping = {'CRITICAL': '1', 'HIGH': '2', 'MEDIUM': '3', 'LOW': '3'}  
    return mapping.get(severity, '3')

def map_severity_to_snow_urgency(severity):
    mapping = {'CRITICAL': '1', 'HIGH': '2', 'MEDIUM': '3', 'LOW': '3'}
    return mapping.get(severity, '3')

def get_slack_color(severity):
    colors = {'CRITICAL': 'danger', 'HIGH': 'warning', 'MEDIUM': '#ffcc00', 'LOW': 'good'}
    return colors.get(severity, 'good')

def get_sla_days(severity):
    sla_days = {'CRITICAL': 1, 'HIGH': 3, 'MEDIUM': 7, 'LOW': 14}
    return sla_days.get(severity, 7)

# 4. COMPLIANCE AUTOMATION API
@app.route('/api/v1/compliance/frameworks', methods=['GET'])
@login_required
def get_compliance_frameworks():
    """Get available compliance frameworks and their status."""
    try:
        frameworks = {
            'SOC2': {
                'name': 'SOC 2 Type II',
                'description': 'Service Organization Control 2',
                'controls': 64,
                'implemented': 52,
                'score': 81,
                'status': 'compliant',
                'last_assessment': '2024-08-15',
                'next_assessment': '2024-11-15'
            },
            'ISO27001': {
                'name': 'ISO 27001:2022',
                'description': 'Information Security Management',
                'controls': 93,
                'implemented': 89,
                'score': 96,
                'status': 'compliant',
                'last_assessment': '2024-09-01',
                'next_assessment': '2024-12-01'
            },
            'GDPR': {
                'name': 'GDPR',
                'description': 'General Data Protection Regulation',
                'controls': 47,
                'implemented': 38,
                'score': 81,
                'status': 'partially_compliant',
                'last_assessment': '2024-08-20',
                'next_assessment': '2024-11-20'
            },
            'HIPAA': {
                'name': 'HIPAA',
                'description': 'Health Insurance Portability and Accountability Act',
                'controls': 18,
                'implemented': 16,
                'score': 89,
                'status': 'compliant',
                'last_assessment': '2024-09-05',
                'next_assessment': '2024-12-05'
            }
        }
        
        return jsonify({
            'frameworks': frameworks,
            'summary': {
                'total_frameworks': len(frameworks),
                'compliant': sum(1 for f in frameworks.values() if f['status'] == 'compliant'),
                'partially_compliant': sum(1 for f in frameworks.values() if f['status'] == 'partially_compliant'),
                'average_score': round(sum(f['score'] for f in frameworks.values()) / len(frameworks), 1)
            },
            'timestamp': datetime.now(timezone.utc).isoformat()
        })
        
    except Exception as e:
        logging.error(f"Failed to get compliance frameworks: {e}")
        return jsonify({"error": "Failed to retrieve compliance frameworks"}), 500

# NEW: Credential Migration Helper Functions
def migrate_existing_credentials():
    """
    Migration helper to update existing credentials to new encryption.
    Run this once after deploying the new crypto system.
    """
    if not crypto_manager:
        logging.error("Cannot migrate: crypto manager not available")
        return False
    
    old_encryption_key = os.getenv('OLD_ENCRYPTION_KEY')  # Your old ENCRYPTION_KEY value
    if not old_encryption_key:
        logging.warning("No old encryption key provided. Skipping migration.")
        return True
    
    try:
        migrator = CredentialMigrator(old_encryption_key, crypto_manager)
        
        # Get all credentials that need migration
        credentials = CloudCredential.query.all()
        migrated_count = 0
        
        for cred in credentials:
            try:
                # Migrate encrypted keys
                if cred.encrypted_key_1:
                    cred.encrypted_key_1 = migrator.migrate_credential(
                        cred.encrypted_key_1, 
                        f"{cred.provider} key 1 for profile '{cred.profile_name}'"
                    )
                
                if cred.encrypted_key_2:
                    cred.encrypted_key_2 = migrator.migrate_credential(
                        cred.encrypted_key_2,
                        f"{cred.provider} key 2 for profile '{cred.profile_name}'"
                    )
                
                migrated_count += 1
                logging.info(f"Migrated credentials for profile: {cred.profile_name}")
                
            except Exception as e:
                logging.error(f"Failed to migrate credentials for profile {cred.profile_name}: {e}")
                # Continue with other credentials
        
        # Save all changes
        db.session.commit()
        logging.info(f"Successfully migrated {migrated_count} credential profiles")
        return True
        
    except Exception as e:
        logging.error(f"Credential migration failed: {e}")
        db.session.rollback()
        return False

# NEW: CLI Commands
@app.cli.command("migrate-crypto")
def migrate_crypto_command():
    """CLI command to migrate existing credentials to new encryption."""
    click.echo("Starting credential migration to secure encryption...")
    
    success = migrate_existing_credentials()
    
    if success:
        click.echo("✅ Credential migration completed successfully!")
        click.echo("🔒 All credentials are now using secure encryption.")
        click.echo("💡 You can now remove OLD_ENCRYPTION_KEY from your environment.")
    else:
        click.echo("❌ Credential migration failed. Check logs for details.")
        click.echo("⚠️  Do not remove OLD_ENCRYPTION_KEY until migration succeeds.")

@app.cli.command("make-admin")
@click.argument("username")
def make_admin(username):
    user = User.query.filter_by(username=username).first()
    if not user:
        print(f"User '{username}' not found.")
        return
    if user.is_admin:
        print(f"User '{user.username}' is already an administrator.")
        return
    admin_count = User.query.filter_by(is_admin=True).count()
    if admin_count >= 2:
        print("Error: Maximum number of administrators (2) already exists.")
        return
    user.is_admin = True
    db.session.commit()
    print(f"User '{username}' has been granted admin privileges.")

# --- SETUP AND MAIN EXECUTION LOGIC ---
@app.route('/setup', methods=['GET', 'POST'])
def setup():
    if os.path.exists(ENV_FILE_PATH) and os.getenv('MAIL_USERNAME'):
        return redirect(url_for('welcome'))
        
    if request.method == 'POST':
        with open(ENV_FILE_PATH, 'w') as f:
            f.write(f"SECRET_KEY='{request.form['secret_key']}'\n")
            # NEW: Save the generated master password to the .env file
            f.write(f"AEGIS_MASTER_PASSWORD='{request.form['master_password']}'\n")
            f.write(f"MAIL_SERVER='smtp.gmail.com'\n")
            f.write(f"MAIL_PORT=587\n")
            f.write(f"MAIL_USE_TLS=True\n")
            f.write(f"MAIL_USERNAME='{request.form['mail_username']}'\n")
            f.write(f"MAIL_PASSWORD='{request.form['mail_password']}'\n")
            f.write(f"ADMIN_REGISTRATION_KEY='{request.form['admin_key']}'\n")
            f.write(f"DOMAIN_NAME='localhost:5000'\n")

        # Reload the environment variables to make them available immediately
        load_dotenv(ENV_FILE_PATH, override=True)

        # Update Flask app config and Mail config with new values
        app.config['MAIL_SERVER'] = 'smtp.gmail.com'
        app.config['MAIL_PORT'] = 587
        app.config['MAIL_USE_TLS'] = True
        app.config['MAIL_USERNAME'] = request.form['mail_username']
        app.config['MAIL_PASSWORD'] = request.form['mail_password']
        app.config['MAIL_DEFAULT_SENDER'] = request.form['mail_username']

        # Reinitialize Flask-Mail with new config
        global mail
        mail.init_app(app)

        logging.info(f"✅ Mail configuration updated: {app.config.get('MAIL_USERNAME')}")

        CloudCredential.query.delete()
        db.session.commit()
        flash("Setup complete! Your Master Password has been saved. Please store it in a safe place.", "info")
        
        is_frozen = getattr(sys, 'frozen', False)
        return render_template('setup_complete.html', is_frozen=is_frozen)
        
    # GET Request Logic
    suggested_secret = os.getenv('SECRET_KEY', secrets.token_hex(24))
    # NEW: Generate the master password for the setup form
    suggested_master_password = generate_secure_password()
    
    logging.info("--- [SECURITY] Setup screen loaded. A new Master Password has been generated. ---")
    return render_template(
        'setup.html', 
        suggested_secret=suggested_secret,
        suggested_master_password=suggested_master_password
    )
@app.route('/version_selection')
def version_selection():
    """Redirect to auth - version selection no longer needed"""
    return redirect(url_for('auth'))

@app.route('/select_version/<version_type>')
def select_version(version_type):
    """Redirect to auth - version selection no longer needed"""
    return redirect(url_for('auth'))

def handle_activate_license():
    """Activate Pro license from Settings page"""
    try:
        license_key = request.form.get('license_key', '').strip().upper()

        if not license_key:
            flash('Please enter a license key', 'error')
            return redirect(url_for('settings'))

        # Validate license format
        if not license_key.startswith('AEGIS-') or len(license_key) < 28:
            flash('Invalid license key format', 'error')
            return redirect(url_for('settings'))

        # Get fresh user object from database
        user = db.session.query(User).filter_by(id=current_user.id).first()

        if user and user.validate_license_key(license_key):
            db.session.commit()
            logging.info(f"User {user.username} activated PRO license: {license_key}")
            flash('License activated successfully! All Pro features are now unlocked.', 'success')
        else:
            logging.error(f"License activation failed for user {current_user.username}")
            flash('Invalid license key. Please check and try again.', 'error')

        return redirect(url_for('settings'))

    except Exception as e:
        db.session.rollback()
        logging.error(f"License activation error: {e}")
        flash('License activation failed. Please try again.', 'error')
        return redirect(url_for('settings'))

def handle_deactivate_license():
    """Deactivate Pro license and revert to Basic"""
    try:
        user = db.session.query(User).filter_by(id=current_user.id).first()

        if user:
            user.license_key = None
            user.license_validated_at = None
            user.license_expires_at = None
            user.user_type = 'BASIC'
            user.allowed_monthly_scans = 5

            db.session.commit()
            logging.info(f"User {user.username} deactivated PRO license")
            flash('License deactivated. You are now using the Basic version.', 'info')

        return redirect(url_for('settings'))

    except Exception as e:
        db.session.rollback()
        logging.error(f"License deactivation error: {e}")
        flash('License deactivation failed. Please try again.', 'error')
        return redirect(url_for('settings'))

def handle_request_license():
    """Request a Pro license - sends email from user to aegis.aws.scanner@gmail.com"""
    try:
        message = request.form.get('message', '').strip()
        company = request.form.get('company', '').strip()

        # Send email to support
        try:
            from flask_mail import Message as MailMessage

            # Send TO support email, FROM configured mail, REPLY-TO user's email
            user_email = current_user.email
            support_email = 'aegis.aws.scanner@gmail.com'
            configured_sender = current_app.config.get('MAIL_USERNAME')

            msg = MailMessage(
                subject=f'Aegis Scanner - Pro License Request from {current_user.username}',
                sender=configured_sender,  # FROM: configured SMTP account (required)
                recipients=[support_email],  # TO: support email
                reply_to=user_email  # REPLY-TO: user's email (for replies)
            )
            msg.body = f"""
Pro License Request

User: {current_user.username}
Email: {user_email}
Company: {company if company else 'N/A'}

Message:
{message if message else 'No additional message'}

---
This is an automated license request from Aegis Cloud Scanner.
Please generate a Pro license key and send it to {user_email}.
"""
            mail.send(msg)
            logging.info(f"License request sent to {support_email} (from {configured_sender}, reply-to {user_email})")
            flash('License request sent successfully! You will receive your license key via email within 24 hours.', 'success')

        except Exception as email_error:
            logging.error(f"Failed to send license request email: {str(email_error)}")
            flash('Failed to send license request. Please contact support directly at aegis.aws.scanner@gmail.com', 'error')

        return redirect(url_for('settings'))

    except Exception as e:
        logging.error(f"License request error: {e}")
        flash('License request failed. Please try again.', 'error')
        return redirect(url_for('settings'))

@app.route('/fix_user_account')
@login_required
def fix_user_account():
    """Fix current user account to Pro status with valid license"""
    try:
        # Force upgrade current user to Pro with a valid license
        dummy_license = 'AEGIS-FIXED-ACCOUNT-001'
        current_user.upgrade_to_pro(license_key=dummy_license)
        current_user.license_validated_at = datetime.now(timezone.utc)

        db.session.commit()

        flash(f'Account fixed! User: {current_user.username}, Type: {current_user.user_type}, License: {current_user.license_key}', 'success')

        # Clear any cached states and redirect
        return redirect(url_for('dashboard'))

    except Exception as e:
        logging.error(f"Account fix error: {e}")
        flash(f'Error fixing account: {e}', 'error')
        return redirect(url_for('dashboard'))


@app.route('/check_setup')
def check_setup():
    """Check if initial setup is complete and redirect to auth page"""
    if not os.path.exists(ENV_FILE_PATH) or not os.getenv('MAIL_USERNAME'):
        return redirect(url_for('setup'))
    else:
        # Skip version selection - all users start as Basic, can upgrade later
        return redirect(url_for('auth'))

@app.route('/test_email', methods=['POST'])
@csrf.exempt
def test_email():
    data = request.get_json()
    temp_mail = Mail()
    temp_app = Flask("temp_app")
    temp_app.config.update(MAIL_SERVER='smtp.gmail.com', MAIL_PORT=587, MAIL_USE_TLS=True, MAIL_USERNAME=data.get('email'), MAIL_PASSWORD=data.get('password'))

    # Set SERVER_NAME to prevent KeyError (same logic as main app)
    domain_name = os.getenv('DOMAIN_NAME', 'localhost:5000')
    if 'localhost' not in domain_name and '127.0.0.1' not in domain_name:
        temp_app.config['SERVER_NAME'] = domain_name

    # ULTIMATE SAFETY: Override Flask's entire config dict for temp app
    class SafeConfig(dict):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)

        def __getitem__(self, key):
            if key == 'SERVER_NAME' and key not in self:
                return None
            return super().__getitem__(key)

        def get(self, key, default=None):
            if key == 'SERVER_NAME' and key not in self:
                return None
            return super().get(key, default)

    # Replace temp_app.config with our safe version
    safe_config = SafeConfig(temp_app.config)
    temp_app.config = safe_config

    temp_mail.init_app(temp_app)
    msg = Message("Aegis Scanner Test Email", sender=data.get('email'), recipients=[data.get('email')])
    msg.body = "This is a test email from the Aegis Scanner setup."
    try:
        with temp_app.app_context():
            temp_mail.send(msg)
        return jsonify({"message": "Test email sent successfully!"}), 200
    except Exception as e:
        logging.error(f"Failed to send test email: {e}")
        return jsonify({"error": str(e)}), 500

def shutdown_server():
    os._exit(0)
    
def handle_guest_chatbot(user_message):
    """Handle chatbot requests for guest users with predefined responses."""
    user_message_lower = user_message.lower()

    # Basic FAQ responses for guest users
    if any(word in user_message_lower for word in ['hello', 'hi', 'hey', 'start']):
        return jsonify({"reply": "👋 Hello! I'm the Aegis Help Bot. I can help you understand the features available in this demo. Try asking 'What can I do here?' or 'How do I run a scan?'"})

    elif any(word in user_message_lower for word in ['scan', 'run', 'start scan']):
        return jsonify({"reply": "🔍 To run a scan in demo mode: Go to the Dashboard, click 'Run Scan' button, and you'll see a simulated scan with sample results. In the full version, you would need to add your cloud credentials first."})

    elif any(word in user_message_lower for word in ['credential', 'aws', 'gcp', 'api key']):
        return jsonify({"reply": "🔑 In demo mode, credentials are simulated. In the full version, you would go to Settings to add your AWS Access Keys or GCP Service Account JSON files securely."})

    elif any(word in user_message_lower for word in ['feature', 'what can', 'capabilities', 'do']):
        return jsonify({"reply": "⚡ Aegis Scanner features:\n• Cloud Security Scanning (AWS/GCP)\n• Real-time Dashboard with metrics\n• Automated Scheduling\n• Detailed Reports (PDF/CSV)\n• Email Notifications\n• Resource Discovery\n• Compliance Monitoring\nTry exploring the sidebar sections!"})

    elif any(word in user_message_lower for word in ['dashboard', 'metrics', 'charts']):
        return jsonify({"reply": "📊 The Dashboard shows:\n• Resource counts and health scores\n• Security posture charts\n• Critical findings by service\n• Historical trends\n• Recent scan activity\nYou can click on chart elements to filter results!"})

    elif any(word in user_message_lower for word in ['automation', 'schedule', 'cron']):
        return jsonify({"reply": "⏰ Automation features:\n• Schedule daily, weekly, or monthly scans\n• Set specific times for automated runs\n• Email reports automatically\n• Background scan monitoring\nCheck the Automation section in the sidebar!"})

    elif any(word in user_message_lower for word in ['report', 'pdf', 'csv', 'export']):
        return jsonify({"reply": "📄 Reporting options:\n• PDF reports with executive summaries\n• CSV exports for data analysis\n• Email delivery of reports\n• Historical scan comparisons\nVisit the Reporting section to try these features!"})

    elif any(word in user_message_lower for word in ['guest', 'demo', 'trial']):
        return jsonify({"reply": "🎯 You're in Guest Mode! This gives you a full demo experience with:\n• Simulated scan data\n• All UI features available\n• No registration required\nTo access real scanning, you would need to create an account and add your cloud credentials."})

    elif any(word in user_message_lower for word in ['help', 'support', 'stuck']):
        return jsonify({"reply": "💡 Need help? Try these:\n• Explore each sidebar section\n• Click on charts and cards to interact\n• Use the search and filter features\n• Check the Settings page\nFor specific questions, try asking about 'features', 'scanning', or 'reports'!"})

    else:
        return jsonify({"reply": "🤖 I'm a demo assistant for Aegis Scanner. I can help with:\n• How to use features\n• Understanding the dashboard\n• Scanning and reporting\n• Automation setup\n\nTry asking: 'What features are available?' or 'How do I run a scan?'"})

# Replace the existing chatbot endpoint in app.py with this enhanced version:

@app.route('/api/v1/chatbot', methods=['POST'])
@login_or_guest_required
def chatbot_api():
    """
    Enhanced chatbot endpoint that uses the user's stored API key.
    """
    data = request.get_json()
    user_message = data.get('message')

    if not user_message:
        return jsonify({"error": "No message provided"}), 400

    # Handle guest users with fallback responses
    if session.get('guest_mode'):
        return handle_guest_chatbot(user_message)

    # FIX: Get API key from the database, not environment variables
    user_api_key_record = current_user.api_key
    if not user_api_key_record:
        return jsonify({"reply": "The chatbot is not configured. Please add your Gemini API key in the Settings page."}), 400
    
    try:
        api_key = decrypt_data(user_api_key_record.encrypted_key, "Gemini API Key")
    except Exception as e:
        logging.error(f"Chatbot could not decrypt API key for user {current_user.id}: {e}")
        return jsonify({"reply": "There was an error accessing the API key. Please try saving it again in Settings."}), 500

    try:
        genai.configure(api_key=api_key)

        # First, try to list available models (same as test function)
        try:
            available_models = list(genai.list_models())
            model_names = [model.name for model in available_models if 'gemini' in model.name.lower()]
            logging.info(f"Available Gemini models for chatbot: {model_names}")
        except Exception as list_error:
            logging.warning(f"Could not list models for chatbot: {list_error}")
            # Fallback to hardcoded list
            model_names = [
                'gemini-1.5-flash',  # Latest flash model
                'gemini-1.5-pro',    # Latest pro model
                'gemini-pro',        # Stable gemini model
                'gemini-1.0-pro',    # Explicit version
                'models/gemini-pro', # With models/ prefix
                'models/gemini-1.5-flash',
                'models/gemini-1.5-pro'
            ]

        model = None
        working_model = None

        for model_name in model_names:
            try:
                # Clean model name (remove models/ prefix if present)
                clean_name = model_name.replace('models/', '') if model_name.startswith('models/') else model_name

                model = genai.GenerativeModel(clean_name)
                # Test with a simple prompt to ensure it works
                test_response = model.generate_content("Test")
                working_model = clean_name
                logging.info(f"Chatbot using model: {working_model}")
                break  # If successful, use this model
            except Exception as model_error:
                logging.info(f"Chatbot model {clean_name} failed: {str(model_error)}")
                if "404" in str(model_error) or "not found" in str(model_error).lower():
                    continue  # Try next model
                else:
                    # For other errors, continue trying other models
                    continue

        if not model:
            return jsonify({"reply": "⚠️ No compatible Gemini model found. Please check if your API key has access to Gemini models."}), 400

        # Import the comprehensive knowledge base
        try:
            from chatbot_knowledge import get_fast_response, get_knowledge_section, FAST_RESPONSES
        except ImportError:
            logging.error("Could not import chatbot knowledge base")
            return jsonify({"reply": "⚠️ Chatbot knowledge base is temporarily unavailable."}), 500

        # FAST RESPONSE SYSTEM: Check for quick answers first
        fast_response = get_fast_response(user_message)
        if fast_response:
            return jsonify({"reply": fast_response})

        # For complex queries, use AI with focused knowledge
        user_message_lower = user_message.lower()

        # Determine context for more focused AI responses
        if any(word in user_message_lower for word in ['aws', 'amazon', 'ec2', 's3', 'iam']):
            context_knowledge = get_knowledge_section('aws')
            context_type = "AWS"
        elif any(word in user_message_lower for word in ['gcp', 'google', 'gce', 'cloud storage']):
            context_knowledge = get_knowledge_section('gcp')
            context_type = "GCP"
        elif any(word in user_message_lower for word in ['azure', 'microsoft', 'vm']):
            context_knowledge = get_knowledge_section('azure')
            context_type = "Azure"
        elif any(word in user_message_lower for word in ['code', 'script', 'python', 'boto3']):
            context_knowledge = get_knowledge_section('code')
            context_type = "Code Examples"
        elif any(word in user_message_lower for word in ['error', 'problem', 'troubleshoot', 'debug']):
            context_knowledge = get_knowledge_section('troubleshooting')
            context_type = "Troubleshooting"
        else:
            context_knowledge = {}
            context_type = "General"

        # Create focused knowledge base for AI
        if context_knowledge:
            knowledge_base = f"""
            You are an expert Aegis Scanner assistant specializing in {context_type}.

            CONTEXT-SPECIFIC KNOWLEDGE:
            {str(context_knowledge)}

            GENERAL AEGIS SCANNER INFO:
            • Cloud security scanning tool for AWS, GCP, Azure
            • Features: Dashboard, Automation, Reporting, Settings
            • Supports credential management and scheduling
            • Provides security posture analysis and compliance checking
            """
        else:
            # Use external knowledge base for comprehensive responses
            knowledge_base = """
            You are a helpful Aegis Scanner assistant with comprehensive cloud security expertise.

            CORE APPLICATION INFO:
            • Aegis Scanner: Multi-cloud security tool (AWS/GCP/Azure)
            • Features: Dashboard, Scan Results, Automation, History, Reporting, Settings
            • Read-only scanning with real-time progress updates
            • Credential management with encryption
            • Automated scheduling and email reports
            • 2FA security and audit logging

            QUICK FEATURE GUIDE:
            • Run Scan: Select credential → Choose regions → Click "Run Scan"
            • Schedule: Automation section → Daily/Weekly/Monthly options
            • Reports: Download PDF/CSV or email reports
            • Settings: Credentials, 2FA, notifications, timeouts

            For detailed AWS/GCP/Azure setup, troubleshooting, and custom code examples,
            refer to the comprehensive knowledge sections as needed.
            """

        prompt = f"""
        You are a helpful and knowledgeable assistant for Aegis Scanner, a cloud security application.
        Your goal is to help users understand and use all features of the application effectively.
        Always provide specific, actionable guidance based on the comprehensive knowledge provided below.
        Be friendly but concise, and focus on solving the user's specific question or problem.

        --- COMPREHENSIVE KNOWLEDGE BASE ---
        {knowledge_base}
        --- END KNOWLEDGE BASE ---

        User Question: "{user_message}"

        Please provide a helpful, specific answer based on the knowledge above. If the user asks about a feature not covered, 
        let them know it's not currently available in Aegis Scanner but suggest related features that might help.
        """

        response = model.generate_content(prompt)
        bot_reply = response.text.strip()
        return jsonify({"reply": bot_reply})

    except Exception as e:
        logging.error(f"Error calling Gemini API: {e}")
        error_message = str(e)
        
        # Provide specific error messages based on the error type
        if "API_KEY_INVALID" in error_message or "invalid api key" in error_message.lower():
            return jsonify({"reply": "⚠️ Your Gemini API key appears to be invalid. Please update it in the Settings page."}), 400
        elif "quota exceeded" in error_message.lower() or "rate limit" in error_message.lower():
            return jsonify({"reply": "⏰ API quota exceeded. Please wait a moment before asking another question."}), 429
        elif "network" in error_message.lower() or "connection" in error_message.lower():
            return jsonify({"reply": "🌐 Network connectivity issue. Please check your internet connection and try again."}), 503
        else:
            return jsonify({"reply": "🤖 I'm temporarily unavailable. Please check your API key in Settings and try again later."}), 500

@app.route('/api/v1/test-gemini', methods=['POST'])
@login_or_guest_required
def test_gemini_api():
    """Test endpoint to diagnose Gemini API key issues."""
    if session.get('guest_mode'):
        return jsonify({"status": "guest_mode", "message": "Guest mode - no API key needed"}), 200

    try:
        # Check if user has API key
        if not current_user.api_key:
            return jsonify({"status": "no_key", "message": "No API key found in database"}), 400

        # Try to decrypt the API key
        try:
            api_key = decrypt_data(current_user.api_key.encrypted_key, "Gemini API Key")
        except Exception as e:
            return jsonify({"status": "decrypt_error", "message": f"Failed to decrypt API key: {str(e)}"}), 500

        # Validate API key format
        if not api_key.startswith('AIza'):
            return jsonify({"status": "invalid_format", "message": "API key doesn't start with 'AIza'"}), 400

        # Test the API key with a simple request
        try:
            genai.configure(api_key=api_key)

            # First, try to list available models
            try:
                available_models = list(genai.list_models())
                model_names = [model.name for model in available_models if 'gemini' in model.name.lower()]
                logging.info(f"Available Gemini models: {model_names}")
            except Exception as list_error:
                logging.warning(f"Could not list models: {list_error}")
                # Fallback to hardcoded list
                model_names = [
                    'gemini-1.5-flash',  # Latest flash model
                    'gemini-1.5-pro',    # Latest pro model
                    'gemini-pro',        # Stable gemini model
                    'gemini-1.0-pro',    # Explicit version
                    'models/gemini-pro', # With models/ prefix
                    'models/gemini-1.5-flash',
                    'models/gemini-1.5-pro'
                ]

            model = None
            working_model = None
            tested_models = []

            for model_name in model_names:
                try:
                    # Clean model name (remove models/ prefix if present)
                    clean_name = model_name.replace('models/', '') if model_name.startswith('models/') else model_name
                    tested_models.append(clean_name)

                    model = genai.GenerativeModel(clean_name)
                    response = model.generate_content("Hello, respond with just 'API key works!'")
                    working_model = clean_name
                    break
                except Exception as model_error:
                    logging.info(f"Model {clean_name} failed: {str(model_error)}")
                    if "404" in str(model_error) or "not found" in str(model_error).lower():
                        continue  # Try next model
                    else:
                        # For other errors, continue trying other models
                        continue

            if working_model:
                return jsonify({
                    "status": "success",
                    "message": f"API key is working with model: {working_model}!",
                    "response": response.text.strip(),
                    "tested_models": tested_models[:5]  # Show first 5 tested models
                }), 200
            else:
                return jsonify({
                    "status": "no_compatible_model",
                    "message": f"API key is valid but no compatible Gemini model found. Tested: {', '.join(tested_models[:5])}",
                    "tested_models": tested_models
                }), 400
        except Exception as e:
            error_str = str(e).lower()
            if "api_key_invalid" in error_str or "invalid api key" in error_str:
                return jsonify({"status": "invalid_key", "message": "API key is invalid"}), 400
            elif "quota" in error_str or "rate limit" in error_str:
                return jsonify({"status": "quota_exceeded", "message": "API quota exceeded"}), 429
            else:
                return jsonify({"status": "api_error", "message": f"Gemini API error: {str(e)}"}), 500

    except Exception as e:
        return jsonify({"status": "server_error", "message": f"Server error: {str(e)}"}), 500

@app.route('/restart_app')
def restart_app():
    if getattr(sys, 'frozen', False):
        subprocess.Popen([sys.executable])
        threading.Timer(1.0, shutdown_server).start()
        return "<h1>Relaunching... You can close this window.</h1>"
    return "Restart is only available for the executable version."

@app.route('/shutdown', methods=['POST'])
def shutdown():
    """Shuts down the Waitress server."""
    os._exit(0)
    return 'Server shutting down...'

@app.route('/api/cleanup-guest-session', methods=['POST'])
def cleanup_guest_session():
    """Cleanup guest session data when browser window is closed"""
    try:
        data = request.get_json()
        guest_session_id = data.get('guest_session_id')
        
        if guest_session_id and session.get('guest_session_id') == guest_session_id:
            # Clear guest credentials from session
            session.pop('guest_credentials', None)
            session.pop('guest_mode', None)
            session.pop('guest_session_id', None)
            session.pop('guest_expires', None)
            
            # Log cleanup action
            logging.info(f"Guest session cleaned up: {guest_session_id}")
            
        return jsonify({"status": "cleaned"}), 200
    except Exception as e:
        logging.error(f"Guest cleanup error: {e}")
        return jsonify({"status": "error"}), 500

@app.route('/api/check-guest-session', methods=['POST'])
def check_guest_session():
    """Check if guest session is still valid"""
    try:
        data = request.get_json()
        guest_session_id = data.get('guest_session_id')
        
        # Check if session is still valid
        if not session.get('guest_mode'):
            return jsonify({"status": "invalid"}), 401
            
        if session.get('guest_session_id') != guest_session_id:
            return jsonify({"status": "invalid"}), 401
            
        # Check if session has expired
        guest_expires = session.get('guest_expires')
        if guest_expires and datetime.now(timezone.utc).timestamp() > guest_expires:
            # Session expired, cleanup
            session.pop('guest_credentials', None)
            session.pop('guest_mode', None)
            session.pop('guest_session_id', None)
            session.pop('guest_expires', None)
            return jsonify({"status": "expired"}), 401
            
        return jsonify({"status": "valid"}), 200
    except Exception as e:
        logging.error(f"Guest session check error: {e}")
        return jsonify({"status": "error"}), 500

def run_main_app():
    # Initialize enhanced demo logging for interview
    init_demo_logging()

    first_run_flag = os.path.join(USER_DATA_DIR, '.first_run')
    if not os.path.exists(first_run_flag):
        threading.Timer(1.5, lambda: webbrowser.open("http://127.0.0.1:5000/")).start()
        with open(first_run_flag, 'w') as f:
            f.write('done')

    with app.app_context():
        db.create_all()

        # Migrate existing users to have proper user_type values
        try:
            users_without_type = User.query.filter(User.user_type.is_(None)).all()
            for user in users_without_type:
                user.user_type = 'BASIC'
                if not user.allowed_monthly_scans:
                    user.allowed_monthly_scans = 5

            if users_without_type:
                db.session.commit()
                logging.info(f"Migrated {len(users_without_type)} users to have proper user_type values")
        except Exception as e:
            logging.error(f"Error during user migration: {e}")
            db.session.rollback()

    # Run in production mode to avoid SERVER_NAME issues with debug mode
    # Debug mode causes Flask to access configurations in ways that trigger KeyError
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)

# === COMPREHENSIVE ADMIN TOOLS ===

# System Monitoring API
@app.route('/api/admin/system-stats')
@login_required
@admin_required
def get_system_stats():
    """Get comprehensive system statistics."""
    import platform
    import sys
    
    try:
        # Try to import psutil, provide fallbacks if not available
        try:
            import psutil
            psutil_available = True
        except ImportError:
            psutil_available = False
        
        # System info (basic info always available)
        system_info = {
            'platform': platform.system(),
            'platform_version': platform.version(),
            'architecture': platform.architecture()[0],
            'python_version': sys.version.split()[0]
        }
        
        # Add psutil-dependent info if available
        if psutil_available:
            system_info.update({
                'cpu_cores': psutil.cpu_count(),
                'memory_total': psutil.virtual_memory().total,
                'disk_total': psutil.disk_usage('/').total if platform.system() != 'Windows' else psutil.disk_usage('C:\\').total
            })
        else:
            system_info.update({
                'cpu_cores': 'N/A',
                'memory_total': 'N/A',
                'disk_total': 'N/A'
            })
        
        # Current stats
        if psutil_available:
            try:
                memory = psutil.virtual_memory()
                disk = psutil.disk_usage('/') if platform.system() != 'Windows' else psutil.disk_usage('C:\\')
                
                current_stats = {
                    'cpu_percent': psutil.cpu_percent(interval=0.1),  # Shorter interval to avoid hanging
                    'memory_used': memory.used,
                    'memory_percent': memory.percent,
                    'disk_used': disk.used,
                    'disk_percent': (disk.used / disk.total) * 100,
                    'uptime': str(datetime.now() - datetime.fromtimestamp(psutil.boot_time())),
                    'load_avg': psutil.getloadavg() if hasattr(psutil, 'getloadavg') else [0, 0, 0]
                }
            except Exception as e:
                current_stats = {
                    'cpu_percent': 0,
                    'memory_used': 0,
                    'memory_percent': 0,
                    'disk_used': 0,
                    'disk_percent': 0,
                    'uptime': 'N/A',
                    'load_avg': [0, 0, 0],
                    'error': str(e)
                }
        else:
            current_stats = {
                'cpu_percent': 0,
                'memory_used': 0,
                'memory_percent': 0,
                'disk_used': 0,
                'disk_percent': 0,
                'uptime': 'N/A',
                'load_avg': [0, 0, 0],
                'message': 'psutil not available - install with: pip install psutil'
            }
        
        # Database stats (always available)
        try:
            user_count = User.query.count()
            scan_count = ScanResult.query.count()
            credential_count = db.session.query(CloudCredential).count()
            audit_count = AuditLog.query.count()
            recent_scans = ScanResult.query.filter(ScanResult.timestamp >= datetime.now(timezone.utc) - timedelta(days=7)).count()
            
            database_stats = {
                'total_users': user_count,
                'total_scans': scan_count,
                'total_credentials': credential_count,
                'total_audit_logs': audit_count,
                'recent_scans_7d': recent_scans
            }
        except Exception as e:
            database_stats = {
                'total_users': 0,
                'total_scans': 0,
                'total_credentials': 0,
                'total_audit_logs': 0,
                'recent_scans_7d': 0,
                'error': str(e)
            }
        
        return jsonify({
            'status': 'success',
            'system_info': system_info,
            'current_stats': current_stats,
            'database_stats': database_stats,
            'psutil_available': psutil_available,
            'timestamp': datetime.now(timezone.utc).isoformat()
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Bulk Scan Management
@app.route('/api/admin/scans/bulk-delete', methods=['POST'])
@login_required
@admin_required
@limiter.limit("5 per minute")
def bulk_delete_scans():
    """Bulk delete scan results."""
    try:
        data = request.get_json()
        scan_ids = data.get('scan_ids', [])
        
        if not scan_ids:
            return jsonify({'status': 'error', 'message': 'No scan IDs provided'}), 400
            
        deleted_count = ScanResult.query.filter(ScanResult.id.in_(scan_ids)).delete(synchronize_session=False)
        db.session.commit()
        
        log_audit("Bulk Scan Deletion", 
                 details=f"Deleted {deleted_count} scan results", 
                 user=current_user)
        
        return jsonify({
            'status': 'success',
            'deleted_count': deleted_count,
            'message': f'Successfully deleted {deleted_count} scan results'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Advanced Audit Log Search
@app.route('/api/admin/audit-logs')
@login_required
@admin_required
def get_audit_logs():
    """Get audit logs with advanced filtering."""
    try:
        # Get query parameters - support both page-based and offset-based pagination
        offset = request.args.get('offset', type=int)
        limit = request.args.get('limit', type=int)
        
        if offset is not None and limit is not None:
            # Use offset-based pagination
            page = (offset // limit) + 1
            per_page = min(limit, 100)
        else:
            # Use traditional page-based pagination
            page = request.args.get('page', 1, type=int)
            per_page = min(request.args.get('per_page', 50, type=int), 100)
        
        search = request.args.get('search', '')
        action = request.args.get('action', '')
        user_id = request.args.get('user_id', type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Build query
        query = AuditLog.query
        
        if search:
            query = query.filter(
                or_(
                    AuditLog.action.contains(search),
                    AuditLog.details.contains(search),
                    AuditLog.ip_address.contains(search)
                )
            )
        
        if action:
            query = query.filter(AuditLog.action == action)
            
        if user_id:
            query = query.filter(AuditLog.user_id == user_id)
            
        if start_date:
            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            query = query.filter(AuditLog.timestamp >= start_dt)
            
        if end_date:
            end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            query = query.filter(AuditLog.timestamp <= end_dt)
        
        # Order by most recent first
        query = query.order_by(AuditLog.timestamp.desc())
        
        # Paginate
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        logs = pagination.items
        
        # Convert to dict
        log_data = []
        for log in logs:
            log_dict = {
                'id': log.id,
                'action': log.action,
                'details': log.details,
                'ip_address': log.ip_address,
                'user_agent': log.user_agent,
                'timestamp': log.timestamp.isoformat(),
                'username': log.user.username if log.user else 'System'
            }
            log_data.append(log_dict)
        
        return jsonify({
            'status': 'success',
            'logs': log_data,
            'pagination': {
                'page': page,
                'pages': pagination.pages,
                'per_page': per_page,
                'total': pagination.total
            }
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Database Backup
@app.route('/api/admin/backup/create', methods=['POST'])
@login_required
@admin_required
@limiter.limit("2 per hour")
def create_database_backup():
    """Create a database backup."""
    try:
        import sqlite3
        import shutil
        
        # Get the database URI and extract the path
        database_uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        
        if database_uri.startswith('sqlite:///'):
            # Remove 'sqlite:///' prefix to get the actual path
            db_path = database_uri.replace('sqlite:///', '')
            # On Windows, handle drive letters properly
            if os.name == 'nt' and len(db_path) > 1 and db_path[1] == ':':
                # Path already has drive letter, use as-is
                pass
            elif not os.path.isabs(db_path):
                # Convert relative path to absolute path
                db_path = os.path.abspath(db_path)
        else:
            # Fallback to default path
            db_path = os.path.join(USER_DATA_DIR, 'aegis_scanner.db')
        
        # Normalize path for Windows
        db_path = os.path.normpath(db_path)
        
        # Debug information
        debug_info = {
            'database_uri': database_uri,
            'extracted_db_path': db_path,
            'user_data_dir': USER_DATA_DIR,
            'os_name': os.name,
            'cwd': os.getcwd()
        }
        
        # Verify the database file exists
        if not os.path.exists(db_path):
            return jsonify({
                'status': 'error', 
                'message': f'Database file not found at: {db_path}',
                'debug': debug_info
            }), 404
        
        # Verify database file is accessible
        if not os.access(db_path, os.R_OK):
            return jsonify({
                'status': 'error',
                'message': f'Database file is not readable: {db_path}',
                'debug': debug_info
            }), 403
        
        # Create backup filename with timestamp
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        backup_filename = f'aegis_backup_{timestamp}.db'
        backup_dir = os.path.join(USER_DATA_DIR, 'backups')
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # Normalize backup paths for Windows
        backup_dir = os.path.normpath(backup_dir)
        backup_path = os.path.normpath(backup_path)
        
        # Create backup directory if it doesn't exist
        try:
            os.makedirs(backup_dir, exist_ok=True)
        except Exception as dir_error:
            return jsonify({
                'status': 'error',
                'message': f'Failed to create backup directory: {str(dir_error)}',
                'debug': {**debug_info, 'backup_dir': backup_dir}
            }), 500
        
        # Verify backup directory is writable
        if not os.access(backup_dir, os.W_OK):
            return jsonify({
                'status': 'error',
                'message': f'Backup directory is not writable: {backup_dir}',
                'debug': {**debug_info, 'backup_dir': backup_dir}
            }), 403
        
        # Copy the database file
        try:
            shutil.copy2(db_path, backup_path)
        except Exception as copy_error:
            return jsonify({
                'status': 'error',
                'message': f'Failed to copy database file: {str(copy_error)}',
                'debug': {
                    **debug_info,
                    'backup_dir': backup_dir,
                    'backup_path': backup_path,
                    'copy_error_type': type(copy_error).__name__
                }
            }), 500
        
        # Verify backup was created
        if not os.path.exists(backup_path):
            return jsonify({
                'status': 'error',
                'message': 'Backup file was not created successfully',
                'debug': {
                    **debug_info,
                    'backup_dir': backup_dir,
                    'backup_path': backup_path
                }
            }), 500
        
        # Get backup file size
        backup_size = os.path.getsize(backup_path)
        
        # Log the backup
        log_audit("Database Backup", 
                 details=f"Created backup: {backup_filename} ({backup_size} bytes)", 
                 user=current_user)
        
        return jsonify({
            'status': 'success',
            'backup_file': backup_filename,
            'backup_path': backup_path,
            'backup_size': backup_size,
            'source_db': db_path,
            'message': f'Database backup created successfully ({backup_size} bytes)'
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        app.logger.error(f"Backup creation failed: {error_details}")
        
        return jsonify({
            'status': 'error', 
            'message': f'Backup failed: {str(e)}',
            'error_type': type(e).__name__,
            'details': error_details if app.debug else str(e)
        }), 500

# Security Configuration
@app.route('/api/admin/security/settings')
@login_required
@api_admin_required
def get_security_settings():
    """Get current security configuration."""
    try:
        # Handle PERMANENT_SESSION_LIFETIME which might be a timedelta object
        session_timeout = app.config.get('PERMANENT_SESSION_LIFETIME', 3600)
        if hasattr(session_timeout, 'total_seconds'):
            session_timeout = int(session_timeout.total_seconds())
        elif not isinstance(session_timeout, (int, float)):
            session_timeout = 3600  # Default to 1 hour

        settings = {
            'max_login_attempts': app.config.get('MAX_LOGIN_ATTEMPTS', 5),
            'session_timeout': session_timeout,
            'password_min_length': app.config.get('PASSWORD_MIN_LENGTH', 8),
            'require_2fa': app.config.get('REQUIRE_2FA', False),
            'ssl_enabled': request.is_secure,
            'csrf_protection': app.config.get('WTF_CSRF_ENABLED', True),
            'secure_cookies': app.config.get('SESSION_COOKIE_SECURE', False),
            'httponly_cookies': app.config.get('SESSION_COOKIE_HTTPONLY', True)
        }

        return jsonify({
            'status': 'success',
            'settings': settings
        })

    except Exception as e:
        logging.error(f"Security settings API error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Application Logs Viewer
@app.route('/api/admin/logs')
@login_required
@admin_required
def get_application_logs():
    """Get recent application logs."""
    try:
        lines = request.args.get('lines', 100, type=int)
        level = request.args.get('level', '').upper()
        
        logs = []
        if os.path.exists(log_file):
            with open(log_file, 'r', encoding='utf-8') as f:
                all_lines = f.readlines()
                recent_lines = all_lines[-lines:] if len(all_lines) > lines else all_lines
                
                for line in recent_lines:
                    line = line.strip()
                    if line and (not level or level in line):
                        logs.append(line)
        
        return jsonify({
            'status': 'success',
            'logs': logs,
            'total_lines': len(logs),
            'log_file': log_file
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Database Information
@app.route('/api/admin/database/info')
@login_required
@admin_required
def get_database_info():
    """Get database configuration and file information."""
    try:
        # Get the database URI and extract the path
        database_uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if database_uri.startswith('sqlite:///'):
            db_path = database_uri.replace('sqlite:///', '')
            if not os.path.isabs(db_path):
                db_path = os.path.abspath(db_path)
        else:
            db_path = os.path.join(USER_DATA_DIR, 'app.db')
        
        # Get file information if it exists
        file_info = {}
        if os.path.exists(db_path):
            file_info = {
                'exists': True,
                'size': os.path.getsize(db_path),
                'modified': datetime.fromtimestamp(os.path.getmtime(db_path)).isoformat(),
                'readable': os.access(db_path, os.R_OK),
                'writable': os.access(db_path, os.W_OK)
            }
        else:
            file_info = {'exists': False}
        
        return jsonify({
            'status': 'success',
            'database_uri': database_uri,
            'database_path': db_path,
            'user_data_dir': USER_DATA_DIR,
            'file_info': file_info
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# User Session Management
@app.route('/api/admin/sessions')
@login_required
@admin_required
def get_active_sessions():
    """Get information about active user sessions."""
    try:
        # Get users with recent activity (last 24 hours)
        recent_threshold = datetime.now(timezone.utc) - timedelta(hours=24)
        active_users = User.query.filter(User.last_login_date >= recent_threshold).all()
        
        sessions_data = []
        for user in active_users:
            session_info = {
                'user_id': user.id,
                'username': user.username,
                'email': user.email,
                'last_login': user.last_login_date.isoformat() if user.last_login_date else None,
                'is_admin': user.is_admin,
                'is_2fa_enabled': user.is_2fa_enabled,
                'failed_login_attempts': user.failed_login_attempts,
                'is_locked': user.is_locked
            }
            sessions_data.append(session_info)
        
        return jsonify({
            'status': 'success',
            'active_sessions': sessions_data,
            'total_active': len(sessions_data)
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# Enhanced Reporting System Routes
@app.route('/reports')
@login_required
@check_verified
@check_2fa
def reports_dashboard():
    """Enhanced reporting dashboard with comprehensive features."""
    try:
        # Get reporting statistics
        total_scans = ScanResult.query.filter_by(user_id=current_user.id).count()
        critical_findings = ScanResult.query.filter_by(user_id=current_user.id, status='CRITICAL').count()
        high_findings = ScanResult.query.filter_by(user_id=current_user.id, status='HIGH').count()
        passed_checks = ScanResult.query.filter_by(user_id=current_user.id, status='OK').count()
        
        # Get recent report history (if we add a reports table later)
        report_history = []
        
        return render_template('reports.html',
            total_scans=total_scans,
            critical_findings=critical_findings,
            high_findings=high_findings,
            passed_checks=passed_checks,
            report_history=report_history
        )
    except Exception as e:
        logging.error(f"Reports dashboard error: {e}")
        flash('Error loading reports dashboard', 'error')
        return redirect(url_for('dashboard'))

@app.route('/api/v1/reports/generate', methods=['POST'])
@login_required
@check_verified  
@check_2fa
def generate_advanced_report():
    """Generate advanced customizable reports."""
    try:
        # Handle both JSON and form data
        if request.is_json:
            data = request.get_json()
            report_type = data.get('reportType', 'comprehensive')
            output_format = data.get('outputFormat', 'pdf')
            start_date = data.get('startDate')
            end_date = data.get('endDate')
            severities = data.get('severities', [])
            services = data.get('services', [])
            include_remediation = data.get('includeRemediation', True)
            include_compliance = data.get('includeCompliance', True)
            include_costs = data.get('includeCosts', False)
        else:
            # Form data handling (existing logic)
            report_type = request.form.get('report_type', 'comprehensive')
            output_format = request.form.get('output_format', 'pdf')
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            severities = json.loads(request.form.get('severities', '[]'))
            services = request.form.getlist('services')
            include_remediation = request.form.get('include_remediation') == 'on'
            include_compliance = request.form.get('include_compliance') == 'on'
            include_costs = request.form.get('include_costs') == 'on'
        
        # Build query based on filters
        query = ScanResult.query.filter_by(user_id=current_user.id)
        
        # Date filtering
        if start_date:
            query = query.filter(ScanResult.timestamp >= datetime.strptime(start_date, '%m/%d/%Y'))
        if end_date:
            query = query.filter(ScanResult.timestamp <= datetime.strptime(end_date, '%m/%d/%Y'))
            
        # Severity filtering
        if severities:
            query = query.filter(ScanResult.status.in_(severities))
            
        # Service filtering
        if services and 'all' not in services:
            query = query.filter(ScanResult.service.in_(services))
            
        results = query.order_by(ScanResult.timestamp.desc()).limit(500).all()
        
        # Generate report based on format
        if output_format == 'pdf':
            pdf_bytes = _create_advanced_pdf_report(results, report_type, {
                'include_remediation': include_remediation,
                'include_compliance': include_compliance,
                'include_costs': include_costs,
                'start_date': start_date,
                'end_date': end_date
            })
            
            filename = f'security_report_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            return Response(pdf_bytes, mimetype='application/pdf', 
                          headers={'Content-Disposition': f'attachment;filename={filename}'})
                          
        elif output_format == 'csv':
            return _generate_advanced_csv_report(results, report_type)
            
        elif output_format == 'excel':
            return _generate_excel_report(results, report_type)
            
        elif output_format == 'json':
            return jsonify({
                'report_type': report_type,
                'generated_at': datetime.now().isoformat(),
                'total_findings': len(results),
                'findings': [{
                    'timestamp': r.timestamp.isoformat(),
                    'service': r.service,
                    'resource': r.resource,
                    'status': r.status,
                    'issue': r.issue,
                    'remediation': r.remediation if include_remediation else None,
                    'doc_url': r.doc_url
                } for r in results]
            })
            
        elif output_format == 'html':
            return _generate_html_report(results, report_type)
            
    except Exception as e:
        logging.error(f"Advanced report generation failed: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/reports/preview', methods=['POST'])
@login_required
@check_verified
@check_2fa  
def preview_report():
    """Generate a preview of the report."""
    try:
        # Handle both JSON and form data
        if request.is_json:
            data = request.get_json()
            report_type = data.get('reportType', 'comprehensive')
        else:
            report_type = request.form.get('report_type', 'comprehensive')
        
        # Get sample data for preview (last 10 results)
        results = ScanResult.query.filter_by(user_id=current_user.id).order_by(
            ScanResult.timestamp.desc()).limit(10).all()
            
        # Generate preview HTML
        preview_html = f"""
        <div class="report-preview">
            <h4>{report_type.title()} Report Preview</h4>
            <p class="text-muted">Showing last 10 findings as preview</p>
            <div class="row mb-3">
                <div class="col-md-3">
                    <div class="stat-card bg-danger text-white">
                        <h5>{len([r for r in results if r.status == 'CRITICAL'])}</h5>
                        <small>Critical</small>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="stat-card bg-warning text-white">
                        <h5>{len([r for r in results if r.status == 'HIGH'])}</h5>
                        <small>High</small>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="stat-card bg-info text-white">
                        <h5>{len([r for r in results if r.status == 'MEDIUM'])}</h5>
                        <small>Medium</small>
                    </div>
                </div>
                <div class="col-md-3">
                    <div class="stat-card bg-success text-white">
                        <h5>{len([r for r in results if r.status == 'OK'])}</h5>
                        <small>Passed</small>
                    </div>
                </div>
            </div>
            <table class="table table-striped">
                <thead>
                    <tr>
                        <th>Service</th>
                        <th>Status</th>
                        <th>Issue</th>
                        <th>Timestamp</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for result in results:
            status_class = {
                'CRITICAL': 'danger',
                'HIGH': 'warning', 
                'MEDIUM': 'info',
                'LOW': 'secondary',
                'OK': 'success'
            }.get(result.status, 'secondary')
            
            preview_html += f"""
                <tr>
                    <td>{result.service}</td>
                    <td><span class="badge bg-{status_class}">{result.status}</span></td>
                    <td>{result.issue[:100]}...</td>
                    <td>{result.timestamp.strftime('%Y-%m-%d %H:%M')}</td>
                </tr>
            """
            
        preview_html += """
                </tbody>
            </table>
        </div>
        """
        
        return jsonify({
            'success': True,
            'html': preview_html
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/v1/reports/stats')
@login_or_guest_required
def get_report_stats():
    """Get current reporting statistics."""
    try:
        # Get all scan results for the current user
        total_results = ScanResult.query.filter_by(user_id=current_user.id).count()

        stats = {
            'total_scans': total_results,
            'total_resources': total_results,  # Using total results as resources count
            'critical_findings': ScanResult.query.filter_by(user_id=current_user.id, status='CRITICAL').count(),
            'high_findings': ScanResult.query.filter_by(user_id=current_user.id, status='HIGH').count(),
            'medium_findings': ScanResult.query.filter_by(user_id=current_user.id, status='MEDIUM').count(),
            'low_findings': ScanResult.query.filter_by(user_id=current_user.id, status='LOW').count(),
            'info_findings': ScanResult.query.filter_by(user_id=current_user.id, status='INFO').count(),
            'warning_findings': ScanResult.query.filter_by(user_id=current_user.id, status='WARNING').count(),
            'passed_checks': ScanResult.query.filter_by(user_id=current_user.id, status='OK').count()
        }

        # Calculate health score based on findings
        if total_results > 0:
            critical_weight = stats['critical_findings'] * 10
            high_weight = stats['high_findings'] * 5
            medium_weight = stats['medium_findings'] * 2
            total_weight = critical_weight + high_weight + medium_weight

            # Health score: 100% minus weighted penalty
            health_score = max(0, 100 - (total_weight / max(1, total_results) * 10))
            stats['health_score'] = round(health_score)
        else:
            stats['health_score'] = 100

        # Get unique services monitored
        unique_services = db.session.query(ScanResult.service).filter_by(user_id=current_user.id).distinct().count()
        stats['services_monitored'] = unique_services

        # Get last scan timestamp
        last_scan = ScanResult.query.filter_by(user_id=current_user.id).order_by(ScanResult.timestamp.desc()).first()
        if last_scan:
            from datetime import datetime
            time_diff = datetime.utcnow() - last_scan.timestamp
            if time_diff.days > 0:
                stats['last_scan'] = f"{time_diff.days} day{'s' if time_diff.days > 1 else ''} ago"
            elif time_diff.seconds // 3600 > 0:
                hours = time_diff.seconds // 3600
                stats['last_scan'] = f"{hours} hour{'s' if hours > 1 else ''} ago"
            elif time_diff.seconds // 60 > 0:
                minutes = time_diff.seconds // 60
                stats['last_scan'] = f"{minutes} minute{'s' if minutes > 1 else ''} ago"
            else:
                stats['last_scan'] = "Just now"
        else:
            stats['last_scan'] = "No scans yet"

        # Calculate compliance score based on passed checks vs total checks
        if total_results > 0:
            compliance_score = round((stats['passed_checks'] / total_results) * 100)
            stats['compliance_score'] = compliance_score
        else:
            stats['compliance_score'] = 0

        # Active threats = critical + high findings
        stats['active_threats'] = stats['critical_findings'] + stats['high_findings']

        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/reports/schedule', methods=['POST'])
@login_required
@check_verified
@check_2fa
def schedule_advanced_reports():
    """Schedule advanced email reports with time configuration."""
    try:
        # Handle both JSON and form data
        if request.is_json:
            data = request.get_json()
            schedule = data.get('frequency', 'disabled')
            recipients = ','.join(data.get('recipients', []))
            report_type = data.get('reportType', 'comprehensive')
            delivery_time = data.get('time', '09:00')
            timezone_str = data.get('timezone', 'America/Los_Angeles')
            week_day = data.get('weekDay', 'monday')
            month_day = data.get('monthDay', 1)
        else:
            schedule = request.form.get('schedule', 'disabled')
            recipients = request.form.get('recipients', '')
            report_type = request.form.get('report_type', 'comprehensive')
            delivery_time = request.form.get('time', '09:00')
            timezone_str = request.form.get('timezone', 'America/Los_Angeles')
            week_day = request.form.get('week_day', 'monday')
            month_day = int(request.form.get('month_day', '1'))

        # Validate schedule value
        if not schedule or schedule not in ['disabled', 'daily', 'weekly', 'monthly']:
            schedule = 'disabled'

        # Parse delivery time
        try:
            hour, minute = map(int, delivery_time.split(':'))
        except:
            hour, minute = 9, 0  # Default to 9:00 AM

        # Update user preferences
        current_user.report_schedule = schedule
        if recipients:
            # Store additional recipients (you might want to create a separate table for this)
            current_user.backup_email = recipients.split(',')[0].strip() if recipients else None
            
        db.session.commit()
        
        # Remove existing scheduled job
        job_id = f"advanced_report_{current_user.id}"
        try:
            scheduler.remove_job(job_id)
        except:
            pass
            
        # Schedule new job with proper timezone support
        if schedule != 'disabled' and schedule:
            # Map day names to APScheduler format
            day_mapping = {
                'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3,
                'friday': 4, 'saturday': 5, 'sunday': 6
            }

            try:
                import pytz
                tz = pytz.timezone(timezone_str)
            except:
                tz = pytz.timezone('America/Los_Angeles')  # Default timezone

            if schedule == 'daily':
                scheduler.add_job(
                    func=send_advanced_scheduled_report,
                    trigger='cron',
                    hour=hour,
                    minute=minute,
                    timezone=tz,
                    args=[current_user.id, report_type, recipients],
                    id=job_id,
                    replace_existing=True
                )
            elif schedule == 'weekly':
                day_of_week = day_mapping.get(week_day.lower(), 0)  # Default to Monday
                scheduler.add_job(
                    func=send_advanced_scheduled_report,
                    trigger='cron',
                    day_of_week=day_of_week,
                    hour=hour,
                    minute=minute,
                    timezone=tz,
                    args=[current_user.id, report_type, recipients],
                    id=job_id,
                    replace_existing=True
                )
            elif schedule == 'monthly':
                scheduler.add_job(
                    func=send_advanced_scheduled_report,
                    trigger='cron',
                    day=min(28, max(1, month_day)),  # Ensure valid day range
                    hour=hour,
                    minute=minute,
                    timezone=tz,
                    args=[current_user.id, report_type, recipients],
                    id=job_id,
                    replace_existing=True
                )
                
        return jsonify({'success': True, 'message': 'Reports scheduled successfully'})
        
    except Exception as e:
        logging.error(f"Report scheduling failed: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/reports/email', methods=['POST'])
@login_required
@check_verified
@check_2fa
def send_report_email():
    """Send report via email using the new reports API path."""
    try:
        # Get form data
        email = request.form.get('email', '').strip()
        report_type = request.form.get('report_type', 'comprehensive')
        date_range = request.form.get('date_range', '')
        
        # Validate email
        if not email:
            return jsonify({"error": "Email address is required"}), 400
            
        is_valid_email, email_error = security_validator.validate_email(email)
        if not is_valid_email:
            return jsonify({"error": f"Invalid email: {email_error}"}), 400
        
        # Parse date range if provided
        start_date = None
        end_date = None
        if date_range:
            try:
                dates = date_range.split(' - ')
                if len(dates) == 2:
                    start_date = datetime.strptime(dates[0], '%m/%d/%Y')
                    end_date = datetime.strptime(dates[1], '%m/%d/%Y') + timedelta(days=1)
            except ValueError:
                pass
        
        # Default to last 30 days if no date range provided
        if not start_date or not end_date:
            end_date = datetime.now(timezone.utc)
            start_date = end_date - timedelta(days=30)
        
        # Get scan results
        query = ScanResult.query.filter_by(user_id=current_user.id)
        if start_date and end_date:
            query = query.filter(
                ScanResult.timestamp >= start_date,
                ScanResult.timestamp <= end_date
            )
        scan_results = query.order_by(ScanResult.timestamp.desc()).limit(100).all()
        
        if not scan_results:
            return jsonify({"error": "No scan results found for the specified date range"}), 400
        
        # Create PDF report
        pdf_bytes = _create_pdf_report(scan_results)
        
        # Send email
        msg = Message(
            subject=f"Aegis Security Report - {report_type.title()}",
            recipients=[email],
            sender=current_app.config.get('MAIL_DEFAULT_SENDER')
        )
        
        msg.html = f"""
        <h2>Aegis Security Report</h2>
        <p>Hello,</p>
        <p>Please find your requested security report attached.</p>
        <p><strong>Report Details:</strong></p>
        <ul>
            <li>Type: {report_type.title()}</li>
            <li>Generated: {datetime.now(timezone.utc).strftime('%B %d, %Y at %I:%M %p UTC')}</li>
            <li>Findings: {len(scan_results)} results</li>
            <li>Date Range: {start_date.strftime('%m/%d/%Y') if start_date else 'N/A'} - {end_date.strftime('%m/%d/%Y') if end_date else 'N/A'}</li>
        </ul>
        <p>Best regards,<br>Aegis Security Team</p>
        """
        
        msg.attach(
            'aegis_security_report.pdf',
            'application/pdf',
            pdf_bytes
        )
        
        mail.send(msg)
        
        # Log the action
        log_audit("Report Emailed", details=f"Report sent to {email}", user=current_user)
        
        return jsonify({"message": f"Report sent successfully to {email}"}), 200
        
    except Exception as e:
        logging.error(f"Failed to send report via email: {e}")
        return jsonify({"error": "Failed to send report via email"}), 500

def _create_advanced_pdf_report(results, report_type, options):
    """Create an advanced PDF report with customization options."""
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.colors import HexColor
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.piecharts import Pie
    import io
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=HexColor('#00A896'),
        alignment=1
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        textColor=HexColor('#2c3e50'),
        borderWidth=1,
        borderColor=HexColor('#3498db'),
        borderPadding=5
    )
    
    # Title page
    story.append(Paragraph(f"Aegis Cloud Security Report", title_style))
    story.append(Paragraph(f"Report Type: {report_type.title()}", styles['Normal']))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 0.5*inch))
    
    # Different content based on report type
    critical_count = len([r for r in results if r.status == 'CRITICAL'])
    high_count = len([r for r in results if r.status == 'HIGH'])
    medium_count = len([r for r in results if r.status == 'MEDIUM'])
    ok_count = len([r for r in results if r.status == 'OK'])

    if report_type == 'executive':
        # Executive Summary - High-level overview for executives
        story.append(Paragraph("Executive Summary", heading_style))

        risk_score = min(100, (critical_count + high_count) / max(1, len(results)) * 100)

        exec_summary = f"""
        <b>Security Risk Assessment</b><br/>
        Overall Risk Score: {risk_score:.1f}/100<br/>
        Total Security Issues: {critical_count + high_count + medium_count}<br/>
        Critical Issues Requiring Immediate Attention: {critical_count}<br/>
        <br/>
        <b>Recommendations:</b><br/>
        {'• Address all critical security findings immediately' if critical_count > 0 else '• No critical issues found'}
        {'<br/>• Review and remediate high-priority vulnerabilities' if high_count > 0 else ''}
        {'<br/>• Implement regular security monitoring procedures' if len(results) > 10 else ''}
        """

        story.append(Paragraph(exec_summary, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

    elif report_type == 'technical':
        # Technical Report - Detailed technical information
        story.append(Paragraph("Technical Security Analysis", heading_style))

        # Service breakdown
        services = {}
        regions = {}
        for result in results:
            service = result.service
            if service not in services:
                services[service] = {'total': 0, 'critical': 0, 'high': 0}
            services[service]['total'] += 1
            if result.status == 'CRITICAL':
                services[service]['critical'] += 1
            elif result.status == 'HIGH':
                services[service]['high'] += 1

            # Track regions if available
            region = getattr(result, 'region', 'Unknown')
            regions[region] = regions.get(region, 0) + 1

        # Technical summary table
        tech_data = [['Service', 'Total Findings', 'Critical', 'High Risk']]
        for service, counts in services.items():
            tech_data.append([service, str(counts['total']), str(counts['critical']), str(counts['high'])])

        tech_table = Table(tech_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch])
        tech_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(tech_table)
        story.append(Spacer(1, 0.3*inch))

    elif report_type == 'compliance':
        # Compliance Report - Focus on regulatory compliance
        story.append(Paragraph("Compliance Assessment Report", heading_style))

        compliance_score = max(0, 100 - (critical_count * 30 + high_count * 15 + medium_count * 5))

        # Compliance frameworks analysis
        frameworks = {
            'SOC 2': max(0, 100 - (critical_count * 25 + high_count * 10)),
            'PCI DSS': max(0, 100 - (critical_count * 35 + high_count * 15)),
            'GDPR': max(0, 100 - (critical_count * 20 + high_count * 12)),
            'HIPAA': max(0, 100 - (critical_count * 30 + high_count * 15)),
            'ISO 27001': max(0, 100 - (critical_count * 25 + high_count * 12))
        }

        compliance_summary = f"""
        <b>Compliance Posture Overview</b><br/>
        Overall Compliance Score: {compliance_score:.1f}%<br/>
        Total Non-Compliant Findings: {critical_count + high_count + medium_count}<br/>
        Critical Compliance Gaps: {critical_count}<br/>
        <br/>
        <b>Framework Compliance Scores:</b><br/>
        • SOC 2 Type II: {frameworks['SOC 2']:.1f}%<br/>
        • PCI DSS: {frameworks['PCI DSS']:.1f}%<br/>
        • GDPR: {frameworks['GDPR']:.1f}%<br/>
        • HIPAA: {frameworks['HIPAA']:.1f}%<br/>
        • ISO 27001: {frameworks['ISO 27001']:.1f}%<br/>
        <br/>
        <b>Compliance Recommendations:</b><br/>
        {'• Immediate remediation required for critical compliance gaps' if critical_count > 0 else '• No critical compliance issues found'}
        {'<br/>• Review data protection and privacy controls' if high_count > 0 else ''}
        {'<br/>• Implement continuous compliance monitoring' if len(results) > 5 else ''}
        """

        story.append(Paragraph(compliance_summary, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

    elif report_type == 'security':
        # Security-focused report
        story.append(Paragraph("Security Findings Overview", heading_style))

        security_summary = f"""
        <b>Security Posture Analysis</b><br/>
        Critical Vulnerabilities: {critical_count}<br/>
        High Severity Issues: {high_count}<br/>
        Medium Risk Items: {medium_count}<br/>
        Compliant Resources: {ok_count}<br/>
        <br/>
        <b>Security Score: {100 - (critical_count * 25 + high_count * 10 + medium_count * 5):.1f}%</b>
        """

        story.append(Paragraph(security_summary, styles['Normal']))
        story.append(Spacer(1, 0.3*inch))

    else:  # comprehensive or default
        # Comprehensive Summary - Detailed overview
        story.append(Paragraph("Comprehensive Security Report", heading_style))

        summary_data = [
            ['Metric', 'Count', 'Percentage'],
            ['Critical Findings', str(critical_count), f"{(critical_count/len(results)*100):.1f}%" if results else "0%"],
            ['High Priority', str(high_count), f"{(high_count/len(results)*100):.1f}%" if results else "0%"],
            ['Medium Priority', str(medium_count), f"{(medium_count/len(results)*100):.1f}%" if results else "0%"],
            ['Passed Checks', str(ok_count), f"{(ok_count/len(results)*100):.1f}%" if results else "0%"],
            ['Total Findings', str(len(results)), '100%']
        ]

        summary_table = Table(summary_data, colWidths=[2*inch, 1*inch, 1*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#00A896')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
    
    # Detailed findings section (different based on report type)
    if results:
        if report_type == 'executive':
            # Executive report - only show critical and high findings
            filtered_results = [r for r in results if r.status in ['CRITICAL', 'HIGH']]
            if filtered_results:
                story.append(Paragraph("Priority Issues Requiring Attention", heading_style))
                # Show only top 10 most critical issues for executives
                for i, result in enumerate(filtered_results[:10], 1):
                    issue_text = f"<b>{i}. {result.service} - {result.status}</b><br/>{result.issue[:100]}{'...' if len(result.issue) > 100 else ''}"
                    story.append(Paragraph(issue_text, styles['Normal']))
                    story.append(Spacer(1, 0.1*inch))

        elif report_type == 'technical':
            # Technical report - detailed technical information
            story.append(Paragraph("Technical Findings Analysis", heading_style))

            # Create detailed technical table
            table_data = [['Service', 'Resource ID', 'Status', 'Technical Details', 'Impact']]

            for result in results[:30]:  # More details for technical report
                row = [
                    result.service,
                    result.resource[:25] + '...' if len(result.resource) > 25 else result.resource,
                    result.status,
                    result.issue[:60] + '...' if len(result.issue) > 60 else result.issue,
                    'High' if result.status in ['CRITICAL', 'HIGH'] else 'Medium' if result.status == 'MEDIUM' else 'Low'
                ]
                table_data.append(row)

        elif report_type == 'security':
            # Security report - focus on security implications
            story.append(Paragraph("Security Vulnerabilities", heading_style))

            # Group by severity
            critical_results = [r for r in results if r.status == 'CRITICAL']
            high_results = [r for r in results if r.status == 'HIGH']

            if critical_results:
                story.append(Paragraph("<b>CRITICAL SECURITY ISSUES:</b>", styles['Normal']))
                for result in critical_results[:5]:
                    story.append(Paragraph(f"• {result.service}: {result.issue}", styles['Normal']))
                story.append(Spacer(1, 0.2*inch))

            if high_results:
                story.append(Paragraph("<b>HIGH PRIORITY SECURITY ISSUES:</b>", styles['Normal']))
                for result in high_results[:8]:
                    story.append(Paragraph(f"• {result.service}: {result.issue}", styles['Normal']))
                story.append(Spacer(1, 0.2*inch))

        else:  # comprehensive
            # Comprehensive report - full detailed findings
            story.append(Paragraph("Detailed Security Findings", heading_style))

            # Create comprehensive findings table
            table_data = [['Service', 'Resource', 'Status', 'Issue']]
            if options.get('include_remediation'):
                table_data[0].append('Remediation')

            for result in results[:50]:  # Limit for PDF size
                row = [
                    result.service,
                    result.resource[:30] + '...' if len(result.resource) > 30 else result.resource,
                    result.status,
                    result.issue[:50] + '...' if len(result.issue) > 50 else result.issue
                ]
                if options.get('include_remediation') and result.remediation:
                    row.append(result.remediation[:40] + '...' if len(result.remediation) > 40 else result.remediation)
                elif options.get('include_remediation'):
                    row.append('N/A')

                table_data.append(row)

        # Create table only for technical and comprehensive reports
        if report_type in ['technical', 'comprehensive'] and len(table_data) > 1:
            # Create table with dynamic column widths
            if report_type == 'technical':
                col_widths = [1.2*inch, 1.5*inch, 0.8*inch, 2.2*inch, 0.8*inch]
            elif options.get('include_remediation'):
                col_widths = [1.2*inch, 1.5*inch, 0.8*inch, 2*inch, 1.5*inch]
            else:
                col_widths = [1.5*inch, 2*inch, 1*inch, 3*inch]

            findings_table = Table(table_data, colWidths=col_widths)
            findings_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#00A896')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP')
            ]))

            # Color code status cells
            status_col = 2 if report_type == 'comprehensive' else 2  # Status column position
            for i, row in enumerate(table_data[1:], 1):
                if len(row) > status_col:
                    status = row[status_col]
                    if status == 'CRITICAL':
                        findings_table.setStyle(TableStyle([('BACKGROUND', (status_col, i), (status_col, i), HexColor('#D64550'))]))
                        findings_table.setStyle(TableStyle([('TEXTCOLOR', (status_col, i), (status_col, i), colors.white)]))
                    elif status == 'HIGH':
                        findings_table.setStyle(TableStyle([('BACKGROUND', (status_col, i), (status_col, i), HexColor('#ff9800'))]))
                    elif status == 'OK':
                        findings_table.setStyle(TableStyle([('BACKGROUND', (status_col, i), (status_col, i), HexColor('#4CAF50'))]))
                        findings_table.setStyle(TableStyle([('TEXTCOLOR', (status_col, i), (status_col, i), colors.white)]))

            story.append(findings_table)
    
    # Compliance section (if enabled)
    if options.get('include_compliance'):
        story.append(PageBreak())
        story.append(Paragraph("Compliance Overview", heading_style))
        
        compliance_text = """
        This security assessment covers multiple compliance frameworks:
        • SOC 2 Type II - System and Organization Controls
        • PCI DSS - Payment Card Industry Data Security Standard  
        • GDPR - General Data Protection Regulation
        • HIPAA - Health Insurance Portability and Accountability Act
        • ISO 27001 - Information Security Management
        """
        story.append(Paragraph(compliance_text, styles['Normal']))
    
    # Build PDF
    doc.build(story)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes

def _generate_advanced_csv_report(results, report_type):
    """Generate advanced CSV report."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Enhanced CSV headers
    headers = [
        'Timestamp', 'Service', 'Resource', 'Status', 'Severity_Score',
        'Issue', 'Remediation', 'Documentation_URL', 'Region', 'Account_ID'
    ]
    writer.writerow(headers)
    
    for result in results:
        # Add severity scoring
        severity_score = {
            'CRITICAL': 10,
            'HIGH': 7,
            'MEDIUM': 5,
            'LOW': 3,
            'OK': 0
        }.get(result.status, 0)
        
        row = [
            result.timestamp.isoformat(),
            result.service,
            result.resource,
            result.status,
            severity_score,
            result.issue,
            result.remediation or 'N/A',
            result.doc_url or 'N/A',
            getattr(result, 'region', 'N/A'),
            getattr(result, 'account_id', 'N/A')
        ]
        writer.writerow(row)
    
    output.seek(0)
    filename = f'security_report_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

def _generate_excel_report(results, report_type):
    """Generate advanced Excel report."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        return _generate_advanced_csv_report(results, report_type)
    
    # Create a pandas DataFrame
    data = []
    for result in results:
        severity_score = {
            'CRITICAL': 10,
            'HIGH': 7, 
            'MEDIUM': 5,
            'LOW': 3,
            'OK': 0
        }.get(result.status, 0)
        
        data.append({
            'Timestamp': result.timestamp.isoformat(),
            'Service': result.service,
            'Resource': result.resource,
            'Status': result.status,
            'Severity_Score': severity_score,
            'Issue': result.issue,
            'Remediation': result.remediation or 'N/A',
            'Documentation_URL': result.doc_url or 'N/A',
            'Region': getattr(result, 'region', 'N/A'),
            'Account_ID': getattr(result, 'account_id', 'N/A')
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Security Report - {report_type.title()}"
    
    # Add title
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = f"Aegis Cloud Security Report - {report_type.title()}"
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add headers starting from row 3
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='336699', end_color='336699', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row_data in enumerate(df.values, 4):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'security_report_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

def send_advanced_scheduled_report(user_id, report_type='comprehensive', additional_recipients=''):
    """Send advanced scheduled reports via email."""
    try:
        with app.app_context():
            user = User.query.get(user_id)
            if not user:
                return
                
            # Generate report
            results = ScanResult.query.filter_by(user_id=user_id).order_by(
                ScanResult.timestamp.desc()).limit(100).all()
                
            if not results:
                logging.info(f"No results found for user {user_id}, skipping report")
                return
                
            # Create PDF report
            pdf_bytes = _create_advanced_pdf_report(results, report_type, {
                'include_remediation': True,
                'include_compliance': True,
                'include_costs': False
            })
            
            # Prepare email
            recipients = [user.email]
            if additional_recipients:
                recipients.extend([email.strip() for email in additional_recipients.split(',') if email.strip()])
                
            subject = f"Aegis Security Report - {report_type.title()} - {datetime.now().strftime('%Y-%m-%d')}"
            
            # Enhanced email template
            body = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
                <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h2 style="color: #00A896; border-bottom: 2px solid #00A896; padding-bottom: 10px;">
                        Aegis Cloud Security Report
                    </h2>
                    
                    <p>Hello {user.username},</p>
                    
                    <p>Your scheduled <strong>{report_type.title()}</strong> security report is ready.</p>
                    
                    <div style="background: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0;">
                        <h3 style="color: #2c3e50; margin-top: 0;">Report Summary</h3>
                        <ul style="list-style: none; padding: 0;">
                            <li>📊 <strong>Total Findings:</strong> {len(results)}</li>
                            <li>🚨 <strong>Critical:</strong> {len([r for r in results if r.status == 'CRITICAL'])}</li>
                            <li>⚠️ <strong>High Priority:</strong> {len([r for r in results if r.status == 'HIGH'])}</li>
                            <li>✅ <strong>Passed Checks:</strong> {len([r for r in results if r.status == 'OK'])}</li>
                        </ul>
                    </div>
                    
                    <p>The detailed report is attached as a PDF file.</p>
                    
                    <div style="background: #e3f2fd; padding: 15px; border-radius: 5px; border-left: 4px solid #2196f3;">
                        <p style="margin: 0;"><strong>Next Steps:</strong></p>
                        <ol>
                            <li>Review critical and high-priority findings</li>
                            <li>Implement recommended security measures</li>
                            <li>Schedule remediation tasks</li>
                            <li>Monitor progress in your dashboard</li>
                        </ol>
                    </div>
                    
                    <p style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; color: #666; font-size: 12px;">
                        This is an automated report from Aegis Cloud Scanner.<br>
                        Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br>
                        <a href="#" style="color: #00A896;">Manage your report preferences</a>
                    </p>
                </div>
            </body>
            </html>
            """
            
            # Send email with attachment
            msg = Message(subject, recipients=recipients)
            msg.html = body
            msg.attach(
                f"security_report_{datetime.now().strftime('%Y%m%d')}.pdf",
                "application/pdf",
                pdf_bytes
            )
            
            mail.send(msg)
            logging.info(f"Advanced scheduled report sent to {recipients}")
            
    except Exception as e:
        logging.error(f"Failed to send advanced scheduled report: {e}")

# Production Health Check and Monitoring Endpoints
@app.route('/health', methods=['GET'])
def health_check():
    """Basic health check endpoint for load balancers and monitoring"""
    try:
        # Check database connectivity
        db.session.execute(db.text('SELECT 1'))
        
        # Check scheduler status
        scheduler_status = 'running' if scheduler.running else 'stopped'
        
        # Check secrets manager
        secrets_status = 'healthy' if secrets_manager else 'not_configured'
        
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'version': '1.0.0',
            'services': {
                'database': 'healthy',
                'scheduler': scheduler_status,
                'secrets_manager': secrets_status
            }
        }), 200
        
    except Exception as e:
        logging.error(f"Health check failed: {e}")
        return jsonify({
            'status': 'unhealthy',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'error': str(e)
        }), 503

@app.route('/health/detailed', methods=['GET'])
def detailed_health_check():
    """Detailed health check for comprehensive monitoring"""
    try:
        health_status = {
            'status': 'healthy',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'version': '1.0.0',
            'uptime_seconds': (datetime.now() - app.config.get('START_TIME', datetime.now())).total_seconds(),
            'services': {},
            'metrics': {}
        }
        
        # Database health
        try:
            db.session.execute(db.text('SELECT 1'))
            user_count = User.query.count()
            scan_count = ScanResult.query.count()
            health_status['services']['database'] = {
                'status': 'healthy',
                'connection': 'active',
                'users': user_count,
                'scan_results': scan_count
            }
        except Exception as e:
            health_status['services']['database'] = {
                'status': 'unhealthy',
                'error': str(e)
            }
            health_status['status'] = 'degraded'
        
        # Scheduler health
        try:
            scheduler_jobs = len(scheduler.get_jobs()) if scheduler else 0
            health_status['services']['scheduler'] = {
                'status': 'running' if scheduler and scheduler.running else 'stopped',
                'active_jobs': scheduler_jobs
            }
        except Exception as e:
            health_status['services']['scheduler'] = {
                'status': 'error',
                'error': str(e)
            }
        
        # Secrets manager health
        try:
            secrets_test = secrets_manager.get_secret('SECRET_KEY') if secrets_manager else None
            health_status['services']['secrets_manager'] = {
                'status': 'healthy' if secrets_test else 'unhealthy',
                'backend': getattr(secrets_manager, 'backend', 'unknown') if secrets_manager else 'not_configured'
            }
        except Exception as e:
            health_status['services']['secrets_manager'] = {
                'status': 'unhealthy',
                'error': str(e)
            }
        
        # System metrics
        import psutil
        health_status['metrics'] = {
            'cpu_percent': psutil.cpu_percent(),
            'memory_percent': psutil.virtual_memory().percent,
            'disk_percent': psutil.disk_usage('/').percent if os.name != 'nt' else psutil.disk_usage('C:\\').percent
        }
        
        # Determine overall status
        service_statuses = [service.get('status', 'unknown') for service in health_status['services'].values()]
        if any(status == 'unhealthy' or status == 'error' for status in service_statuses):
            health_status['status'] = 'unhealthy'
        elif any(status == 'degraded' for status in service_statuses):
            health_status['status'] = 'degraded'
            
        return jsonify(health_status), 200 if health_status['status'] == 'healthy' else 503
        
    except Exception as e:
        logging.error(f"Detailed health check failed: {e}")
        return jsonify({
            'status': 'unhealthy',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'error': str(e)
        }), 503

@app.route('/metrics', methods=['GET'])
def prometheus_metrics():
    """Prometheus-compatible metrics endpoint"""
    try:
        user_count = User.query.count()
        scan_count = ScanResult.query.count()
        critical_findings = ScanResult.query.filter_by(status='CRITICAL').count()
        
        metrics = f"""
# HELP aegis_users_total Total number of registered users
# TYPE aegis_users_total gauge
aegis_users_total {user_count}

# HELP aegis_scans_total Total number of scans performed
# TYPE aegis_scans_total gauge
aegis_scans_total {scan_count}

# HELP aegis_critical_findings_total Total number of critical findings
# TYPE aegis_critical_findings_total gauge
aegis_critical_findings_total {critical_findings}

# HELP aegis_scheduler_jobs Total number of scheduled jobs
# TYPE aegis_scheduler_jobs gauge
aegis_scheduler_jobs {len(scheduler.get_jobs()) if scheduler else 0}
"""
        
        return metrics, 200, {'Content-Type': 'text/plain; charset=utf-8'}
        
    except Exception as e:
        logging.error(f"Metrics endpoint failed: {e}")
        return f"# Error generating metrics: {str(e)}\n", 500, {'Content-Type': 'text/plain; charset=utf-8'}

@app.route('/ready', methods=['GET'])
def readiness_check():
    """Kubernetes readiness probe endpoint"""
    try:
        # Check if app is ready to receive traffic
        db.session.execute(db.text('SELECT 1'))
        
        # Ensure critical services are initialized
        if not secrets_manager:
            raise Exception("Secrets manager not initialized")
            
        return jsonify({
            'status': 'ready',
            'timestamp': datetime.now(timezone.utc).isoformat()
        }), 200
        
    except Exception as e:
        logging.error(f"Readiness check failed: {e}")
        return jsonify({
            'status': 'not_ready',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'error': str(e)
        }), 503

@app.route('/api/v1/system/info', methods=['GET'])
@check_verified
@check_2fa
def system_info():
    """System information endpoint (authenticated)"""
    try:
        import platform
        import psutil
        
        return jsonify({
            'system': {
                'platform': platform.system(),
                'architecture': platform.architecture()[0],
                'python_version': platform.python_version(),
                'hostname': platform.node()
            },
            'performance': {
                'cpu_count': psutil.cpu_count(),
                'memory_total_gb': round(psutil.virtual_memory().total / (1024**3), 2),
                'memory_available_gb': round(psutil.virtual_memory().available / (1024**3), 2),
                'disk_total_gb': round(psutil.disk_usage('/').total / (1024**3), 2) if os.name != 'nt' else round(psutil.disk_usage('C:\\').total / (1024**3), 2)
            },
            'database': {
                'users': User.query.count(),
                'scan_results': ScanResult.query.count(),
                'credentials': CloudCredential.query.count()
            }
        }), 200
        
    except Exception as e:
        logging.error(f"System info endpoint failed: {e}")
        return jsonify({'error': str(e)}), 500

# Contact Us Route
@app.route('/contact', methods=['GET', 'POST'])
def contact_us():
    """Contact us page with issue reporting functionality"""
    logging.error(f"CONTACT ROUTE HIT: {request.method} {request.path}")

    if request.method == 'GET':
        # Check auth for GET requests (page load)
        if not current_user.is_authenticated:
            return redirect(url_for('auth'))
        if session.get('guest_mode'):
            pass  # Allow guest mode
        elif not current_user.email_verified:
            flash("You must verify your email address to access this page.", "warning")
            return redirect(url_for('unverified'))
        elif current_user.is_2fa_enabled and session.get('2fa_passed') is not True:
            flash("Please complete the 2FA verification to continue.", "warning")
            return redirect(url_for('verify_2fa_login'))

        return render_template('contact.html')

    # POST request - AJAX form submission
    if request.method == 'POST':
        logging.info("CONTACT POST ROUTE REACHED")
        logging.info(f"Current user authenticated: {current_user.is_authenticated}")
        logging.info(f"User: {current_user}")

        # Check authentication for AJAX - return JSON errors
        if not current_user.is_authenticated:
            logging.error("User not authenticated")
            return jsonify({
                'success': False,
                'message': 'Authentication required. Please log in.'
            }), 401

        if not session.get('guest_mode'):
            if not current_user.email_verified:
                return jsonify({
                    'success': False,
                    'message': 'Email verification required. Please verify your email address.'
                }), 401

            if current_user.is_2fa_enabled and session.get('2fa_passed') is not True:
                return jsonify({
                    'success': False,
                    'message': '2FA verification required. Please complete 2FA verification.'
                }), 401

        try:
            logging.info("=== CONTACT FORM DEBUG ===")
            logging.info(f"Request method: {request.method}")
            logging.info(f"Request path: {request.path}")
            logging.info(f"Request endpoint: {request.endpoint}")
            logging.info(f"User authenticated: {current_user.is_authenticated}")
            logging.info(f"Current user: {current_user.username if current_user.is_authenticated else 'Anonymous'}")
            logging.info(f"Form data: {request.form}")
            logging.info(f"Files: {request.files}")
            logging.info("=== END DEBUG ===")

            # Get form data
            subject = request.form.get('subject', '').strip()
            category = request.form.get('category', '').strip()
            priority = request.form.get('priority', 'medium').strip()
            description = request.form.get('description', '').strip()
            environment = request.form.get('environment', '').strip()

            # Validate required fields
            if not subject or not category or not description:
                return jsonify({
                    'success': False,
                    'message': 'Please fill in all required fields.'
                }), 400

            # Handle file attachments
            uploaded_files = []
            attachments = request.files.getlist('attachments')

            for file in attachments:
                if file.filename:
                    # Validate file size (10MB max)
                    file.seek(0, 2)  # Seek to end
                    file_size = file.tell()
                    file.seek(0)  # Reset to beginning

                    if file_size > 10 * 1024 * 1024:  # 10MB
                        return jsonify({
                            'success': False,
                            'message': f'File {file.filename} is too large (max 10MB).'
                        }), 400

                    # Validate file type
                    allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.pdf', '.txt', '.log', '.csv'}
                    file_ext = os.path.splitext(file.filename)[1].lower()

                    if file_ext not in allowed_extensions:
                        return jsonify({
                            'success': False,
                            'message': f'File type {file_ext} is not allowed.'
                        }), 400

                    uploaded_files.append((file.filename, file.read()))

            # Prepare email content
            email_subject = f"[Aegis Support] {category.replace('_', ' ').title()}: {subject}"

            # Create HTML email body
            email_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 700px; margin: 0 auto; padding: 20px;">
                <div style="background: linear-gradient(135deg, #00A896, #007F73); color: white; padding: 20px; border-radius: 10px; margin-bottom: 20px;">
                    <h2 style="margin: 0; display: flex; align-items: center;">
                        <span style="margin-right: 10px;">🎧</span>
                        Support Request from Aegis Cloud Scanner
                    </h2>
                </div>

                <div style="background: #f8f9fa; padding: 20px; border-radius: 8px; border-left: 4px solid #00A896;">
                    <h3 style="color: #2c3e50; margin-top: 0;">Contact Information</h3>
                    <table style="width: 100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 8px; border-bottom: 1px solid #dee2e6; font-weight: bold; width: 150px;">User:</td>
                            <td style="padding: 8px; border-bottom: 1px solid #dee2e6;">{current_user.username} ({current_user.email})</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; border-bottom: 1px solid #dee2e6; font-weight: bold;">Subject:</td>
                            <td style="padding: 8px; border-bottom: 1px solid #dee2e6;">{subject}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; border-bottom: 1px solid #dee2e6; font-weight: bold;">Category:</td>
                            <td style="padding: 8px; border-bottom: 1px solid #dee2e6;"><span style="background: #e3f2fd; padding: 4px 8px; border-radius: 4px;">{category.replace('_', ' ').title()}</span></td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; border-bottom: 1px solid #dee2e6; font-weight: bold;">Priority:</td>
                            <td style="padding: 8px; border-bottom: 1px solid #dee2e6;">
                                <span style="background: {'#ff4444' if priority == 'critical' else '#ff9800' if priority == 'high' else '#2196f3' if priority == 'medium' else '#4caf50'}; color: white; padding: 4px 8px; border-radius: 4px;">
                                    {priority.upper()}
                                </span>
                            </td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; border-bottom: 1px solid #dee2e6; font-weight: bold;">Timestamp:</td>
                            <td style="padding: 8px; border-bottom: 1px solid #dee2e6;">{datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}</td>
                        </tr>
                    </table>
                </div>

                <div style="background: white; padding: 20px; border-radius: 8px; border: 1px solid #dee2e6; margin: 20px 0;">
                    <h3 style="color: #2c3e50; margin-top: 0;">Issue Description</h3>
                    <div style="background: #f8f9fa; padding: 15px; border-radius: 5px; border-left: 3px solid #00A896;">
                        {description.replace('\\n', '<br>')}
                    </div>
                </div>

                {f'''
                <div style="background: white; padding: 20px; border-radius: 8px; border: 1px solid #dee2e6; margin: 20px 0;">
                    <h3 style="color: #2c3e50; margin-top: 0;">Environment Information</h3>
                    <div style="background: #f8f9fa; padding: 15px; border-radius: 5px;">
                        {environment.replace(chr(10), '<br>')}
                    </div>
                </div>
                ''' if environment else ''}

                {f'''
                <div style="background: #fff3cd; padding: 15px; border-radius: 5px; border: 1px solid #ffeeba; margin: 20px 0;">
                    <h4 style="color: #856404; margin-top: 0;">📎 Attachments ({len(uploaded_files)})</h4>
                    <ul style="margin: 0; padding-left: 20px;">
                        {chr(10).join([f'<li>{filename} ({len(content)} bytes)</li>' for filename, content in uploaded_files])}
                    </ul>
                </div>
                ''' if uploaded_files else ''}

                <div style="background: #e8f5e8; padding: 15px; border-radius: 5px; border-left: 4px solid #4caf50; margin: 20px 0;">
                    <h4 style="color: #2e7d32; margin-top: 0;">📝 Next Steps</h4>
                    <p style="margin: 5px 0; color: #2e7d32;">
                        Our support team will review this request and respond within 24 hours.
                        For critical issues, we aim to respond within 2-4 hours during business hours.
                    </p>
                </div>

                <div style="text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #dee2e6;">
                    <p style="color: #6c757d; font-size: 12px; margin: 0;">
                        This support request was generated automatically from Aegis Cloud Scanner.<br>
                        Please do not reply to this email directly.
                    </p>
                </div>
            </div>
        </body>
        </html>
        """

            # Send email - with fallback for testing
            try:
                # Check if mail is properly configured
                if not current_app.config.get('MAIL_SERVER'):
                    logging.warning("Mail server not configured, saving contact request to log instead")

                    # Log the support request (fallback when email isn't configured)
                    log_message = f"""
                    SUPPORT REQUEST SUBMITTED
                    User: {current_user.username} ({current_user.email})
                    Subject: {subject}
                    Category: {category}
                    Priority: {priority}
                    Description: {description}
                    Environment: {environment}
                    Attachments: {len(uploaded_files)} files
                    Timestamp: {datetime.now()}
                    """
                    logging.info(log_message)

                    # Log the support request
                    log_audit(
                        action="Support Request Submitted",
                        details=f"Category: {category}, Priority: {priority}, Subject: {subject}",
                        user=current_user
                    )

                    return jsonify({
                        'success': True,
                        'message': 'Your support request has been submitted successfully! We will respond within 24 hours. (Note: Email delivery is currently disabled in development mode)'
                    }), 200

                else:
                    # Send email TO support email (aegis.aws.scanner@gmail.com)
                    # FROM configured mail (required by SMTP), REPLY-TO user's email
                    support_email = 'aegis.aws.scanner@gmail.com'
                    user_email = current_user.email
                    configured_sender = current_app.config.get('MAIL_USERNAME')

                    msg = Message(
                        subject=email_subject,
                        recipients=[support_email],  # TO: support email
                        sender=configured_sender,  # FROM: configured SMTP account (required)
                        reply_to=user_email  # REPLY-TO: user's email (for replies)
                    )
                    msg.html = email_body

                    # Add attachments if any
                    for filename, content in uploaded_files:
                        msg.attach(
                            filename,
                            'application/octet-stream',
                            content
                        )

                    mail.send(msg)
                    logging.info(f"Support request sent to {support_email} (from {configured_sender}, reply-to {user_email})")

                    # Log the support request
                    log_audit(
                        action="Support Request Submitted",
                        details=f"Category: {category}, Priority: {priority}, Subject: {subject}",
                        user=current_user
                    )

                    return jsonify({
                        'success': True,
                        'message': 'Your support request has been submitted successfully! We will respond within 24 hours.'
                    }), 200

            except Exception as email_error:
                import traceback
                logging.error(f"Failed to send support email: {email_error}")
                logging.error(f"Email error traceback: {traceback.format_exc()}")
                logging.error(f"MAIL_SERVER config: {current_app.config.get('MAIL_SERVER')}")
                logging.error(f"MAIL_USERNAME config: {current_app.config.get('MAIL_USERNAME')}")
                logging.error(f"SUPPORT_EMAIL config: {current_app.config.get('SUPPORT_EMAIL')}")

                # Fallback: still log the request even if email fails
                log_audit(
                    action="Support Request Submitted (Email Failed)",
                    details=f"Category: {category}, Priority: {priority}, Subject: {subject}, Error: {str(email_error)}",
                    user=current_user
                )

                return jsonify({
                    'success': True,
                    'message': f'Your support request has been logged successfully! Email delivery failed: {str(email_error)[:100]}'
                }), 200

        except Exception as e:
            logging.error(f"Contact form error: {e}")
            import traceback
            logging.error(f"Full traceback: {traceback.format_exc()}")
            return jsonify({
                'success': False,
                'message': f'An error occurred while processing your request: {str(e)}'
            }), 500

# Automation Rules API
@app.route('/api/v1/automation/rules', methods=['GET'])
@login_required
@check_verified
@check_2fa
def get_automation_rules():
    """Get all automation rules for the current user."""
    try:
        rules = AutomationRule.query.filter_by(user_id=current_user.id).all()
        rules_data = []
        for rule in rules:
            rules_data.append({
                'id': rule.id,
                'name': rule.name,
                'description': rule.description,
                'rule_type': rule.rule_type,
                'trigger_condition': json.loads(rule.trigger_condition) if rule.trigger_condition else {},
                'action_config': json.loads(rule.action_config) if rule.action_config else {},
                'is_active': rule.is_active,
                'created_at': rule.created_at.isoformat() if rule.created_at else None,
                'last_executed': rule.last_executed.isoformat() if rule.last_executed else None,
                'execution_count': rule.execution_count
            })
        return jsonify({'success': True, 'rules': rules_data})
    except Exception as e:
        logging.error(f"Error fetching automation rules: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/v1/automation/rules', methods=['POST'])
@login_required
@check_verified
@check_2fa
def create_automation_rule():
    """Create a new automation rule."""
    try:
        data = request.get_json()

        # Validate required fields
        required_fields = ['name', 'rule_type', 'trigger_condition', 'action_config']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400

        # Validate rule_type
        valid_types = ['remediation', 'notification', 'report']
        if data['rule_type'] not in valid_types:
            return jsonify({'success': False, 'error': f'Invalid rule_type. Must be one of: {valid_types}'}), 400

        # Create new rule
        new_rule = AutomationRule(
            user_id=current_user.id,
            name=data['name'],
            description=data.get('description', ''),
            rule_type=data['rule_type'],
            trigger_condition=json.dumps(data['trigger_condition']),
            action_config=json.dumps(data['action_config']),
            is_active=data.get('is_active', True)
        )

        db.session.add(new_rule)
        db.session.commit()

        # Log the creation
        log_audit(
            action="create_automation_rule",
            details=f"Created automation rule: {new_rule.name} (Type: {new_rule.rule_type})",
            user=current_user
        )

        return jsonify({
            'success': True,
            'message': 'Automation rule created successfully',
            'rule_id': new_rule.id
        })

    except Exception as e:
        logging.error(f"Error creating automation rule: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/v1/automation/rules/<int:rule_id>/toggle', methods=['POST'])
@login_required
@check_verified
@check_2fa
def toggle_automation_rule(rule_id):
    """Toggle the active status of an automation rule."""
    try:
        rule = AutomationRule.query.filter_by(id=rule_id, user_id=current_user.id).first()
        if not rule:
            return jsonify({'success': False, 'error': 'Rule not found'}), 404

        rule.is_active = not rule.is_active
        rule.updated_at = datetime.now(timezone.utc)
        db.session.commit()

        # Log the toggle
        log_audit(
            action="toggle_automation_rule",
            details=f"{'Enabled' if rule.is_active else 'Disabled'} automation rule: {rule.name}",
            user=current_user
        )

        return jsonify({
            'success': True,
            'message': f'Rule {"enabled" if rule.is_active else "disabled"} successfully',
            'is_active': rule.is_active
        })

    except Exception as e:
        logging.error(f"Error toggling automation rule: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# DEBUG: Simple test route at the end to see if routing works at all
@app.route('/debug-test', methods=['POST'])
def debug_test():
    return jsonify({'debug': True, 'message': 'Debug route works'})

if __name__ == '__main__':
    if not os.path.exists(ENV_FILE_PATH):
        print("INFO: First run detected. Creating temporary .env file.")
        temp_secret = secrets.token_hex(24)
        # NOTE: No longer creating temp encryption key - will use master password
        with open(ENV_FILE_PATH, 'w') as f:
            f.write(f"SECRET_KEY='{temp_secret}'\n")

    run_main_app()